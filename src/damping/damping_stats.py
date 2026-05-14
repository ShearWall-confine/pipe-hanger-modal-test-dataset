"""Damping statistics and plotting utilities.

These functions are GUI-friendly: no plt.show(), drawing into provided Figure.
"""

from __future__ import annotations

import argparse
import logging
from pathlib import Path
from typing import Dict, List, Tuple

import h5py
import numpy as np
import openpyxl
from matplotlib import cm
from matplotlib.colors import Normalize
from matplotlib.figure import Figure
from matplotlib.lines import Line2D
from scipy import stats

from damping.amplitude_anomaly import filter_result_rows_for_amplitude
from damping.damping_core import _bandpass_butter, extract_damping_single
from damping.damping_h5_schema import read_cycle_damping, read_damping_results
from damping.signal_access import load_preferred_signal
from core.modal_db import DEFAULT_H5_PATH
from tools.paper_figure_style import SUPPORT_COLORS, SUPPORT_MARKERS, support_marker_handles

LOGGER = logging.getLogger(__name__)

MODE_ORDER = ["radial", "axial", "planar_rotation"]
SUPPORT_ORDER = ["CA", "TB", "CSBD", "CSB"]
MODE_LABEL = {
    "radial": "(a) Radial",
    "axial": "(b) Axial",
    "planar_rotation": "(c) In-plane rotation",
}
DIR_COLOR = {"A": "#005f73", "R": "#0a9396", "AA": "#ca6702", "AR": "#ee9b00", "RA": "#bb3e03", "RR": "#9b2226"}
SUPPORT_COLOR = SUPPORT_COLORS


def _apply_pub_axis_style(ax, *, xlabel: str | None = None, ylabel: str | None = None) -> None:
    ax.grid(False)
    if xlabel:
        ax.set_xlabel(xlabel, fontweight="bold")
    if ylabel:
        ax.set_ylabel(ylabel, fontweight="bold")
    ax.tick_params(labelsize=8, length=3.2, width=0.8)
    for spine in ax.spines.values():
        spine.set_linewidth(0.8)


def _load_rod_length_map(compare_path: Path) -> Dict[str, float]:
    """Load mapping: {component_direction -> rod_length_from_col_N(l)} from COMPARE_6.xlsx."""
    if not compare_path.exists():
        return {}
    wb = openpyxl.load_workbook(compare_path, data_only=True)
    mapping: Dict[str, float] = {}
    for sheet in SUPPORT_ORDER:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        current_component = ""
        current_direction = ""
        current_length: float | None = None
        for r in range(2, ws.max_row + 1):
            b = ws.cell(r, 2).value
            c = ws.cell(r, 3).value
            m = ws.cell(r, 14).value  # column N = l(m)
            if isinstance(b, str) and b.strip():
                current_component = b.strip()
            if isinstance(c, str) and c.strip():
                current_direction = c.strip().upper()
                current_length = None
            if isinstance(m, (int, float)):
                current_length = float(m)
            if current_component and current_direction and current_length is not None:
                mapping[f"{current_component}_{current_direction}"] = float(current_length)
    return mapping


def _attach_rod_length(rows: List[dict], compare_path: Path) -> List[dict]:
    mapping = _load_rod_length_map(compare_path)
    for r in rows:
        key = f"{r.get('component_id', '')}_{r.get('direction', '')}"
        r["rod_length"] = mapping.get(key, np.nan)
    return rows


def _remove_group_outliers_iqr(
    rows: List[dict],
    value_key: str,
    group_keys: tuple[str, ...],
    k: float = 1.5,
) -> List[dict]:
    """Remove outliers by group using IQR rule."""
    grouped: Dict[tuple, List[dict]] = {}
    for r in rows:
        g = tuple(r.get(gk, None) for gk in group_keys)
        grouped.setdefault(g, []).append(r)

    kept: List[dict] = []
    for _g, rs in grouped.items():
        vals = np.asarray([float(x.get(value_key, np.nan)) for x in rs], dtype=np.float64)
        finite_mask = np.isfinite(vals)
        if np.sum(finite_mask) < 4:
            kept.extend(rs)
            continue
        v = vals[finite_mask]
        q1 = float(np.percentile(v, 25))
        q3 = float(np.percentile(v, 75))
        iqr = q3 - q1
        if iqr <= 0:
            kept.extend(rs)
            continue
        lo = q1 - k * iqr
        hi = q3 + k * iqr
        for r in rs:
            x = float(r.get(value_key, np.nan))
            if np.isfinite(x) and lo <= x <= hi:
                kept.append(r)
    return kept


def _collect_valid_rows(h5_path: Path) -> List[dict]:
    all_rows = read_damping_results(h5_path, exp_id=None)
    rows: List[dict] = []
    for exp_id, rs in all_rows.items():
        for r in rs:
            if r.get("valid", False):
                r2 = dict(r)
                r2["exp_id"] = exp_id
                rows.append(r2)
    return rows


def compute_damping_statistics(h5_path: Path) -> Dict[str, dict]:
    """Compute grouped statistics with bootstrap CI and distribution checks."""
    rows = _collect_valid_rows(h5_path)
    groups: Dict[Tuple[str, str], List[float]] = {}
    for r in rows:
        key = (r["support_type"], r["mode"])
        groups.setdefault(key, []).append(float(r["zeta_hilbert"]))

    out: Dict[str, dict] = {}
    rng = np.random.default_rng(1234)
    for (support, mode), values in groups.items():
        arr = np.asarray(values, dtype=np.float64)
        n = int(arr.size)
        if n == 0:
            continue
        mean = float(np.mean(arr))
        std = float(np.std(arr, ddof=1)) if n > 1 else 0.0
        cov_abs = float(std / abs(mean)) if abs(mean) > 1e-12 else np.nan
        median = float(np.median(arr))

        # Bootstrap CI for mean
        boot_means = np.zeros(10000, dtype=np.float64)
        for i in range(10000):
            sample = rng.choice(arr, size=n, replace=True)
            boot_means[i] = np.mean(sample)
        ci_low = float(np.percentile(boot_means, 2.5))
        ci_high = float(np.percentile(boot_means, 97.5))

        distribution_fit = "normal"
        fit_params: dict = {}
        if n >= 3:
            shapiro_stat, shapiro_p = stats.shapiro(arr)
            fit_params["shapiro_p"] = float(shapiro_p)
            fit_params["shapiro_stat"] = float(shapiro_stat)
            if shapiro_p <= 0.05 and np.all(arr > 0):
                shape, loc, scale = stats.lognorm.fit(arr, floc=0)
                ks_stat, ks_p = stats.kstest(arr, "lognorm", args=(shape, loc, scale))
                distribution_fit = "lognormal"
                fit_params.update(
                    {
                        "shape": float(shape),
                        "loc": float(loc),
                        "scale": float(scale),
                        "ks_stat": float(ks_stat),
                        "ks_p": float(ks_p),
                    }
                )

        out[f"{support}_{mode}"] = {
            "n": n,
            "mean": mean,
            "std": std,
            "cov_abs": cov_abs,
            "ci_95_lower": ci_low,
            "ci_95_upper": ci_high,
            "median": median,
            "distribution_fit": distribution_fit,
            "fit_params": fit_params,
        }

    # Drift statistics (valid damping + valid drift only)
    drift_groups: Dict[Tuple[str, str], List[float]] = {}
    for r in rows:
        if not r.get("drift_valid", False):
            continue
        if float(r.get("drift_r_squared", 0.0)) < 0.05:
            continue
        key = (r["support_type"], r["mode"])
        drift_groups.setdefault(key, []).append(float(r["drift_ratio"]))

    for support in SUPPORT_ORDER:
        for mode in MODE_ORDER:
            k = f"{support}_{mode}"
            if k not in out:
                out[k] = {
                    "n": 0,
                    "mean": np.nan,
                    "std": np.nan,
                    "cov_abs": np.nan,
                    "ci_95_lower": np.nan,
                    "ci_95_upper": np.nan,
                    "median": np.nan,
                    "distribution_fit": "normal",
                    "fit_params": {},
                }
            darr = np.asarray(drift_groups.get((support, mode), []), dtype=np.float64)
            if darr.size == 0:
                out[k].update(
                    {
                        "drift_mean": np.nan,
                        "drift_std": np.nan,
                        "drift_cov_abs": np.nan,
                        "drift_ci_95_lower": np.nan,
                        "drift_ci_95_upper": np.nan,
                        "drift_interpretation": "linear",
                    }
                )
                continue
            dmean = float(np.mean(darr))
            dstd = float(np.std(darr, ddof=1)) if darr.size > 1 else 0.0
            dcov_abs = float(dstd / abs(dmean)) if abs(dmean) > 1e-12 else np.nan
            boot = np.zeros(10000, dtype=np.float64)
            for i in range(10000):
                s = rng.choice(darr, size=darr.size, replace=True)
                boot[i] = np.mean(s)
            dlow = float(np.percentile(boot, 2.5))
            dhigh = float(np.percentile(boot, 97.5))
            if abs(dmean) < 0.01:
                interp = "linear"
            elif dmean < -0.01:
                interp = "softening"
            else:
                interp = "hardening"
            out[k].update(
                {
                    "drift_mean": dmean,
                    "drift_std": dstd,
                    "drift_cov_abs": dcov_abs,
                    "drift_ci_95_lower": dlow,
                    "drift_ci_95_upper": dhigh,
                    "drift_interpretation": interp,
                }
            )
    return out


def plot_damping_summary_by_support(
    h5_path: Path,
    fig: Figure,
    mode_filter: str | None = None,
) -> None:
    """2x2 support panels with boxplot + jitter + mean/std/COV annotation."""
    rows = _collect_valid_rows(h5_path)
    rows = _attach_rod_length(rows, h5_path.parent / "COMPARE_6.xlsx")
    rows = _remove_group_outliers_iqr(rows, value_key="zeta_hilbert", group_keys=("support_type", "mode"))
    if mode_filter:
        rows = [r for r in rows if r["mode"] == mode_filter]
    rng = np.random.default_rng(1234)

    rod_vals = np.asarray([float(r.get("rod_length", np.nan)) for r in rows], dtype=np.float64)
    finite = np.isfinite(rod_vals)
    norm = Normalize(vmin=float(np.min(rod_vals[finite])), vmax=float(np.max(rod_vals[finite]))) if np.any(finite) else None
    cmap = cm.get_cmap("viridis")

    fig.clear()
    axs = fig.subplots(2, 2, squeeze=False)
    support_axes = dict(zip(SUPPORT_ORDER, axs.flatten()))
    for support in SUPPORT_ORDER:
        ax = support_axes[support]
        _apply_pub_axis_style(ax, ylabel="Damping Ratio, $\\zeta$")
        ax.set_title(f"{support}", fontsize=10, fontweight="bold")
        mode_data = []
        for m in MODE_ORDER:
            vals = [float(r["zeta_hilbert"]) for r in rows if r["support_type"] == support and r["mode"] == m]
            mode_data.append(vals)
        positions = np.arange(1, len(MODE_ORDER) + 1)
        ax.boxplot(mode_data, positions=positions, widths=0.5, patch_artist=False, manage_ticks=False)
        for i, m in enumerate(MODE_ORDER, start=1):
            subset = [r for r in rows if r["support_type"] == support and r["mode"] == m]
            for r in subset:
                x = i + rng.uniform(-0.12, 0.12)
                rv = float(r.get("rod_length", np.nan))
                c = cmap(norm(rv)) if norm is not None and np.isfinite(rv) else "#6c757d"
                ax.scatter(
                    x,
                    float(r["zeta_hilbert"]),
                    s=10,
                    alpha=0.7,
                    color=c,
                )
            vals = np.asarray([float(r["zeta_hilbert"]) for r in subset], dtype=np.float64)
            if vals.size > 0:
                mean = float(np.mean(vals))
                std = float(np.std(vals, ddof=1)) if vals.size > 1 else 0.0
                cov_abs = std / abs(mean) if abs(mean) > 1e-12 else np.nan
                y = float(np.max(vals)) * 1.05
                ax.text(i, y, f"{mean:.4f} ± {std:.4f}\nCOV={cov_abs:.2f}", ha="center", va="bottom", fontsize=7)
        ax.set_xticks(positions)
        ax.set_xticklabels([MODE_LABEL[m] for m in MODE_ORDER], rotation=12, ha="right", fontsize=8)

    if norm is not None:
        sm = cm.ScalarMappable(norm=norm, cmap=cmap)
        cax = fig.add_axes([0.935, 0.12, 0.018, 0.76])  # independent colorbar axis
        cbar = fig.colorbar(sm, cax=cax)
        cbar.set_label("Rod Length, $l$ (m)", fontsize=8, fontweight="bold")
        cbar.ax.tick_params(labelsize=7, length=2.5, width=0.7)
    # Explicit margins; reserve right side for standalone colorbar.
    fig.suptitle("Damping Ratio Distribution by Support Type and Mode", fontsize=11, fontweight="bold", y=0.98)
    fig.subplots_adjust(left=0.08, right=0.90, bottom=0.14, top=0.90, wspace=0.24, hspace=0.34)


def plot_xi_amplitude(
    h5_path: Path,
    fig: Figure,
    support_type: str | None = None,
    mode: str | None = None,
    exclude_amp_anomalies: bool = False,
    enhanced: bool = False,
    threshold_amp: float = 50.0,
) -> None:
    """Scatter of local damping vs amplitude, with per-support power-law fitting."""
    fig.clear()
    rows = read_damping_results(h5_path, exp_id=None)
    selected_modes = [mode] if mode else MODE_ORDER
    axes = fig.subplots(1, len(selected_modes), squeeze=False, sharey=True).flatten()
    all_amp_values: List[float] = []
    all_zeta_values: List[float] = []

    for exp_id, rs in rows.items():
        if exclude_amp_anomalies:
            rs = filter_result_rows_for_amplitude(rs, h5_path)
        exp_support = rs[0]["support_type"] if rs else ""
        if support_type and exp_support != support_type:
            continue
        for r in rs:
            if not r.get("valid", False) or r["mode"] not in selected_modes:
                continue
            cyc = read_cycle_damping(h5_path, exp_id, r["channel_id"], r["mode"])
            if cyc is None:
                continue
            amp, zeta_local, _ = cyc
            if amp.size == 0:
                continue
            finite_mask = np.isfinite(amp) & np.isfinite(zeta_local)
            if not np.any(finite_mask):
                continue
            all_amp_values.extend(np.asarray(amp[finite_mask], dtype=np.float64).tolist())
            all_zeta_values.extend(np.asarray(zeta_local[finite_mask], dtype=np.float64).tolist())

    x_limits = None
    if len(all_amp_values) > 1:
        amp_arr = np.asarray(all_amp_values, dtype=np.float64)
        x_lo = float(np.min(amp_arr))
        x_hi = float(np.max(amp_arr))
        pad = max(0.05 * (x_hi - x_lo), 1e-4)
        x_limits = (max(0.0, x_lo - pad), x_hi + pad)

    y_limits = None
    if len(all_zeta_values) > 1:
        zeta_arr = np.asarray(all_zeta_values, dtype=np.float64)
        y_hi = float(np.max(zeta_arr))
        pad = max(0.05 * y_hi, 1e-4)
        y_limits = (0.0, y_hi + pad)

    for idx, (ax, m) in enumerate(zip(axes, selected_modes)):
        points_by_support: Dict[str, List[tuple[float, float]]] = {k: [] for k in SUPPORT_COLOR.keys()}
        for exp_id, rs in rows.items():
            if exclude_amp_anomalies:
                rs = filter_result_rows_for_amplitude(rs, h5_path)
            exp_support = rs[0]["support_type"] if rs else ""
            if support_type and exp_support != support_type:
                continue
            for r in rs:
                if not r.get("valid", False) or r["mode"] != m:
                    continue
                cyc = read_cycle_damping(h5_path, exp_id, r["channel_id"], m)
                if cyc is None:
                    continue
                amp, zeta_local, _ = cyc
                if amp.size == 0:
                    continue
                sup = r["support_type"]
                for a, z in zip(amp.tolist(), zeta_local.tolist()):
                    points_by_support.setdefault(sup, []).append((float(a), float(z)))

        # Interleaved reorder, then one-shot batched scatter
        support_keys = [k for k in SUPPORT_ORDER if points_by_support.get(k)]
        idx_map = {k: 0 for k in support_keys}
        interleaved_support: List[str] = []
        interleaved_amp: List[float] = []
        interleaved_zeta: List[float] = []
        remaining = True
        while remaining:
            remaining = False
            for sup in support_keys:
                idx = idx_map[sup]
                pts = points_by_support.get(sup, [])
                if idx < len(pts):
                    a, z = pts[idx]
                    interleaved_support.append(sup)
                    interleaved_amp.append(a)
                    interleaved_zeta.append(z)
                    idx_map[sup] = idx + 1
                    remaining = True
        draw_order = [sup for sup in ("TB", "CA", "CSBD", "CSB") if sup in support_keys]
        for sup in draw_order:
            indices = [i for i, s in enumerate(interleaved_support) if s == sup]
            if not indices:
                continue
            arr_x = np.asarray([interleaved_amp[i] for i in indices], dtype=np.float64)
            arr_y = np.asarray([interleaved_zeta[i] for i in indices], dtype=np.float64)
            rare_support = len(indices) < 250
            common_ca = sup == "CA"
            ax.scatter(
                arr_x,
                arr_y,
                s=18 if rare_support else 10,
                alpha=0.62 if rare_support else (0.50 if common_ca else 0.36),
                color=SUPPORT_COLOR.get(sup, "#6c757d"),
                marker=SUPPORT_MARKERS.get(sup, "o"),
                edgecolors="white" if rare_support else "none",
                linewidths=0.25 if rare_support else 0.0,
                rasterized=True,
                zorder=4 if rare_support else (3 if common_ca else 2),
            )

        if enhanced:
            ax.axvspan(0.0, threshold_amp, color="gray", alpha=0.05, zorder=0)
            ax.axvline(
                x=threshold_amp,
                color="gray",
                linestyle=":",
                linewidth=1.0,
                alpha=0.65,
                zorder=4,
            )

            for sup in ["CA", "TB", "CSB"]:
                pts = points_by_support.get(sup, [])
                if len(pts) < 5:
                    continue
                arr = np.asarray(pts, dtype=np.float64)
                amp_arr = arr[:, 0]
                zeta_arr = arr[:, 1]
                mask = (
                    np.isfinite(amp_arr)
                    & np.isfinite(zeta_arr)
                    & (amp_arr > max(30.0, threshold_amp * 0.6))
                    & (zeta_arr > 0.0)
                )
                if int(np.sum(mask)) < 5:
                    continue
                c_val = float(np.median(amp_arr[mask] * zeta_arr[mask]))
                if not np.isfinite(c_val) or c_val <= 0.0:
                    continue
                x0 = max(30.0, float(np.nanmin(amp_arr[mask])))
                x1 = float(np.nanmax(amp_arr[mask]))
                if not (np.isfinite(x0) and np.isfinite(x1) and x1 > x0):
                    continue
                a_theory = np.linspace(x0, x1, 200)
                z_theory = c_val / a_theory
                if y_limits is not None:
                    z_theory = np.clip(z_theory, y_limits[0], y_limits[1])
                ax.plot(
                    a_theory,
                    z_theory,
                    color=SUPPORT_COLOR.get(sup, "#6c757d"),
                    linestyle="--",
                    linewidth=1.2,
                    alpha=0.75,
                    zorder=5,
                )

            if y_limits is not None:
                ax.text(
                    threshold_amp + 2.0,
                    y_limits[1] * 0.95,
                    "NLI 5%\nthreshold",
                    fontsize=6.5,
                    color="gray",
                    va="top",
                    ha="left",
                    zorder=6,
                )
                ax.text(
                    0.97,
                    0.06,
                    "Dashed: Coulomb fit $\\zeta=C/A$",
                    transform=ax.transAxes,
                    fontsize=6.5,
                    color="#495057",
                    ha="right",
                    va="bottom",
                    bbox=dict(facecolor="white", edgecolor="none", alpha=0.65, boxstyle="round,pad=0.18"),
                    zorder=6,
                )

        xlabel = ""
        ylabel = "Local Damping Ratio, $\\zeta$" if idx == 0 else ""
        _apply_pub_axis_style(ax, xlabel=xlabel, ylabel=ylabel)
        ax.set_title(f"{MODE_LABEL.get(m, m)}", fontsize=11, fontweight="bold", pad=8)
        if idx > 0:
            ax.tick_params(axis="y", labelleft=False)
        if x_limits is not None:
            ax.set_xlim(*x_limits)
        if y_limits is not None:
            ax.set_ylim(*y_limits)
        ax.set_box_aspect(0.72)

    if axes.size > 0:
        axes[0].set_ylabel("Local Damping Ratio, $\\zeta$", fontweight="bold")
        axes[0].tick_params(axis="y", labelleft=True)

    support_handles = []
    support_labels = []
    for sup in SUPPORT_ORDER:
        if support_type and sup != support_type:
            continue
        support_handles.append(support_marker_handles((sup,), markersize=6.8)[0])
        support_labels.append(sup)

    if support_handles and axes.size > 0:
        axes[-1].legend(
            support_handles,
            support_labels,
            loc="upper right",
            ncol=1,
            fontsize=7.5,
            frameon=False,
            labelspacing=0.25,
            handletextpad=0.35,
            borderaxespad=0.45,
        )
    if axes.size > 1:
        axes[1].set_xlabel("Amplitude (mm/s)", fontsize=11, fontweight="normal", labelpad=8)
    fig.suptitle("Local Damping Ratio vs Amplitude (mm/s)", fontsize=12, fontweight="bold", y=0.965)
    fig.tight_layout(rect=(0.03, 0.06, 0.98, 0.94), w_pad=1.1)


def plot_drift_ratio_summary(
    h5_path: Path,
    fig: Figure,
    mode_filter: str | None = None,
    exclude_amp_anomalies: bool = False,
) -> None:
    """Grouped box plot of normalized frequency-amplitude slope."""
    fig.clear()
    rows = _collect_valid_rows(h5_path)
    if exclude_amp_anomalies:
        rows = filter_result_rows_for_amplitude(rows, h5_path)
    rows = _attach_rod_length(rows, h5_path.parent / "COMPARE_6.xlsx")
    rows = [r for r in rows if r.get("drift_valid", False) and float(r.get("drift_r_squared", 0.0)) >= 0.05]
    if mode_filter:
        rows = [r for r in rows if r["mode"] == mode_filter]
    rng = np.random.default_rng(1234)

    rod_vals = np.asarray([float(r.get("rod_length", np.nan)) for r in rows], dtype=np.float64)
    finite = np.isfinite(rod_vals)
    norm = Normalize(vmin=float(np.min(rod_vals[finite])), vmax=float(np.max(rod_vals[finite]))) if np.any(finite) else None
    cmap = cm.get_cmap("viridis")

    axs = fig.subplots(2, 2, squeeze=False)
    support_axes = dict(zip(SUPPORT_ORDER, axs.flatten()))
    for support in SUPPORT_ORDER:
        ax = support_axes[support]
        _apply_pub_axis_style(ax, ylabel="Norm. Freq-Amplitude Slope, $(1/f_n)\\,df/dA$")
        ax.set_title(f"{support}", fontsize=10, fontweight="bold")
        mode_data = []
        for m in MODE_ORDER:
            vals = [float(r["drift_ratio"]) for r in rows if r["support_type"] == support and r["mode"] == m]
            mode_data.append(vals)
        positions = np.arange(1, len(MODE_ORDER) + 1)
        ax.boxplot(mode_data, positions=positions, widths=0.5, patch_artist=False, manage_ticks=False)
        for i, m in enumerate(MODE_ORDER, start=1):
            subset = [r for r in rows if r["support_type"] == support and r["mode"] == m]
            for r in subset:
                x = i + rng.uniform(-0.12, 0.12)
                rv = float(r.get("rod_length", np.nan))
                c = cmap(norm(rv)) if norm is not None and np.isfinite(rv) else "#6c757d"
                ax.scatter(
                    x,
                    float(r["drift_ratio"]),
                    s=10,
                    alpha=0.7,
                    color=c,
                )
            vals = np.asarray([float(r["drift_ratio"]) for r in subset], dtype=np.float64)
            if vals.size > 0:
                mean = float(np.mean(vals))
                std = float(np.std(vals, ddof=1)) if vals.size > 1 else 0.0
                cov_abs = float(std / abs(mean)) if abs(mean) > 1e-12 else np.nan
                y = float(np.max(vals)) * 1.05 if np.max(vals) >= 0 else float(np.max(vals)) * 0.95
                ax.text(i, y, f"{mean:.4f} ± {std:.4f}\nCOV={cov_abs:.2f}", ha="center", va="bottom", fontsize=7)
        ax.set_xticks(positions)
        ax.set_xticklabels([MODE_LABEL[m] for m in MODE_ORDER], rotation=12, ha="right", fontsize=8)

    if norm is not None:
        sm = cm.ScalarMappable(norm=norm, cmap=cmap)
        cax = fig.add_axes([0.935, 0.12, 0.018, 0.76])  # independent colorbar axis
        cbar = fig.colorbar(sm, cax=cax)
        cbar.set_label("Rod Length, $l$ (m)", fontsize=8, fontweight="bold")
        cbar.ax.tick_params(labelsize=7, length=2.5, width=0.7)
    # Explicit margins; reserve right side for standalone colorbar.
    fig.suptitle("Normalized Frequency-Amplitude Slope Distribution by Support Type and Mode", fontsize=11, fontweight="bold", y=0.98)
    fig.subplots_adjust(left=0.08, right=0.90, bottom=0.14, top=0.90, wspace=0.24, hspace=0.34)


def plot_single_experiment_diagnostic(
    h5_path: Path,
    exp_id: str,
    fig: Figure,
    channel_id: str,
    mode: str,
) -> None:
    """4-panel diagnostic plot for one experiment/channel/mode."""
    with h5py.File(h5_path, "r") as h5:
        exp_group = h5["experiments"][exp_id]
        time_arr = np.asarray(exp_group["time"][:], dtype=np.float64)
        signal = load_preferred_signal(exp_group["channels"][channel_id])
        fs = float(exp_group.attrs.get("sample_rate", 0.0))
        if fs <= 0 and time_arr.size >= 2:
            dt = float(np.median(np.diff(time_arr)))
            fs = 1.0 / dt if dt > 0 else 0.0
        # pick first frequency for selected mode
        freq_target = None
        if "modal_freq_modes" in exp_group and "modal_freq_values" in exp_group:
            modes = exp_group["modal_freq_modes"][:]
            vals = exp_group["modal_freq_values"][:]
            for m, v in zip(modes, vals):
                ms = m.decode("utf-8", errors="ignore") if isinstance(m, bytes) else str(m)
                if ms == mode:
                    freq_target = float(v)
                    break
        if freq_target is None and "freq_peaks" in exp_group and exp_group["freq_peaks"].size > 0:
            freq_target = float(exp_group["freq_peaks"][0])
        if freq_target is None:
            freq_target = 1.0

    res = extract_damping_single(
        time=time_arr,
        signal=signal,
        freq_target=freq_target,
        channel_id=channel_id,
        mode=mode,
        fs=fs,
    )
    filtered, lo, hi, _ = _bandpass_butter(signal, fs, freq_target, bandwidth_pct=15.0, order=4)

    fig.clear()
    ax11, ax12, ax21, ax22 = fig.subplots(2, 2).flatten()
    # 1) raw + filtered
    ax11.plot(time_arr, signal, lw=0.7, ls="--", color="#6c757d")
    ax11.plot(time_arr, filtered, lw=0.8, color="#005f73")
    ax11.set_title("Raw + Filtered/Decay Window", fontsize=9)
    ax11.grid(False)

    # 2) FFT before/after
    if time_arr.size >= 2:
        dt = float(np.median(np.diff(time_arr)))
        f = np.fft.rfftfreq(signal.size, d=dt)
        mag0 = np.abs(np.fft.rfft(signal))
        mag1 = np.abs(np.fft.rfft(filtered))
        ax12.plot(f, mag0, lw=0.7, ls="--", color="#6c757d")
        ax12.plot(f, mag1, lw=0.8, color="#bb3e03")
    ax12.axvspan(lo, hi, color="#94d2bd", alpha=0.2)
    ax12.axvline(freq_target, color="#d00000", lw=1.0)
    ax12.set_xlim(left=0)
    ax12.set_title("FFT Compare + Band", fontsize=9)
    ax12.grid(False)

    # 3) envelope + fit line
    et = res.envelope_time
    ea = res.envelope_amplitude
    if et.size > 0:
        ax21.plot(et, ea, lw=0.8, color="#005f73")
        fit_mask = (et >= res.t_start) & (et <= res.t_end) & (ea > 0)
        if np.sum(fit_mask) > 3 and np.isfinite(res.zeta_hilbert):
            x = et[fit_mask]
            y0 = float(np.max(ea[fit_mask]))
            y_fit = y0 * np.exp(-res.zeta_hilbert * 2.0 * np.pi * freq_target * (x - x[0]))
            ax21.plot(x, y_fit, lw=1.0, color="#bb3e03")
    ax21.set_title(f"Hilbert Envelope | zeta={res.zeta_hilbert:.4f}, R2={res.r_squared:.3f}", fontsize=9)
    ax21.grid(False)

    # 4) log decrement view
    if et.size > 0:
        ax22.plot(et, ea, lw=0.8, color="#0a9396")
    ax22.set_title(f"Log-Dec | zeta_log={res.zeta_log_dec:.4f}, peaks={res.peak_count}", fontsize=9)
    ax22.grid(False)

    for ax in [ax11, ax12, ax21, ax22]:
        ax.tick_params(labelsize=7, length=2)
    fig.tight_layout()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate damping summary figures.")
    parser.add_argument("--h5-path", type=Path, default=DEFAULT_H5_PATH)
    parser.add_argument("--out-dir", type=Path, default=Path("..") / "processed")
    return parser.parse_args()


def main() -> None:
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(name)s: %(message)s")
    args = parse_args()
    args.out_dir.mkdir(parents=True, exist_ok=True)

    def save_both(fig: Figure, out_base: Path) -> None:
        fig.savefig(out_base.with_suffix(".png"), dpi=320)
        
    fig1 = Figure(figsize=(12, 8))
    plot_damping_summary_by_support(args.h5_path, fig1)
    save_both(fig1, args.out_dir / "damping_summary")

    fig2 = Figure(figsize=(12, 4))
    plot_xi_amplitude(args.h5_path, fig2)
    save_both(fig2, args.out_dir / "damping_xi_amplitude")

    fig3 = Figure(figsize=(12, 8))
    plot_drift_ratio_summary(args.h5_path, fig3)
    save_both(fig3, args.out_dir / "damping_freq_drift")

    stats_dict = compute_damping_statistics(args.h5_path)
    LOGGER.info("Computed damping statistics groups: %d", len(stats_dict))


if __name__ == "__main__":
    main()
