from __future__ import annotations

import argparse
import json
import math
from collections import Counter
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[2]
COMPARE6_PATH = ROOT / "data" / "metadata" / "COMPARE_6.xlsx"
OUT_PATH = ROOT / "data" / "metadata" / "compare6_validation_report.json"

EXPECTED_SHEETS = ["CA", "TB", "CSBD", "CSB"]
EXPECTED_HEADERS = {
    2: "组件",
    3: "组件方向",
    4: "模态方向",
    5: "试验结果",
    10: "摆长",
    13: "m(kg)",
    14: "l(m)",
    15: "r",
}
EXPECTED_MODES = {"径向平动", "轴向平动", "平面转动"}
EXPECTED_DIRECTIONS = {"A", "R", "AA", "AR", "RA", "RR"}


def as_float(v):
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        try:
            v = float(s)
        except ValueError:
            return None
    try:
        fv = float(v)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(fv):
        return None
    return fv


def validate_sheet(ws) -> dict:
    issues: list[dict] = []
    transitions: list[dict] = []
    mode_counter: Counter[str] = Counter()

    current_component = ""
    current_direction = ""
    current_block_rows = 0

    for col, expected in EXPECTED_HEADERS.items():
        actual = ws.cell(1, col).value
        if actual != expected:
            issues.append(
                {
                    "type": "header_mismatch",
                    "row": 1,
                    "col": col,
                    "expected": expected,
                    "actual": actual,
                }
            )

    for r in range(2, ws.max_row + 1):
        component = ws.cell(r, 2).value
        direction = ws.cell(r, 3).value
        mode = ws.cell(r, 4).value
        freq = ws.cell(r, 5).value
        pend_len_mm = ws.cell(r, 10).value
        mass = ws.cell(r, 13).value
        length = ws.cell(r, 14).value
        radius = ws.cell(r, 15).value
        reg_value = ws.cell(r, 19).value

        if isinstance(component, str) and component.strip():
            current_component = component.strip()
            current_block_rows = 0
            transitions.append({"row": r, "component": current_component})
        if isinstance(direction, str) and direction.strip():
            current_direction = direction.strip().upper()
            current_block_rows = 0
            transitions.append({"row": r, "direction": current_direction, "component": current_component})

        current_block_rows += 1

        if mode is None and freq is None and mass is None and length is None and radius is None:
            continue

        if not current_component:
            issues.append({"type": "missing_component_context", "row": r})
        if not current_direction:
            issues.append({"type": "missing_direction_context", "row": r})

        if isinstance(mode, str):
            mode = mode.strip()
        if mode not in EXPECTED_MODES:
            issues.append({"type": "unexpected_mode", "row": r, "value": mode})
        else:
            mode_counter[mode] += 1

        if current_direction and current_direction not in EXPECTED_DIRECTIONS:
            issues.append({"type": "unexpected_direction", "row": r, "value": current_direction})

        # Each direction block is expected to have three mode rows.
        if current_block_rows > 3 and mode is not None:
            issues.append(
                {
                    "type": "direction_block_too_long",
                    "row": r,
                    "component": current_component,
                    "direction": current_direction,
                    "block_rows_seen": current_block_rows,
                }
            )

        ff = as_float(freq)
        if ff is None or ff <= 0:
            issues.append({"type": "invalid_frequency", "row": r, "value": freq})

        mm = as_float(mass)
        ll = as_float(length)
        rr = as_float(radius)

        if mode in {"径向平动", "轴向平动"}:
            if mm is None or mm <= 0:
                issues.append({"type": "invalid_mass", "row": r, "value": mass})
            if ll is None or ll <= 0:
                issues.append({"type": "invalid_length", "row": r, "value": length})
            if rr is None or rr <= 0:
                issues.append({"type": "invalid_radius", "row": r, "value": radius})
        elif mode == "平面转动":
            if mm is None or mm <= 0:
                issues.append({"type": "invalid_mass_planar", "row": r, "value": mass})

        # J column is pendulum length in mm and should roughly match l(m)*1000 where both exist.
        pend_len = as_float(pend_len_mm)
        if pend_len is not None and ll is not None:
            if abs(pend_len - ll * 1000.0) > 1.0:
                issues.append(
                    {
                        "type": "length_mismatch_j_vs_n",
                        "row": r,
                        "j_mm": pend_len,
                        "n_m": ll,
                    }
                )

        if isinstance(reg_value, str) and reg_value.strip().upper() == "#REF!":
            issues.append({"type": "formula_ref_error", "row": r, "col": 19})

    return {
        "sheet": ws.title,
        "rows": ws.max_row,
        "cols": ws.max_column,
        "mode_counter": dict(mode_counter),
        "transition_count": len(transitions),
        "issue_count": len(issues),
        "issue_counter": dict(Counter(x["type"] for x in issues)),
        "issues": issues,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Validate the manual frequency workbook.")
    parser.add_argument("--compare-xlsx", type=Path, default=COMPARE6_PATH)
    parser.add_argument("--out", type=Path, default=OUT_PATH)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    wb = openpyxl.load_workbook(args.compare_xlsx, data_only=True)
    report = {
        "path": str(args.compare_xlsx),
        "sheetnames": wb.sheetnames,
        "missing_expected_sheets": [s for s in EXPECTED_SHEETS if s not in wb.sheetnames],
        "sheets": [],
    }
    for name in EXPECTED_SHEETS:
        if name in wb.sheetnames:
            report["sheets"].append(validate_sheet(wb[name]))

    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    print("report=", args.out)
    print("sheetnames=", wb.sheetnames)
    for sheet in report["sheets"]:
        print(
            sheet["sheet"],
            "issues=",
            sheet["issue_count"],
            "issue_types=",
            sheet["issue_counter"],
        )


if __name__ == "__main__":
    main()
