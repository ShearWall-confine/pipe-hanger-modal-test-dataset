from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Dict, List, Tuple

import h5py
import numpy as np
import openpyxl

from core.modal_db import DEFAULT_FREQ_IMPORT_LOG, DEFAULT_H5_PATH

VALID_SHEETS = {"CA", "TB", "CSBD", "CSB"}
VALID_DIRS = {"AA", "AR", "RA", "RR", "A", "R"}
MODE_ALIAS = {
    "径向平动": "radial",
    "轴向平动": "axial",
    "平面转动": "planar_rotation",
}


def normalize_mode(raw_mode: str) -> str:
    m = (raw_mode or "").strip()
    for key, std in MODE_ALIAS.items():
        if key in m:
            return std
    return "unknown"


def build_freq_mapping(freq_xlsx_path: Path) -> Tuple[Dict[str, List[float]], Dict[str, Dict[str, List[float]]]]:
    wb = openpyxl.load_workbook(freq_xlsx_path, data_only=True)
    mapping: Dict[str, List[float]] = {}
    mode_mapping: Dict[str, Dict[str, List[float]]] = {}
    for sheet_name in wb.sheetnames:
        if sheet_name not in VALID_SHEETS:
            continue
        ws = wb[sheet_name]
        last_component = ""
        last_direction = ""
        for r in range(2, ws.max_row + 1):
            component = ws.cell(r, 2).value
            direction = ws.cell(r, 3).value
            mode = ws.cell(r, 4).value
            freq_val = ws.cell(r, 5).value
            if isinstance(component, str) and component.strip():
                last_component = component.strip()
            component = last_component
            if not component:
                continue
            if not component.startswith(sheet_name + "_"):
                continue
            if isinstance(direction, str) and direction.strip():
                last_direction = direction.strip().upper()
            direction = last_direction
            if not direction:
                continue
            if direction not in VALID_DIRS:
                continue
            if not isinstance(freq_val, (int, float)):
                continue
            key = f"{component}_{direction}"
            fv = float(freq_val)
            mapping.setdefault(key, []).append(fv)
            mode_key = normalize_mode(str(mode) if mode is not None else "")
            mode_mapping.setdefault(key, {}).setdefault(mode_key, []).append(fv)

    for key, values in mapping.items():
        # Keep deterministic order while deduplicating
        deduped = list(dict.fromkeys(round(v, 6) for v in values))
        mapping[key] = deduped
    for key, md in mode_mapping.items():
        for mk, vals in md.items():
            md[mk] = list(dict.fromkeys(round(v, 6) for v in vals))
    return mapping, mode_mapping


def write_freqs_to_h5(
    h5_path: Path,
    mapping: Dict[str, List[float]],
    mode_mapping: Dict[str, Dict[str, List[float]]],
    log_path: Path,
) -> Dict[str, List[str]]:
    missing: List[str] = []
    updated: List[str] = []
    with h5py.File(h5_path, "a") as h5:
        experiments = h5["experiments"]
        exp_ids = sorted(experiments.keys())
        for exp_id in exp_ids:
            exp_group = experiments[exp_id]
            peaks = mapping.get(exp_id, [])
            mode_dict = mode_mapping.get(exp_id, {})
            if "freq_peaks" in exp_group:
                del exp_group["freq_peaks"]
            exp_group.create_dataset(
                "freq_peaks",
                data=np.asarray(peaks, dtype=np.float32),
                maxshape=(None,),
                compression="gzip",
            )
            if "modal_freq_values" in exp_group:
                del exp_group["modal_freq_values"]
            if "modal_freq_modes" in exp_group:
                del exp_group["modal_freq_modes"]
            flat_modes: List[str] = []
            flat_vals: List[float] = []
            for mode_name, vals in mode_dict.items():
                for v in vals:
                    flat_modes.append(mode_name)
                    flat_vals.append(float(v))
            exp_group.create_dataset(
                "modal_freq_modes",
                data=np.asarray(flat_modes, dtype="S32"),
                maxshape=(None,),
                compression="gzip",
            )
            exp_group.create_dataset(
                "modal_freq_values",
                data=np.asarray(flat_vals, dtype=np.float32),
                maxshape=(None,),
                compression="gzip",
            )
            channels = exp_group["channels"]
            for ch_id in channels.keys():
                ch_group = channels[ch_id]
                if "freq_peaks" in ch_group:
                    del ch_group["freq_peaks"]
                ch_group.create_dataset(
                    "freq_peaks",
                    data=np.asarray(peaks, dtype=np.float32),
                    maxshape=(None,),
                    compression="gzip",
                )
            if peaks:
                updated.append(exp_id)
            else:
                missing.append(exp_id)

    report = {
        "updated_count": len(updated),
        "missing_count": len(missing),
        "updated_experiments": updated,
        "missing_experiments": missing,
    }
    log_path.parent.mkdir(parents=True, exist_ok=True)
    log_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    return report


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import manual frequency table into H5.")
    parser.add_argument("--h5-path", type=Path, default=DEFAULT_H5_PATH)
    parser.add_argument("--freq-xlsx", type=Path, required=True)
    parser.add_argument("--log-path", type=Path, default=DEFAULT_FREQ_IMPORT_LOG)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    mapping, mode_mapping = build_freq_mapping(args.freq_xlsx)
    report = write_freqs_to_h5(args.h5_path, mapping, mode_mapping, args.log_path)
    print(f"Mapped experiments from table: {len(mapping)}")
    print(f"Updated: {report['updated_count']}, Missing: {report['missing_count']}")
    print(f"Log: {args.log_path}")


if __name__ == "__main__":
    main()
