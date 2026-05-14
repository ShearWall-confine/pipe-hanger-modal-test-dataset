from __future__ import annotations

"""Notebook-extracted paper plotting workbench logic.

This module is generated from paper_plot_workbench.ipynb so the notebook can
stay thin and Codex-friendly while the implementation lives in .py files.
"""


# ===== extracted from notebook cell 01 =====


from pathlib import Path
import matplotlib.pyplot as plt
from tools.publication_plotting import (
    WORD_A4_TEXT_WIDTH_IN,
    WORD_THEME,
    save_figure_bundle,
    show_saved_figure,
    style_figure,
)
from tools.paper_figure_style import (
    SUPPORT_COLORS,
    SUPPORT_MARKERS,
    apply_insert_text_style,
    apply_insert_text_style_for_figure,
    font_profile_for_latex_insert,
)

WORKBENCH_DIR = Path.cwd().resolve()
CODE_ROOT = WORKBENCH_DIR
PROJECT_DIR = CODE_ROOT.parent
DATA_DIR = PROJECT_DIR / "data"
PROCESSED_DIR = DATA_DIR / "derived"
METADATA_DIR = DATA_DIR / "metadata"
PAPER_OUT_DIR = PROJECT_DIR / "figures"
WORKBENCH_OUT_DIR = PAPER_OUT_DIR
WORKBENCH_PNG_OUT_DIR = PAPER_OUT_DIR / "png"
WORKBENCH_TIFF_OUT_DIR = PROJECT_DIR / "_generated_tiff"
WORKBENCH_PDF_OUT_DIR = PAPER_OUT_DIR / "pdf"
DIM_OUT_DIR = PROCESSED_DIR
DAMP_OUT_DIR = WORKBENCH_OUT_DIR
WORKBENCH_OUT_DIR.mkdir(parents=True, exist_ok=True)
WORKBENCH_PNG_OUT_DIR.mkdir(parents=True, exist_ok=True)
WORKBENCH_TIFF_OUT_DIR.mkdir(parents=True, exist_ok=True)
WORKBENCH_PDF_OUT_DIR.mkdir(parents=True, exist_ok=True)

GLOBAL_H5_PATH = PROCESSED_DIR / "modal_db.h5"
GLOBAL_COMPARE_XLSX = METADATA_DIR / "COMPARE_6.xlsx"

RUN_DAMPING_BATCH = True
BANDWIDTH_PCT = 15.0
BUTTER_ORDER = 4
OVERWRITE = True
WORKERS = None

PUB = {
    "dpi": 600,
    "font.family": "serif",
    "font.serif": ["Times New Roman", "Times", "DejaVu Serif"],
    "mathtext.fontset": "stix",
    "mathtext.default": "it",
    "font.size": 11.0,
    "axes.titlesize": 12.0,
    "axes.labelsize": 11.5,
    "xtick.labelsize": 10.0,
    "ytick.labelsize": 10.0,
    "legend.fontsize": 9.6,
    "axes.linewidth": 1.05,
    "lines.linewidth": 1.35,
    "xtick.major.width": 1.05,
    "ytick.major.width": 1.05,
    "xtick.major.size": 4.8,
    "ytick.major.size": 4.8,
    "xtick.minor.width": 0.8,
    "ytick.minor.width": 0.8,
    "xtick.minor.size": 3.0,
    "ytick.minor.size": 3.0,
    "legend.frameon": False,
    "figure.dpi": 600,
    "savefig.dpi": 600,
}
rc_pub = {k: v for k, v in PUB.items() if k != "dpi"}
plt.rcParams.update(rc_pub)

print("[CONFIG] H5_PATH =", GLOBAL_H5_PATH)
print("[CONFIG] COMPARE_XLSX =", GLOBAL_COMPARE_XLSX)
print("[CONFIG] WORKBENCH_OUT_DIR =", WORKBENCH_OUT_DIR)
print("[CONFIG] DIM_OUT_DIR =", DIM_OUT_DIR)
print("[CONFIG] DAMP_OUT_DIR =", DAMP_OUT_DIR)
print("[CONFIG] RUN_DAMPING_BATCH =", RUN_DAMPING_BATCH)


def get_workbench_paths() -> dict[str, Path]:
    return {
        "workbench_dir": WORKBENCH_DIR,
        "code_root": CODE_ROOT,
        "project_dir": PROJECT_DIR,
        "processed_dir": PROCESSED_DIR,
        "paper_out_dir": PAPER_OUT_DIR,
        "workbench_out_dir": WORKBENCH_OUT_DIR,
        "workbench_png_out_dir": WORKBENCH_PNG_OUT_DIR,
        "workbench_tiff_out_dir": WORKBENCH_TIFF_OUT_DIR,
        "workbench_pdf_out_dir": WORKBENCH_PDF_OUT_DIR,
        "dim_out_dir": DIM_OUT_DIR,
        "damp_out_dir": DAMP_OUT_DIR,
        "h5_path": GLOBAL_H5_PATH,
        "compare_xlsx": GLOBAL_COMPARE_XLSX,
    }



# ===== extracted from notebook cell 03 =====


import csv
import math
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Tuple

import h5py
import numpy as np
import openpyxl
import matplotlib.pyplot as plt


# 参数区
H5_PATH = GLOBAL_H5_PATH
COMPARE_XLSX = GLOBAL_COMPARE_XLSX
OUT_DIR = DIM_OUT_DIR

OUTLIER = "none"  # "none" | "iqr" | "residual"
RESIDUAL_THRESH = 0.20
G_ACCEL = 9.81

MODE_ORDER = ["radial", "axial", "planar_rotation"]
MODE_ALIAS = {
    "径向平动": "radial",
    "轴向平动": "axial",
    "平面转动": "planar_rotation",
}
MODE_CN = {
    "radial": "Radial Translation",
    "axial": "Axial Translation",
    "planar_rotation": "In-plane rotation",
}
MODE_COLOR = {
    "radial": "#d00000",
    "axial": "#3a86ff",
    "planar_rotation": "#2a9d8f",
}
COMPONENT_ORDER_RAW = [
    "CA_07", "CA_13", "CA_07_Top", "CA_13_Top", "CA_07+13_Mid", "CA_13+07_Mid",
    "TB_07", "TB_13", "TB_07_Top", "TB_13_Top", "TB_07+07_Mid", "TB_07+13_Mid",
    "TB_13+07_Mid", "TB_07+07_MidTop", "TB_07+13_MidTop", "TB_13+07_MidTop",
    "TR_07", "TR_13", "CS_07", "CS_13",
]
DIRECTION_RANK = {"A": 0, "R": 1, "AA": 2, "AR": 3, "RA": 4, "RR": 5}

OUT_DIR.mkdir(parents=True, exist_ok=True)
plt.rcParams.update({
    "font.size": 10,
    "axes.titlesize": 12.8,
    "axes.labelsize": 11.2,
    "xtick.labelsize": 9.4,
    "ytick.labelsize": 9.4,
    "grid.alpha": 0.2,
})

print("[CONFIG] OUT_DIR =", OUT_DIR)
print("[CONFIG] OUTLIER =", OUTLIER)


# 全局字体、样式与绘图工具函数（统一规范）
FONT_FAMILY = "Times New Roman"  # fallback: Times / DejaVu Serif
FONT_SIZE_BASE = 12.6
FONT_SIZE_TITLE = 13.8
FONT_SIZE_LABEL = 12.8
FONT_SIZE_TICK = 10.8
FONT_SIZE_LEGEND = 10.8
FONT_SIZE_TEXTBOX = 9.8
FONT_SIZE_XTICK = 10.0

GROUP_COLOR = {
    "CA": "#d62828",
    "OTHERS": "#495057",
}
FIT_GROUP_LABEL = {
    "CA": "CA",
    "OTHERS": "SR",
}
SUPPORT_COLOR = {
    "CA": "#d62828",
    "TB": "#0077b6",
    "CSBD": "#6a994e",
    "CSB": "#8d6a9f",
}
SPLIT_EXP_COLOR = {
    "CA": "#b91c1c",
    "TB": "#005f8f",
    "CSBD": "#4f7f3a",
    "CSB": "#6f4d7c",
}
ERROR_COLOR = {
    "pendulum": "#8c92ac",
    "fitted": "#e07a5f",
}

plt.rcParams.update({
    "font.family": "serif",
    "font.serif": [FONT_FAMILY, "Times", "DejaVu Serif"],
    "mathtext.fontset": "stix",
    "mathtext.default": "it",
    "font.size": FONT_SIZE_BASE,
    "axes.titlesize": FONT_SIZE_TITLE,
    "axes.labelsize": FONT_SIZE_LABEL,
    "xtick.labelsize": FONT_SIZE_TICK,
    "ytick.labelsize": FONT_SIZE_TICK,
    "legend.fontsize": FONT_SIZE_LEGEND,
    "figure.titlesize": FONT_SIZE_TITLE,
    "lines.linewidth": 1.35,
    "grid.alpha": 0.0,
})


def apply_axis_style(ax, *, grid: bool = False) -> None:
    """Apply consistent spine/tick style to one axis."""
    for side in ["left", "right", "top", "bottom"]:
        ax.spines[side].set_linewidth(1.2)
    ax.tick_params(direction="in", length=4, width=1.2)
    ax.grid(False)


def compute_prefix_spans(exp_union: list[str]) -> list[tuple[int, int]]:
    """Return contiguous spans grouped by exp_id prefix (CA/TB/CSBD/CSB...)."""
    if not exp_union:
        return []
    spans: list[tuple[int, int]] = []
    start = 0
    prev = exp_union[0].split("_")[0] if "_" in exp_union[0] else exp_union[0]
    for i in range(1, len(exp_union) + 1):
        cur = None
        if i < len(exp_union):
            cur = exp_union[i].split("_")[0] if "_" in exp_union[i] else exp_union[i]
        if i == len(exp_union) or cur != prev:
            spans.append((start, i - 1))
            if i < len(exp_union):
                start = i
                prev = cur
    return spans


def compute_component_spans(exp_union: list[str]) -> list[tuple[int, int, str]]:
    """Return contiguous spans grouped by component, leaving direction as leaf labels."""
    if not exp_union:
        return []
    spans: list[tuple[int, int, str]] = []
    start = 0
    prev_component, _ = _split_exp_id(exp_union[0])
    for i in range(1, len(exp_union) + 1):
        cur_component = None
        if i < len(exp_union):
            cur_component, _ = _split_exp_id(exp_union[i])
        if i == len(exp_union) or cur_component != prev_component:
            spans.append((start, i - 1, prev_component))
            if i < len(exp_union):
                start = i
                prev_component = cur_component
    return spans


def component_tree_label(component: str) -> str:
    """Short component label for hierarchical experiment axes."""
    text = str(component)
    support = get_support_group(f"{text}_A")
    prefix = f"{support}_"
    if text.startswith(prefix):
        text = text[len(prefix):]
    text = text.replace("_", " ")
    return text


def component_axis_parts(component: str) -> tuple[str, str]:
    """Map a component id to compact Table-1-style spring and rod labels."""
    text = str(component)
    support = get_support_group(f"{text}_A")
    prefix = f"{support}_"
    if text.startswith(prefix):
        text = text[len(prefix):]
    parts = [p for p in text.split("_") if p]
    if not parts:
        return "", text

    spring_terms = {"Top", "Mid", "TopMid", "MidTop"}
    if parts[-1] in spring_terms:
        spring = parts[-1]
        rod = "+".join(parts[:-1])
    else:
        spring = "TR" if support == "CSBD" else "Single"
        rod = "+".join(parts)
    spring_short = {
        "Single": "S",
        "Top": "Top",
        "Mid": "Mid",
        "TopMid": "TM",
        "MidTop": "TM",
        "TR": "TR",
    }.get(spring, spring)
    return spring_short, rod


def orientation_color(direction: str) -> str:
    return {
        "A": "#b91c1c",
        "R": "#1d4ed8",
        "AA": "#b91c1c",
        "AR": "#b91c1c",
        "RA": "#1d4ed8",
        "RR": "#1d4ed8",
    }.get(str(direction), "#374151")


def support_span_label(exp_id: str) -> str:
    return get_support_group(exp_id)


def orientation_label(exp_id: str) -> str:
    _, direction = _split_exp_id(exp_id)
    return direction


def draw_experiment_tree_axis(ax, exp_union: list[str]):
    """Draw compact tree-style experiment labels on a dedicated label axis."""
    from matplotlib.patches import Rectangle

    component_spans = compute_component_spans(exp_union)
    support_spans = compute_prefix_spans(exp_union)
    x = np.arange(len(exp_union), dtype=float)
    ax.set_xlim(-0.6, len(exp_union) - 0.4)
    ax.set_ylim(0.0, 1.0)
    ax.set_xticks([])
    ax.set_yticks([])
    for spine in ax.spines.values():
        spine.set_visible(False)

    trans = ax.transData
    y_orientation_band = 0.88
    y_rod_line = 0.67
    y_rod_text = 0.57
    y_spring_line = 0.40
    y_spring_text = 0.30
    y_support_line = 0.13
    y_support_text = 0.02

    def draw_orientation_swatches(center_x: float, direction: str) -> None:
        seq = list(direction) if direction else ["?"]
        n = len(seq)
        gap = 0.04
        total_width = 0.58
        swatch_width = (total_width - gap * max(n - 1, 0)) / n
        left = center_x - 0.5 * total_width
        for idx, token in enumerate(seq):
            ax.add_patch(
                Rectangle(
                    (left + idx * (swatch_width + gap), y_orientation_band - 0.038),
                    swatch_width,
                    0.076,
                    transform=trans,
                    facecolor=orientation_color(token),
                    edgecolor="white",
                    linewidth=0.35,
                    alpha=0.96,
                    clip_on=False,
                )
            )

    for i, exp_id in enumerate(exp_union):
        direction = orientation_label(exp_id)
        draw_orientation_swatches(float(i), direction)

    legend_y = 1.03
    experiment_label = ax.text(
        0.00,
        legend_y,
        "Orientation:",
        transform=ax.transAxes,
        ha="left",
        va="center",
        fontsize=8.4,
        color="#111827",
        clip_on=False,
    )
    for x0, token, label in [(0.135, "A", "A"), (0.195, "R", "R")]:
        ax.add_patch(
            Rectangle(
                (x0, legend_y - 0.018),
                0.028,
                0.036,
                transform=ax.transAxes,
                facecolor=orientation_color(token),
                edgecolor="white",
                linewidth=0.45,
                clip_on=False,
            )
        )
        ax.text(
            x0 + 0.036,
            legend_y,
            label,
            transform=ax.transAxes,
            ha="left",
            va="center",
            fontsize=8.3,
            color="#111827",
            clip_on=False,
        )

    spring_spans: list[tuple[int, int, str]] = []
    for s, e, component in component_spans:
        spring_label, _ = component_axis_parts(component)
        if spring_spans and spring_spans[-1][2] == spring_label and spring_spans[-1][1] + 1 == s:
            prev_s, _prev_e, prev_label = spring_spans[-1]
            spring_spans[-1] = (prev_s, e, prev_label)
        else:
            spring_spans.append((s, e, spring_label))

    for s, e, component in component_spans:
        x0 = s - 0.45
        x1 = e + 0.45
        xc = 0.5 * (s + e)
        _, rod_label = component_axis_parts(component)
        ax.plot([x0, x1], [y_rod_line, y_rod_line], transform=trans, color="#64748b", lw=0.6, clip_on=False)
        ax.plot([x0, x0], [y_rod_line, y_rod_line - 0.06], transform=trans, color="#64748b", lw=0.6, clip_on=False)
        ax.plot([x1, x1], [y_rod_line, y_rod_line - 0.06], transform=trans, color="#64748b", lw=0.6, clip_on=False)
        ax.text(
            xc,
            y_rod_text,
            rod_label,
            transform=trans,
            ha="center",
            va="center",
            rotation=0,
            fontsize=8.8,
            bbox=dict(facecolor="white", edgecolor="none", boxstyle="square,pad=0.08", alpha=0.95),
            clip_on=False,
        )

    for s, e, spring_label in spring_spans:
        x0 = s - 0.45
        x1 = e + 0.45
        xc = 0.5 * (s + e)
        ax.plot([x0, x1], [y_spring_line, y_spring_line], transform=trans, color="#4b5563", lw=0.7, clip_on=False)
        ax.plot([x0, x0], [y_spring_line, y_spring_line - 0.06], transform=trans, color="#4b5563", lw=0.7, clip_on=False)
        ax.plot([x1, x1], [y_spring_line, y_spring_line - 0.06], transform=trans, color="#4b5563", lw=0.7, clip_on=False)
        ax.text(
            xc,
            y_spring_text,
            spring_label,
            transform=trans,
            ha="center",
            va="center",
            rotation=0,
            fontsize=9.0,
            fontweight="bold",
            bbox=dict(facecolor="white", edgecolor="none", boxstyle="square,pad=0.08", alpha=0.95),
            clip_on=False,
        )

    for s, e in support_spans:
        x0 = s - 0.5
        x1 = e + 0.5
        xc = 0.5 * (s + e)
        ax.plot([x0, x1], [y_support_line, y_support_line], transform=trans, color="#111827", lw=0.8, clip_on=False)
        ax.plot([x0, x0], [y_support_line, y_support_line + 0.08], transform=trans, color="#111827", lw=0.8, clip_on=False)
        ax.plot([x1, x1], [y_support_line, y_support_line + 0.08], transform=trans, color="#111827", lw=0.8, clip_on=False)
        ax.text(
            xc,
            y_support_text,
            support_span_label(exp_union[s]),
            transform=trans,
            ha="center",
            va="top",
            fontsize=10.0,
            fontweight="bold",
            bbox=dict(facecolor="white", edgecolor="none", boxstyle="square,pad=0.10", alpha=0.95),
            clip_on=False,
        )
    experiment_label = ax.text(
        0.5,
        -0.14,
        "Experiment",
        transform=ax.transAxes,
        ha="center",
        va="top",
        fontsize=12.5,
        fontweight="semibold",
        clip_on=False,
    )
    experiment_label.set_gid("keep-fontsize")
    return experiment_label


def build_mode_stat_text(mode_stats: dict) -> str:
    """Format two-group fit stats text for one mode."""
    ca_st = mode_stats.get("CA")
    oth_st = mode_stats.get("OTHERS")
    lines = []
    if ca_st is not None:
        lines.append(f"CA: A={ca_st.A:.4g}, B={ca_st.B:.4g}, R2={ca_st.r2:.3f}, RMSE={ca_st.rmse:.4f}")
    if oth_st is not None:
        lines.append(
            f"SR: A={oth_st.A:.4g}, B={oth_st.B:.4g}, R2={oth_st.r2:.3f}, RMSE={oth_st.rmse:.4f}"
        )
    return "\n".join(lines) if lines else "No fit stats"


print("[FONT] Applied:", FONT_FAMILY)


# 点大小统一配置（只改这里即可全局生效）
POINT_SIZE = {

    "split_exp": 26,                 # Split Panels: 实验频率点
    "split_pend": 20*0.80,           # Split Panels: 单摆频率点
    "split_fit": 30,                 # Split Panels: 拟合频率点
    "split_err_pend_ms": 4*0.88,     # Split Panels: 单摆误差顶点 ms
    "split_err_fit_ms": 5.2,         # Split Panels: 拟合误差顶点 ms
    "norm_pend": 20*1.15,            # 归一化图: 单摆点
    "norm_fit_ca": 30*1.10,          # 归一化图: 拟合点-CA
    "norm_fit_others": 28*1.10,      # 归一化图: 拟合点-TB/CSBD/CSB
}



@dataclass
class FitStats:
    mode: str
    fit_group: str
    n_total: int
    n_used: int
    outlier_policy: str
    A: float
    B: float
    r2: float
    rmse: float
    mae: float


def _to_float(v):
    if v is None:
        return None
    try:
        fv = float(v)
        return fv if math.isfinite(fv) else None
    except Exception:
        return None


def _normalize_mode(raw_mode: str) -> str:
    text = (raw_mode or "").strip()
    for key, val in MODE_ALIAS.items():
        if key in text:
            return val
    if text in MODE_ORDER:
        return text
    return "unknown"


def _normalize_component_prefix(component: str) -> str:
    c = (component or "").strip()
    if c.startswith("TR_"):
        return "CSBD_" + c[3:]
    if c.startswith("CS_"):
        return "CSB_" + c[3:]
    return c


def get_fit_group(exp_id: str) -> str:
    return "CA" if str(exp_id).startswith("CA_") else "OTHERS"


def get_support_group(exp_id: str) -> str:
    exp = str(exp_id)
    if exp.startswith("CA_"):
        return "CA"
    if exp.startswith("TB_"):
        return "TB"
    if exp.startswith("CSBD_"):
        return "CSBD"
    if exp.startswith("CSB_"):
        return "CSB"
    return "CSB"


COMPONENT_ORDER = [_normalize_component_prefix(x) for x in COMPONENT_ORDER_RAW]
COMPONENT_RANK = {name: idx for idx, name in enumerate(COMPONENT_ORDER)}


def _split_exp_id(exp_id: str):
    if "_" not in exp_id:
        return exp_id, ""
    comp, direction = exp_id.rsplit("_", 1)
    if direction in DIRECTION_RANK:
        return comp, direction
    return exp_id, ""


def exp_sort_key(exp_id: str):
    comp, direction = _split_exp_id(exp_id)
    return (COMPONENT_RANK.get(comp, 10000), DIRECTION_RANK.get(direction, 10000), exp_id)


PIPE_FIXED_MASS_KG = 388.0
PIPE_YAW_INERTIA_BASE = 259.7
HANGER_ECCENTRICITY_M = 0.9


def _param_key(exp_id: str, mode: str) -> tuple[str, str]:
    return (str(exp_id), str(mode))


def _mean_positive(values) -> float:
    arr = [float(v) for v in values if v is not None and np.isfinite(v) and float(v) > 0]
    return float(np.mean(arr)) if arr else np.nan


def _select_params_for_mode(
    params_mean: dict[tuple[str, str], dict[str, float]],
    exp_id: str,
    mode: str,
) -> dict[str, float]:
    default = {"m": np.nan, "l": np.nan, "r": np.nan}
    direct = params_mean.get(_param_key(exp_id, mode))
    axial = params_mean.get(_param_key(exp_id, "axial"))
    if mode == "planar_rotation" and axial is not None:
        return dict(axial)
    if direct is not None:
        return dict(direct)
    if axial is not None:
        return dict(axial)
    return dict(default)


def _compute_planar_rotation_inertia(mass_total: float) -> float:
    return PIPE_YAW_INERTIA_BASE + (mass_total - PIPE_FIXED_MASS_KG) * (HANGER_ECCENTRICITY_M ** 2)


def _calc_mode_f_pend(mode: str, mass_total: float, l_val: float) -> float:
    if not (np.isfinite(mass_total) and np.isfinite(l_val) and mass_total > 0 and l_val > 0):
        return np.nan
    return float((1.0 / (2.0 * np.pi)) * math.sqrt(G_ACCEL / l_val))


def _calc_mode_k_obs(mode: str, mass_total: float, l_val: float, f_exp: float) -> float:
    if not (np.isfinite(mass_total) and np.isfinite(l_val) and np.isfinite(f_exp)):
        return np.nan
    if mass_total <= 0 or l_val <= 0 or f_exp <= 0:
        return np.nan
    if mode == "planar_rotation":
        f_ideal = _calc_mode_f_pend("axial", mass_total, l_val)
        if not np.isfinite(f_ideal):
            return np.nan
        inertia_z = _compute_planar_rotation_inertia(mass_total)
        return float(4.0 * (np.pi ** 2) * inertia_z * (f_exp ** 2 - f_ideal ** 2))
    omega = 2.0 * np.pi * f_exp
    return float(mass_total * (l_val ** 2) * (omega ** 2 - G_ACCEL / l_val))


def _calc_mode_f_fit(mode: str, mass_total: float, l_val: float, k_fit: float) -> float:
    if not (np.isfinite(mass_total) and np.isfinite(l_val) and np.isfinite(k_fit)):
        return np.nan
    if mass_total <= 0 or l_val <= 0:
        return np.nan
    f_ideal = _calc_mode_f_pend(mode, mass_total, l_val)
    if not np.isfinite(f_ideal):
        return np.nan
    if mode == "planar_rotation":
        inertia_z = _compute_planar_rotation_inertia(mass_total)
        inside = f_ideal ** 2 + k_fit / (4.0 * (np.pi ** 2) * inertia_z)
        return float(math.sqrt(max(inside, 0.0)))
    inside = G_ACCEL / l_val + k_fit / (mass_total * (l_val ** 2))
    return float((1.0 / (2.0 * np.pi)) * math.sqrt(max(inside, 0.0)))


def load_h5_modal_freq_means(h5_path: Path) -> Dict[Tuple[str, str], float]:
    bucket: Dict[Tuple[str, str], List[float]] = {}
    with h5py.File(h5_path, "r") as h5:
        for exp_id in h5["experiments"].keys():
            g = h5["experiments"][exp_id]
            if "modal_freq_modes" not in g or "modal_freq_values" not in g:
                continue
            for m_raw, f_raw in zip(g["modal_freq_modes"][:], g["modal_freq_values"][:]):
                mode = _normalize_mode(m_raw.decode("utf-8", errors="ignore") if isinstance(m_raw, bytes) else str(m_raw))
                fv = _to_float(f_raw)
                if mode in MODE_ORDER and fv is not None and fv > 0:
                    bucket.setdefault((exp_id, mode), []).append(fv)
    return {k: float(np.mean(v)) for k, v in bucket.items()}


def load_compare6(compare6_path: Path):
    wb = openpyxl.load_workbook(compare6_path, data_only=True)
    records = []
    freq_bucket: Dict[Tuple[str, str], List[float]] = {}

    for sname in wb.sheetnames:
        ws = wb[sname]
        last_component, last_direction = "", ""
        last_m, last_l, last_r = None, None, None

        for r in range(2, ws.max_row + 1):
            c_component = ws.cell(r, 2).value
            c_direction = ws.cell(r, 3).value
            c_mode = ws.cell(r, 4).value
            c_freq = ws.cell(r, 5).value
            c_m = ws.cell(r, 13).value  # M
            c_l = ws.cell(r, 14).value  # N
            c_r = ws.cell(r, 15).value  # O

            if isinstance(c_component, str) and c_component.strip():
                last_component = c_component.strip()
            component = _normalize_component_prefix(last_component)
            if not component:
                continue

            if isinstance(c_direction, str) and c_direction.strip():
                last_direction = c_direction.strip().upper()
            direction = last_direction
            if not direction:
                continue

            mv, lv, rv = _to_float(c_m), _to_float(c_l), _to_float(c_r)
            if mv is not None:
                last_m = mv
            if lv is not None:
                last_l = lv
            if rv is not None:
                last_r = rv

            mode = _normalize_mode(str(c_mode) if c_mode is not None else "")
            fv = _to_float(c_freq)
            exp_id = f"{component}_{direction}"

            records.append({
                "exp_id": exp_id,
                "mode": mode,
                "m": last_m,
                "l": last_l,
                "r": last_r,
                "f_excel": fv,
            })

            if mode in MODE_ORDER and fv is not None and fv > 0:
                freq_bucket.setdefault((exp_id, mode), []).append(fv)

    freq_means = {k: float(np.mean(v)) for k, v in freq_bucket.items()}
    return records, freq_means


def build_rows(compare_records, compare_freq_means, h5_freq_means):
    param_bucket = {}
    for rec in compare_records:
        mode = rec["mode"]
        if mode not in MODE_ORDER:
            continue
        key = _param_key(rec["exp_id"], mode)
        p = param_bucket.setdefault(key, {"m": [], "l": [], "r": []})
        for k in ["m", "l", "r"]:
            v = rec[k]
            if v is not None and v > 0:
                p[k].append(float(v))

    params_mean = {
        key: {k: _mean_positive(v[k]) for k in ["m", "l", "r"]}
        for key, v in param_bucket.items()
    }

    rows = []
    all_keys = sorted(set(compare_freq_means.keys()) | set(h5_freq_means.keys()))
    for exp_id, mode in all_keys:
        p = _select_params_for_mode(params_mean, exp_id, mode)
        f_excel = compare_freq_means.get((exp_id, mode), np.nan)
        f_h5 = h5_freq_means.get((exp_id, mode), np.nan)
        rows.append({
            "exp_id": exp_id,
            "mode": mode,
            "m": p["m"],
            "l": p["l"],
            "r": p["r"],
            "f_exp_excel": f_excel,
            "f_exp_h5": f_h5,
            "f_exp": f_excel,
        })
    return rows


def _iqr_mask(values: np.ndarray) -> np.ndarray:
    q1, q3 = np.quantile(values, 0.25), np.quantile(values, 0.75)
    iqr = q3 - q1
    return (values >= q1 - 1.5 * iqr) & (values <= q3 + 1.5 * iqr)


def fit_mode_legacy(rows_mode, outlier="none", residual_thresh=0.2):
    valid = []
    for r in rows_mode:
        if all(np.isfinite(r[k]) for k in ["m", "l", "r", "f_exp"]) and r["m"] > 0 and r["l"] > 0 and r["r"] > 0 and r["f_exp"] > 0:
            rr = dict(r)
            rr["k_obs"] = _calc_mode_k_obs(rr["mode"], rr["m"], rr["l"], rr["f_exp"])
            rr["pi_k"] = rr["k_obs"] / (rr["m"] * G_ACCEL * rr["r"])
            rr["pi_l"] = rr["l"] / rr["r"]
            if np.isfinite(rr["pi_k"]) and rr["pi_k"] > 0 and np.isfinite(rr["pi_l"]) and rr["pi_l"] > 0:
                valid.append(rr)

    if len(valid) < 2:
        return FitStats(rows_mode[0]["mode"], len(rows_mode), 0, outlier, np.nan, np.nan, np.nan, np.nan, np.nan), rows_mode

    x_all = np.log(np.array([r["pi_l"] for r in valid]))
    y_all = np.log(np.array([r["pi_k"] for r in valid]))
    mask = np.ones_like(x_all, dtype=bool)

    if outlier == "iqr" and len(y_all) >= 4:
        mask = _iqr_mask(y_all)
    elif outlier == "residual" and len(y_all) >= 4:
        b0, a0 = np.polyfit(x_all, y_all, 1)
        A0, B0 = float(np.exp(a0)), float(b0)
        rels = []
        for r in valid:
            k_fit0 = A0 * r["m"] * G_ACCEL * r["r"] * ((r["l"] / r["r"]) ** B0)
            f_fit0 = _calc_mode_f_fit(r["mode"], r["m"], r["l"], k_fit0)
            rels.append(abs(f_fit0 - r["f_exp"]) / max(r["f_exp"], 1e-12))
        mask = np.array(rels) <= residual_thresh
        if mask.sum() < 2:
            mask = np.ones_like(mask, dtype=bool)

    x, y = x_all[mask], y_all[mask]
    if len(x) < 2:
        x, y = x_all, y_all

    B, lnA = np.polyfit(x, y, 1)
    A = float(np.exp(lnA))
    B = float(B)

    for r in valid:
        k_fit = A * r["m"] * G_ACCEL * r["r"] * ((r["l"] / r["r"]) ** B)
        r["f_pend"] = _calc_mode_f_pend(r["mode"], r["m"], r["l"])
        r["f_fit"] = _calc_mode_f_fit(r["mode"], r["m"], r["l"], k_fit)
        r["err_abs"] = abs(r["f_fit"] - r["f_exp"])
        r["err_rel"] = r["err_abs"] / max(r["f_exp"], 1e-12)
        r["A"] = A
        r["B"] = B

    y_true = np.array([r["f_exp"] for r in valid])
    y_pred = np.array([r["f_fit"] for r in valid])
    ss_res = float(np.sum((y_true - y_pred) ** 2))
    ss_tot = float(np.sum((y_true - np.mean(y_true)) ** 2))
    r2 = 1 - ss_res / ss_tot if ss_tot > 0 else np.nan
    rmse = float(np.sqrt(np.mean((y_true - y_pred) ** 2)))
    mae = float(np.mean(np.abs(y_true - y_pred)))

    out_map = {(r["exp_id"], r["mode"]): r for r in valid}
    out_rows = []
    for r in rows_mode:
        rr = dict(r)
        v = out_map.get((r["exp_id"], r["mode"]))
        rr["f_pend"] = _calc_mode_f_pend(rr["mode"], rr.get("m", np.nan), rr.get("l", np.nan))
        if v is None:
            rr["f_fit"] = np.nan
            rr["err_abs"] = np.nan
            rr["err_rel"] = np.nan
            rr["A"] = A
            rr["B"] = B
        else:
            rr["f_pend"] = v["f_pend"]
            rr["f_fit"] = v["f_fit"]
            rr["err_abs"] = v["err_abs"]
            rr["err_rel"] = v["err_rel"]
            rr["A"] = v["A"]
            rr["B"] = v["B"]
        out_rows.append(rr)

    stats = FitStats(rows_mode[0]["mode"], len(rows_mode), len(valid), outlier, A, B, r2, rmse, mae)
    return stats, out_rows


def plot_dual_axis_compare(rows_mode, mode_stats, out_dir: Path, save=True, show=True):
    mode = rows_mode[0]["mode"]
    d = [r for r in rows_mode if np.isfinite(r.get("f_exp", np.nan)) and np.isfinite(r.get("f_pend", np.nan))]
    if not d:
        print(f"[SKIP] dual-axis plot: {mode}")
        return
    d = sorted(d, key=lambda z: exp_sort_key(z["exp_id"]))
    x = np.arange(len(d))
    labels = [r["exp_id"] for r in d]
    f_exp = np.array([r["f_exp"] for r in d], dtype=float)
    f_pend = np.array([r["f_pend"] for r in d], dtype=float)
    f_fit = np.array([r.get("f_fit", np.nan) for r in d], dtype=float)
    fit_mask = np.isfinite(f_fit)

    fig, ax1 = plt.subplots(figsize=(14, 6))
    l1 = ax1.scatter(x, f_exp, s=16, marker="o", color="#1f77b4", alpha=0.9, label="Exp Frequency")
    l2 = ax1.scatter(x, f_pend, s=16, marker="s", color="#6c757d", alpha=0.85, label="Pendulum Freq")
    l3 = ax1.scatter(x[fit_mask], f_fit[fit_mask], s=20, marker="^", color=MODE_COLOR.get(mode, "#d00000"), alpha=0.9, label="Fitted Freq")
    y_stack = np.vstack([f_exp, f_pend, f_fit])
    y_min = np.nanmin(y_stack, axis=0)
    y_max = np.nanmax(y_stack, axis=0)
    ax1.vlines(x, ymin=y_min, ymax=y_max, color="gray", alpha=0.3, lw=1.5, zorder=0)
    if np.count_nonzero(~fit_mask) > 0:
        ax1.scatter(x[~fit_mask], f_exp[~fit_mask], marker="x", s=28, color="#ee9b00", label="Not used in fit", zorder=6)
    ax1.set_ylabel("Frequency (Hz)", fontweight="bold")
    ax1.set_xlabel("Experiment", fontweight="bold")
    ax1.set_ylim(0.2, 0.9)
    ax1.set_xticks(x)
    ax1.set_xticklabels(labels, rotation=80, ha="right", fontsize=7)
    ax1.grid(False)

    ax2 = ax1.twinx()
    width = 0.15
    x_pend = x - width
    x_fit = x + width
    err_pend_pct = np.abs(f_pend - f_exp) / np.maximum(f_exp, 1e-12) * 100.0
    err_fit_pct = np.array([100.0 * r.get("err_rel", np.nan) for r in d], dtype=float)
    pend_mask = np.isfinite(err_pend_pct)
    fit_err_mask = np.isfinite(err_fit_pct)
    pend_color = "#8c92ac"
    fit_color = "#e07a5f"
    ax2.vlines(x_pend[pend_mask], ymin=0, ymax=err_pend_pct[pend_mask], color=pend_color, alpha=0.7, lw=2.8)
    ax2.vlines(x_fit[fit_err_mask], ymin=0, ymax=err_fit_pct[fit_err_mask], color=fit_color, alpha=0.7, lw=2.8)
    l4 = ax2.plot(x_pend[pend_mask], err_pend_pct[pend_mask], "o", color=pend_color, ms=4, alpha=0.95, label="Pendulum Error (%)")[0]
    l5 = ax2.plot(x_fit[fit_err_mask], err_fit_pct[fit_err_mask], "o", color=fit_color, ms=4, alpha=0.95, label="Fitted Error (%)")[0]
    max_global_err = np.nanmax(np.concatenate([err_pend_pct[pend_mask], err_fit_pct[fit_err_mask]])) if (np.any(pend_mask) or np.any(fit_err_mask)) else 1.0
    if not np.isfinite(max_global_err) or max_global_err <= 0:
        max_global_err = 1.0
    ax2.set_ylim(0, max_global_err * 2.5)
    ax2.set_ylabel("Relative Error (%)", color=fit_color, fontweight="bold")
    ax2.tick_params(axis="y", colors=fit_color)

    handles = [l1, l2, l3, l4, l5]
    labels_legend = [h.get_label() for h in handles]
    ax1.legend(handles, labels_legend, loc="upper right", frameon=False, fontsize=9, ncol=3)
    ax1.set_title(f"{MODE_CN.get(mode, mode)}: Frequency & Error Dual-Axis (n={len(d)})")

    ca_st = mode_stats.get("CA")
    ot_st = mode_stats.get("OTHERS")
    if ca_st is not None and ot_st is not None:
        stat_txt = (
            f"CA: A={ca_st.A:.4g}, B={ca_st.B:.4g}, R2={ca_st.r2:.3f}, RMSE={ca_st.rmse:.4f}\n"
            f"SR: A={ot_st.A:.4g}, B={ot_st.B:.4g}, R2={ot_st.r2:.3f}, RMSE={ot_st.rmse:.4f}"
        )
    elif ca_st is not None:
        stat_txt = f"CA: A={ca_st.A:.4g}, B={ca_st.B:.4g}, R2={ca_st.r2:.3f}, RMSE={ca_st.rmse:.4f}"
    elif ot_st is not None:
        stat_txt = f"SR: A={ot_st.A:.4g}, B={ot_st.B:.4g}, R2={ot_st.r2:.3f}, RMSE={ot_st.rmse:.4f}"
    else:
        stat_txt = "No fit stats"

    ax1.text(0.01, 0.96, stat_txt,
             transform=ax1.transAxes, ha="left", va="top", fontsize=8,
             bbox={"facecolor": "white", "alpha": 0.8, "edgecolor": "none", "pad": 2})
    fig.subplots_adjust(left=0.05, right=0.95, top=0.92, bottom=0.25)
    if save:
        fig.savefig(out_dir / f"dimscaling_dual_{mode}.png", dpi=600)
    if show:
        plt.show()
    else:
        plt.close(fig)


# 当前使用版本：允许 k_obs<=0 的样本也参与拟合（非线性拟合 pi_k = A*(l/r)^B）
from scipy.optimize import curve_fit

def fit_mode(rows_mode, outlier="none", residual_thresh=0.2, fit_group="ALL"):
    valid = []
    for r in rows_mode:
        if all(np.isfinite(r[k]) for k in ["m", "l", "r", "f_exp"]) and r["m"] > 0 and r["l"] > 0 and r["r"] > 0 and r["f_exp"] > 0:
            rr = dict(r)
            rr["fit_group"] = fit_group
            rr["k_obs"] = _calc_mode_k_obs(rr["mode"], rr["m"], rr["l"], rr["f_exp"])
            rr["pi_k"] = rr["k_obs"] / (rr["m"] * G_ACCEL * rr["r"])
            rr["pi_l"] = rr["l"] / rr["r"]
            if np.isfinite(rr["pi_k"]) and np.isfinite(rr["pi_l"]) and rr["pi_l"] > 0:
                valid.append(rr)

    if len(valid) < 2:
        return FitStats(rows_mode[0]["mode"], fit_group, len(rows_mode), 0, outlier, np.nan, np.nan, np.nan, np.nan, np.nan), rows_mode

    x_raw = np.array([r["pi_l"] for r in valid], dtype=float)
    y_raw = np.array([r["pi_k"] for r in valid], dtype=float)
    mask = np.ones_like(x_raw, dtype=bool)

    if outlier == "iqr" and len(y_raw) >= 4:
        mask = _iqr_mask(y_raw)
    elif outlier == "residual" and len(y_raw) >= 4:
        def _model(x, A, B):
            return A * np.power(x, B)
        A0 = float(np.nanmedian(y_raw)) if len(y_raw) else 1.0
        if not np.isfinite(A0) or A0 == 0:
            A0 = 1.0
        B0 = 0.5
        try:
            popt0, _ = curve_fit(_model, x_raw, y_raw, p0=[A0, B0], maxfev=30000)
            A0, B0 = float(popt0[0]), float(popt0[1])
        except Exception:
            pass
        rels = []
        for r in valid:
            k_fit0 = A0 * r["m"] * G_ACCEL * r["r"] * ((r["l"] / r["r"]) ** B0)
            f_fit0 = _calc_mode_f_fit(r["mode"], r["m"], r["l"], k_fit0)
            rels.append(abs(f_fit0 - r["f_exp"]) / max(r["f_exp"], 1e-12))
        mask = np.array(rels) <= residual_thresh
        if mask.sum() < 2:
            mask = np.ones_like(mask, dtype=bool)

    used_fit_rows = [valid[i] for i, keep in enumerate(mask) if keep]
    x = x_raw[mask]
    y = y_raw[mask]
    if len(x) < 2:
        used_fit_rows = list(valid)
        x = x_raw
        y = y_raw

    def _model(xv, A, B):
        return A * np.power(xv, B)

    A_init = float(np.nanmedian(y)) if len(y) else 1.0
    if not np.isfinite(A_init) or A_init == 0:
        A_init = 1.0
    B_init = 0.5

    try:
        popt, _ = curve_fit(_model, x, y, p0=[A_init, B_init], maxfev=50000)
        A = float(popt[0])
        B = float(popt[1])
    except Exception:
        B = 0.5
        xb = np.power(x, B)
        denom = float(np.dot(xb, xb))
        A = float(np.dot(xb, y) / denom) if denom > 0 else float(A_init)

    used_keys = {(r["exp_id"], r["mode"]) for r in used_fit_rows}
    for r in valid:
        r["fit_used"] = (r["exp_id"], r["mode"]) in used_keys
        k_fit = A * r["m"] * G_ACCEL * r["r"] * ((r["l"] / r["r"]) ** B)
        r["f_pend"] = _calc_mode_f_pend(r["mode"], r["m"], r["l"])
        r["f_fit"] = _calc_mode_f_fit(r["mode"], r["m"], r["l"], k_fit)
        r["err_abs"] = abs(r["f_fit"] - r["f_exp"])
        r["err_rel"] = r["err_abs"] / max(r["f_exp"], 1e-12)
        r["A"] = A
        r["B"] = B
        r["fit_group"] = fit_group

    y_true = np.array([r["f_exp"] for r in valid])
    y_pred = np.array([r["f_fit"] for r in valid])
    ss_res = float(np.sum((y_true - y_pred) ** 2))
    ss_tot = float(np.sum((y_true - np.mean(y_true)) ** 2))
    r2 = 1 - ss_res / ss_tot if ss_tot > 0 else np.nan
    rmse = float(np.sqrt(np.mean((y_true - y_pred) ** 2)))
    mae = float(np.mean(np.abs(y_true - y_pred)))

    out_map = {(r["exp_id"], r["mode"]): r for r in valid}
    out_rows = []
    for r in rows_mode:
        rr = dict(r)
        rr["fit_group"] = fit_group
        rr["f_pend"] = _calc_mode_f_pend(rr["mode"], rr.get("m", np.nan), rr.get("l", np.nan))
        v = out_map.get((r["exp_id"], r["mode"]))
        if v is None:
            rr["f_fit"] = np.nan
            rr["err_abs"] = np.nan
            rr["err_rel"] = np.nan
            rr["fit_used"] = False
            rr["A"] = A
            rr["B"] = B
        else:
            rr["f_pend"] = v["f_pend"]
            rr["f_fit"] = v["f_fit"]
            rr["err_abs"] = v["err_abs"]
            rr["err_rel"] = v["err_rel"]
            rr["fit_used"] = bool(v.get("fit_used", False))
            rr["A"] = v["A"]
            rr["B"] = v["B"]
            rr["fit_group"] = v.get("fit_group", fit_group)
        out_rows.append(rr)

    stats = FitStats(rows_mode[0]["mode"], fit_group, len(rows_mode), len(used_fit_rows), outlier, A, B, r2, rmse, mae)
    return stats, out_rows


# 出版版主函数：Split Panels (6x1) with grouped color/marker mapping
def plot_split_panels_publication(rows_by_mode, stats_by_mode, out_dir: Path):
    from matplotlib.lines import Line2D

    present_modes = [m for m in MODE_ORDER if m in rows_by_mode and m in stats_by_mode]
    if not present_modes:
        print("[SKIP] no modes available")
        return

    # ---- Palette / markers (support colors + series markers) ----
    support_colors = SUPPORT_COLOR

    # ---- x domain ----
    exp_union = sorted(
        {r["exp_id"] for m in present_modes for r in rows_by_mode[m]},
        key=exp_sort_key,
    )
    exp_to_idx = {eid: i for i, eid in enumerate(exp_union)}
    x = np.arange(len(exp_union), dtype=float)
    spans = compute_prefix_spans(exp_union)
    is_ca = np.asarray([str(eid).startswith("CA") for eid in exp_union], dtype=bool)

    # ---- layout ----
    fig = plt.figure(figsize=(7.35, 7.55))
    gs_main = fig.add_gridspec(4, 1, height_ratios=[1.0, 1.0, 1.0, 0.68], hspace=0.23)
    axes_pairs = []
    for i in range(len(present_modes)):
        sub = gs_main[i].subgridspec(2, 1, height_ratios=[3, 1], hspace=0.12)
        if i == 0:
            ax_freq = fig.add_subplot(sub[0])
            ax_err = fig.add_subplot(sub[1], sharex=ax_freq)
        else:
            ax_freq = fig.add_subplot(sub[0], sharex=axes_pairs[0][1])
            ax_err = fig.add_subplot(sub[1], sharex=axes_pairs[0][1])
        axes_pairs.append((ax_freq, ax_err))
    ax_label = fig.add_subplot(gs_main[3], sharex=axes_pairs[0][1])

    fig.subplots_adjust(left=0.105, right=0.985, top=0.855, bottom=0.165)

    for i, mode in enumerate(present_modes):
        ax_freq, ax_err = axes_pairs[i]
        rows = rows_by_mode[mode]
        mode_stats = stats_by_mode[mode]

        f_exp = np.full(len(exp_union), np.nan)
        f_pend = np.full(len(exp_union), np.nan)
        f_fit = np.full(len(exp_union), np.nan)
        err_fit_pct = np.full(len(exp_union), np.nan)

        for r in rows:
            idx = exp_to_idx.get(r["exp_id"])
            if idx is None:
                continue
            f_exp[idx] = r["f_exp"] if np.isfinite(r.get("f_exp", np.nan)) else np.nan
            f_pend[idx] = r["f_pend"] if np.isfinite(r.get("f_pend", np.nan)) else np.nan
            f_fit[idx] = r["f_fit"] if np.isfinite(r.get("f_fit", np.nan)) else np.nan
            err_fit_pct[idx] = 100.0 * r["err_rel"] if np.isfinite(r.get("err_rel", np.nan)) else np.nan

        err_pend_pct = np.abs(f_pend - f_exp) / np.maximum(f_exp, 1e-12) * 100.0

        exp_mask = np.isfinite(f_exp)
        pend_mask = np.isfinite(f_pend)
        fit_mask = np.isfinite(f_fit)
        support_masks = {
            sup: np.asarray([get_support_group(eid) == sup for eid in exp_union], dtype=bool)
            for sup in ["CA", "TB", "CSBD", "CSB"]
        }

        # alternating light background by prefix span
        for j, (s, e) in enumerate(spans):
            if j % 2 == 0:
                ax_freq.axvspan(s - 0.5, e + 0.5, color="#f8f9fa", alpha=1.0, linewidth=0)
                ax_err.axvspan(s - 0.5, e + 0.5, color="#f8f9fa", alpha=1.0, linewidth=0)

        # dumbbell stems
        y_stack = np.vstack([f_exp, f_pend, f_fit])
        y_min = np.nanmin(y_stack, axis=0)
        y_max = np.nanmax(y_stack, axis=0)
        stem_mask = np.isfinite(y_min) & np.isfinite(y_max)
        ax_freq.vlines(x[stem_mask], ymin=y_min[stem_mask], ymax=y_max[stem_mask], color="gray", alpha=0.28, lw=1.1, zorder=1)

        # frequency points
        for sup in ["CA", "TB", "CSBD", "CSB"]:
            sup_mask = support_masks[sup]
            sup_color = support_colors[sup]
            exp_color = SPLIT_EXP_COLOR.get(sup, sup_color)
            ax_freq.scatter(
                x[exp_mask & sup_mask],
                f_exp[exp_mask & sup_mask],
                s=POINT_SIZE["split_exp"],
                marker="o",
                color=exp_color,
                alpha=0.96,
                edgecolors="none",
                linewidths=0.0,
                zorder=4,
            )
            ax_freq.scatter(
                x[pend_mask & sup_mask],
                f_pend[pend_mask & sup_mask],
                s=POINT_SIZE["split_pend"],
                marker="s",
                color=sup_color,
                alpha=0.28,
                edgecolors=sup_color,
                linewidths=0.60,
                zorder=4,
            )
            ax_freq.scatter(
                x[fit_mask & sup_mask],
                f_fit[fit_mask & sup_mask],
                s=POINT_SIZE["split_fit"],
                marker="^",
                color=sup_color,
                alpha=0.95,
                edgecolors="white",
                linewidths=0.30,
                zorder=5,
            )

        ax_freq.set_ylim(0.2, 0.9)
        ax_freq.set_ylabel("Freq (Hz)", fontsize=11.2, labelpad=6)
        ax_freq.yaxis.set_label_coords(-0.052, 0.5)
        ax_freq.set_title("")
        ax_freq.text(
            0.975,
            0.085,
            MODE_CN.get(mode, mode),
            transform=ax_freq.transAxes,
            ha="right",
            va="bottom",
            fontsize=11.2,
            color="#202020",
            bbox=dict(
                facecolor="white",
                edgecolor="#555555",
                linewidth=0.55,
                alpha=0.88,
                boxstyle="square,pad=0.22",
            ),
            zorder=8,
        )
        ax_freq.tick_params(axis="x", labelbottom=False)
        apply_axis_style(ax_freq, grid=False)

        # error lollipop (dodged)
        width = 0.15
        x_pend = x - width
        x_fit = x + width

        err_arrays = []
        for sup in ["CA", "TB", "CSBD", "CSB"]:
            sup_mask = support_masks[sup]
            sup_color = support_colors[sup]
            pend_err_mask = np.isfinite(err_pend_pct) & sup_mask
            fit_err_mask = np.isfinite(err_fit_pct) & sup_mask
            if np.any(pend_err_mask):
                ax_err.vlines(x_pend[pend_err_mask], 0, err_pend_pct[pend_err_mask], color=sup_color, alpha=0.18, lw=2.0, zorder=2)
                ax_err.plot(
                    x_pend[pend_err_mask],
                    err_pend_pct[pend_err_mask],
                    "s",
                    color=sup_color,
                    ms=POINT_SIZE["split_err_pend_ms"],
                    alpha=0.22,
                    markeredgecolor=sup_color,
                    markeredgewidth=0.55,
                    zorder=3,
                )
                err_arrays.append(err_pend_pct[pend_err_mask])
            if np.any(fit_err_mask):
                ax_err.vlines(x_fit[fit_err_mask], 0, err_fit_pct[fit_err_mask], color=sup_color, alpha=0.85, lw=2.0, zorder=2)
                ax_err.plot(
                    x_fit[fit_err_mask],
                    err_fit_pct[fit_err_mask],
                    "^",
                    color=sup_color,
                    ms=POINT_SIZE["split_err_fit_ms"],
                    alpha=0.95,
                    markeredgecolor="white",
                    markeredgewidth=0.35,
                    zorder=3,
                )
                err_arrays.append(err_fit_pct[fit_err_mask])

        max_err = np.nanmax(np.concatenate(err_arrays)) if err_arrays else 1.0
        if not np.isfinite(max_err) or max_err <= 0:
            max_err = 1.0

        ax_err.set_ylim(0, max_err * 1.1)
        ax_err.set_yticks([0, 20, 40])
        ax_err.set_ylabel("Error (%)", fontsize=11.2, labelpad=6)
        ax_err.yaxis.set_label_coords(-0.052, 0.5)
        ax_err.tick_params(axis="x", labelbottom=False)
        apply_axis_style(ax_err, grid=False)

    # bottom x
    experiment_label = draw_experiment_tree_axis(ax_label, exp_union)
    x_limits = (-0.50, len(exp_union) - 0.50)
    for ax_freq, ax_err in axes_pairs:
        ax_freq.set_xlim(*x_limits)
        ax_err.set_xlim(*x_limits)
    ax_label.set_xlim(*x_limits)

    # legend: series markers + support colors
    legend_handles = [
        Line2D([0], [0], marker="o", linestyle="None", markersize=6, markerfacecolor="#333333", markeredgecolor="#333333", label="Exp Freq"),
        Line2D([0], [0], marker="s", linestyle="None", markersize=6, markerfacecolor="#777777", markeredgecolor="#777777", alpha=0.35, label="Pendulum Freq"),
        Line2D([0], [0], marker="^", linestyle="None", markersize=6, markerfacecolor="#777777", markeredgecolor="#777777", label="Fitted Freq"),
        Line2D([0], [0], color=SUPPORT_COLOR["CA"], lw=2.2, label="CA"),
        Line2D([0], [0], color=SUPPORT_COLOR["TB"], lw=2.2, label="TB"),
        Line2D([0], [0], color=SUPPORT_COLOR["CSBD"], lw=2.2, label="CSBD"),
        Line2D([0], [0], color=SUPPORT_COLOR["CSB"], lw=2.2, label="CSB"),
    ]
    fig.subplots_adjust(top=0.885)

    style_figure(fig, theme=WORD_THEME, grid=False)
    apply_insert_text_style_for_figure(fig, latex_text_width_in=6.45, pad_inches=0.0)
    legend = axes_pairs[0][0].legend(
        legend_handles,
        [h.get_label() for h in legend_handles],
        loc="upper center",
        ncol=7,
        frameon=False,
        fontsize=8.8,
        handlelength=1.45,
        columnspacing=0.95,
        borderaxespad=0.18,
        bbox_to_anchor=(0.5, 0.985),
    )
    legend.set_zorder(10)
    experiment_label.set_fontsize(12.5)
    experiment_label.set_fontweight("semibold")
    for ax_freq, ax_err in axes_pairs:
        ax_freq.yaxis.set_label_coords(-0.052, 0.5)
        ax_err.yaxis.set_label_coords(-0.052, 0.5)
    fig.savefig(_workbench_format_path(out_dir / "dimscaling_split_panels_combined.png", "png"), dpi=600, bbox_inches="tight", pad_inches=0.03)
    fig.savefig(
        _workbench_format_path(out_dir / "dimscaling_split_panels_combined.tiff", "tiff"),
        format="tiff",
        bbox_inches="tight",
        pad_inches=0.03,
        pil_kwargs={"compression": "tiff_lzw"},
    )
    fig.savefig(
        _workbench_format_path(out_dir / "dimscaling_split_panels_combined.pdf", "pdf"),
        format="pdf",
        bbox_inches="tight",
        pad_inches=0.0,
    )
    plt.show()
    print("[DONE] split-panels figure exported to:", out_dir)








# 归一化频率-参数图（3x3，分组散点 + 公式反算单摆拟合线）
from matplotlib.lines import Line2D

def _build_inverse_pendulum_curve(
    rows_group: list[dict],
    A: float,
    B: float,
    x_var: str,
    x_grid: np.ndarray,
) -> tuple[np.ndarray, np.ndarray] | None:
    """Build pendulum normalized fit curve by inverse use of fitted formula.

    Fitted formula (already solved):
        k_theta = A * m * g * r * (l/r)^B
        f_fit   = (1/(2*pi))*sqrt(g/l + k_theta/(m*l^2))

    For normalized pendulum points y = f_pend / f_exp,
    inverse-normalize with model-predicted experimental frequency:
        y_curve = f_pend_model / f_fit_model
    """
    if x_grid.size < 2 or not (np.isfinite(A) and np.isfinite(B)):
        return None

    m_vals = np.asarray([float(r.get("m", np.nan)) for r in rows_group], dtype=np.float64)
    l_vals = np.asarray([float(r.get("l", np.nan)) for r in rows_group], dtype=np.float64)
    r_vals = np.asarray([float(r.get("r", np.nan)) for r in rows_group], dtype=np.float64)
    good = np.isfinite(m_vals) & np.isfinite(l_vals) & np.isfinite(r_vals) & (m_vals > 0) & (l_vals > 0) & (r_vals > 0)
    if np.count_nonzero(good) < 2:
        return None

    m_ref = float(np.nanmedian(m_vals[good]))
    l_ref = float(np.nanmedian(l_vals[good]))
    r_ref = float(np.nanmedian(r_vals[good]))
    if not (np.isfinite(m_ref) and np.isfinite(l_ref) and np.isfinite(r_ref)):
        return None
    mode = str(rows_group[0].get("mode", "")) if rows_group else ""

    y = np.full_like(x_grid, np.nan, dtype=np.float64)
    for i, xv in enumerate(x_grid):
        if not np.isfinite(xv) or xv <= 0:
            continue
        if x_var == "m":
            m, l, rr = float(xv), l_ref, r_ref
        elif x_var == "l":
            m, l, rr = m_ref, float(xv), r_ref
        else:  # r
            m, l, rr = m_ref, l_ref, float(xv)
        if m <= 0 or l <= 0 or rr <= 0:
            continue

        f_pend = _calc_mode_f_pend(mode, m, l)
        k_fit = A * m * G_ACCEL * rr * ((l / rr) ** B)
        f_fit = _calc_mode_f_fit(mode, m, l, k_fit)
        if f_fit <= 0:
            continue
        y[i] = f_pend / f_fit

    ok = np.isfinite(y)
    if np.count_nonzero(ok) < 2:
        return None
    return x_grid[ok], y[ok]


def _find_curve_band_entry(
    x_curve: np.ndarray,
    y_curve: np.ndarray,
    y_lower: float,
    y_upper: float,
) -> tuple[float, float] | None:
    """Return the first x where the curve enters the target y-band."""
    if x_curve.size < 2 or y_curve.size < 2:
        return None

    x = np.asarray(x_curve, dtype=np.float64)
    y = np.asarray(y_curve, dtype=np.float64)
    ok = np.isfinite(x) & np.isfinite(y)
    if np.count_nonzero(ok) < 2:
        return None
    x = x[ok]
    y = y[ok]

    inside = (y >= y_lower) & (y <= y_upper)
    inside_idx = np.flatnonzero(inside)
    if inside_idx.size > 0:
        idx = int(inside_idx[0])
        return float(x[idx]), float(y[idx])

    for target in (y_lower, y_upper):
        diff = y - target
        for i in range(len(x) - 1):
            d0 = diff[i]
            d1 = diff[i + 1]
            if not (np.isfinite(d0) and np.isfinite(d1)):
                continue
            if d0 == 0:
                return float(x[i]), float(y[i])
            if d0 * d1 > 0:
                continue
            denom = y[i + 1] - y[i]
            if not np.isfinite(denom) or abs(denom) < 1e-12:
                continue
            t = (target - y[i]) / denom
            if 0.0 <= t <= 1.0:
                x_hit = x[i] + t * (x[i + 1] - x[i])
                return float(x_hit), float(target)
    return None


def plot_normalized_freq_vs_params(mode_rows, stats_by_mode, out_dir: Path, save=True, show=True):
    fig, axes = plt.subplots(3, 3, figsize=(9.0, 5.9), constrained_layout=False, sharex="col", sharey=True)

    ca_color = GROUP_COLOR["CA"]
    oth_color = GROUP_COLOR["OTHERS"]
    oth_plot_label = "SR"
    support_colors = SUPPORT_COLOR
    support_markers = SUPPORT_MARKERS
    band_color = "#ffe8a1"
    mode_short = {
        "radial": "Radial",
        "axial": "Axial",
        "planar_rotation": "Rotation",
    }
    x_specs = [
        ("m", "Mass", "Mass $m$ (kg)"),
        ("l", "Length", "Length $l$ (m)"),
        ("r", "Radius", "Radius $r$ (m)"),
    ]
    band_lower = 0.95
    band_upper = 1.05

    legend_handles = [
        Line2D([0], [0], color="black", lw=1.2, ls="--", label="Exp Baseline (f/f_exp = 1)"),
        Line2D([0], [0], color=band_color, lw=6.0, alpha=0.45, label="Negligible-error band (1.0 +/- 5%)"),
        Line2D([0], [0], marker=support_markers["CA"], color="none", markerfacecolor=support_colors["CA"], markeredgecolor=support_colors["CA"], alpha=0.85, markersize=6, label="CA points"),
        Line2D([0], [0], marker=support_markers["TB"], color="none", markerfacecolor=support_colors["TB"], markeredgecolor=support_colors["TB"], alpha=0.85, markersize=6, label="TB points"),
        Line2D([0], [0], marker=support_markers["CSBD"], color="none", markerfacecolor=support_colors["CSBD"], markeredgecolor=support_colors["CSBD"], alpha=0.85, markersize=6, label="CSBD points"),
        Line2D([0], [0], marker=support_markers["CSB"], color="none", markerfacecolor=support_colors["CSB"], markeredgecolor=support_colors["CSB"], alpha=0.85, markersize=6, label="CSB points"),
        Line2D([0], [0], color=ca_color, lw=2.0, ls="-", alpha=0.95, label="Inverse-fit curve (CA)"),
        Line2D([0], [0], color=oth_color, lw=2.0, ls="-", alpha=0.95, label="Inverse-fit curve (SR)"),
    ]

    for row_idx, mode in enumerate(MODE_ORDER):
        rows = mode_rows.get(mode, [])
        mode_stats = stats_by_mode.get(mode, {})
        ca_st = mode_stats.get("CA")
        oth_st = mode_stats.get("OTHERS")

        for col_idx, (x_var, x_title, x_label) in enumerate(x_specs):
            ax = axes[row_idx, col_idx]
            if x_var == "l":
                ax.axhspan(band_lower, band_upper, color=band_color, alpha=0.45, zorder=0)
                ax.axhline(band_lower, color="#d4a017", lw=0.9, ls=":", alpha=0.8, zorder=1)
                ax.axhline(band_upper, color="#d4a017", lw=0.9, ls=":", alpha=0.8, zorder=1)
            ax.axhline(1.0, color="black", lw=1.2, ls="--")

            if not rows:
                ax.text(0.5, 0.5, "No data", ha="center", va="center", transform=ax.transAxes, fontsize=9)
            else:
                rows_ca, rows_oth = [], []
                support_points = {sup: {"x": [], "yp": [], "yf": []} for sup in ["CA", "TB", "CSBD", "CSB"]}
                l_threshold_notes = []
                x_values_all = []

                for d in rows:
                    f_exp = float(d.get("f_exp", np.nan))
                    f_pend = float(d.get("f_pend", np.nan))
                    f_fit = float(d.get("f_fit", np.nan))
                    m = float(d.get("m", np.nan))
                    l = float(d.get("l", np.nan))
                    rr = float(d.get("r", np.nan))
                    if not (np.isfinite(f_exp) and f_exp > 0 and np.isfinite(f_pend) and np.isfinite(f_fit)):
                        continue
                    if not (np.isfinite(m) and np.isfinite(l) and np.isfinite(rr)):
                        continue
                    x_val = m if x_var == "m" else (l if x_var == "l" else rr)
                    fit_group = d.get("fit_group", get_fit_group(d.get("exp_id", "")))
                    support_group = get_support_group(d.get("exp_id", ""))
                    yp = f_pend / f_exp
                    yf = f_fit / f_exp
                    x_values_all.append(float(x_val))
                    support_points[support_group]["x"].append(x_val)
                    support_points[support_group]["yp"].append(yp)
                    support_points[support_group]["yf"].append(yf)
                    if fit_group == "CA":
                        rows_ca.append(d)
                    else:
                        rows_oth.append(d)

                for sup in ["CA", "TB", "CSBD", "CSB"]:
                    pts = support_points[sup]
                    if not pts["x"]:
                        continue
                    x_np = np.asarray(pts["x"], dtype=np.float64)
                    color = support_colors[sup]
                    marker = support_markers[sup]
                    fit_size = POINT_SIZE["norm_fit_ca"] if sup == "CA" else POINT_SIZE["norm_fit_others"]
                    ax.scatter(x_np, np.asarray(pts["yp"]), s=POINT_SIZE["norm_pend"], marker=marker, color=color, alpha=0.28, edgecolors="none")
                    ax.scatter(x_np, np.asarray(pts["yf"]), s=fit_size, marker=marker, color=color, alpha=0.88, edgecolors="white", linewidths=0.25)

                x_all_np = np.asarray(x_values_all, dtype=np.float64)
                x_all_np = x_all_np[np.isfinite(x_all_np)]
                x_plot_min = None
                x_plot_max = None
                if x_all_np.size:
                    x_plot_min = float(np.nanmin(x_all_np))
                    x_plot_max = float(np.nanmax(x_all_np))
                    if np.isclose(x_plot_min, x_plot_max):
                        span = max(abs(x_plot_min) * 0.05, 1e-6)
                        x_plot_min -= span
                        x_plot_max += span
                    ax.set_xlim(x_plot_min, x_plot_max)

                if rows_ca and ca_st is not None and np.isfinite(ca_st.A) and np.isfinite(ca_st.B):
                    if x_plot_min is not None and x_plot_max is not None:
                        xg = np.linspace(x_plot_min, x_plot_max, 400)
                        curve = _build_inverse_pendulum_curve(rows_ca, ca_st.A, ca_st.B, x_var, xg)
                        if curve is not None:
                            ax.plot(curve[0], curve[1], color=ca_color, lw=2.0, ls="-", alpha=0.95)
                            if x_var == "l":
                                hit = _find_curve_band_entry(curve[0], curve[1], band_lower, band_upper)
                                if hit is not None:
                                    x_hit, y_hit = hit
                                    ax.scatter([x_hit], [y_hit], s=36, marker="o", color=ca_color, edgecolors="white", linewidths=0.8, zorder=6)
                                    ax.axvline(x_hit, color=ca_color, lw=1.0, ls="--", alpha=0.75, zorder=2)
                                    l_threshold_notes.append(("CA", x_hit, y_hit, ca_color))

                if rows_oth and oth_st is not None and np.isfinite(oth_st.A) and np.isfinite(oth_st.B):
                    if x_plot_min is not None and x_plot_max is not None:
                        xg = np.linspace(x_plot_min, x_plot_max, 400)
                        curve = _build_inverse_pendulum_curve(rows_oth, oth_st.A, oth_st.B, x_var, xg)
                        if curve is not None:
                            ax.plot(curve[0], curve[1], color=oth_color, lw=2.0, ls="-", alpha=0.95)
                            if x_var == "l":
                                hit = _find_curve_band_entry(curve[0], curve[1], band_lower, band_upper)
                                if hit is not None:
                                    x_hit, y_hit = hit
                                    ax.scatter([x_hit], [y_hit], s=36, marker="o", color=oth_color, edgecolors="white", linewidths=0.8, zorder=6)
                                    ax.axvline(x_hit, color=oth_color, lw=1.0, ls="--", alpha=0.75, zorder=2)
                                    l_threshold_notes.append((oth_plot_label, x_hit, y_hit, oth_color))

                if x_var == "l" and l_threshold_notes:
                    l_threshold_notes.sort(key=lambda item: item[1])
                    x_min, x_max = ax.get_xlim()
                    y_min, y_max = ax.get_ylim()
                    x_slots = {
                        "CA": x_min + 0.67 * (x_max - x_min),
                        oth_plot_label: x_min + 0.33 * (x_max - x_min),
                    }
                    y_slots = {
                        "CA": y_min + 0.28 * (y_max - y_min),
                        oth_plot_label: y_min + 0.68 * (y_max - y_min),
                    }
                    for group_label, x_hit, y_hit, note_color in l_threshold_notes:
                        x_text = x_slots.get(group_label, x_min + 0.50 * (x_max - x_min))
                        y_text = y_slots.get(group_label, y_min + 0.48 * (y_max - y_min))
                        ax.annotate(
                            f"{group_label}: {x_hit:.3f} m",
                            xy=(x_hit, y_hit),
                            xytext=(x_text, y_text),
                            textcoords="data",
                            ha="center",
                            va="center",
                            fontsize=7.8,
                            color=note_color,
                            arrowprops=dict(arrowstyle="->", color=note_color, lw=0.9, shrinkA=0, shrinkB=3),
                            bbox=dict(facecolor="white", edgecolor=note_color, alpha=0.76, pad=1.2),
                            zorder=7,
                        )

            panel_idx = row_idx * len(x_specs) + col_idx
            panel_label = chr(ord("a") + panel_idx)
            if row_idx == 0:
                ax.set_title(x_title, fontsize=10.0, pad=8)
            ax.set_xlabel(x_label if row_idx == len(MODE_ORDER) - 1 else "", fontsize=10.0, labelpad=6)
            ax.set_ylabel("")
            ax.set_ylim(0.5, 1.5)
            ax.set_yticks([0.5, 1.0, 1.5])
            if col_idx > 0:
                ax.tick_params(axis="y", labelleft=False)
            if row_idx < len(MODE_ORDER) - 1:
                ax.tick_params(axis="x", labelbottom=False)
            if col_idx == 0:
                mode_label = mode_short.get(mode, MODE_CN.get(mode, mode))
                ax.text(
                    -0.20,
                    0.50,
                    mode_label,
                    transform=ax.transAxes,
                    rotation=90,
                    ha="center",
                    va="center",
                    fontsize=11.2,
                    fontweight="semibold",
                    clip_on=False,
                    zorder=8,
                )
            panel_text = ax.text(
                0.98,
                0.96,
                f"({panel_label})",
                transform=ax.transAxes,
                ha="right",
                va="top",
                fontsize=11.2,
                fontweight="semibold",
                bbox=dict(facecolor="white", edgecolor="none", alpha=0.78, pad=0.35),
                zorder=8,
            )
            panel_text.set_gid("keep-fontsize")
            apply_axis_style(ax, grid=False)

    fig.legend(
        handles=legend_handles,
        loc="upper center",
        ncol=4,
        frameon=False,
        bbox_to_anchor=(0.5, 0.94),
        fontsize=8.6,
        columnspacing=1.0,
        handlelength=1.8,
    )
    fig.supylabel(
        "Normalized frequency ratio ($f/f_{\\mathrm{exp}}$)",
        x=0.055,
        fontsize=10.0,
    )
    fig.subplots_adjust(left=0.145, right=0.99, top=0.82, bottom=0.12, hspace=0.18, wspace=0.10)
    if save:
        out_dir.mkdir(parents=True, exist_ok=True)
        style_figure(fig, theme=WORD_THEME, grid=False)
        apply_insert_text_style_for_figure(fig, latex_text_width_in=6.45, pad_inches=0.0)
        fig.savefig(_workbench_format_path(out_dir / "dimscaling_normalized_freq_vs_params.png", "png"), dpi=600, bbox_inches="tight")
        fig.savefig(
            _workbench_format_path(out_dir / "dimscaling_normalized_freq_vs_params.tiff", "tiff"),
            format="tiff",
            bbox_inches="tight",
            pil_kwargs={"compression": "tiff_lzw"},
        )
        fig.savefig(
            _workbench_format_path(out_dir / "dimscaling_normalized_freq_vs_params.pdf", "pdf"),
            format="pdf",
            bbox_inches="tight",
            pad_inches=0.0,
        )
    if show:
        plt.show()
    else:
        plt.close(fig)



# ===== extracted from notebook cell 04 =====


def _rename_if_exists(src: Path, dst: Path) -> None:
    if src.exists():
        dst.parent.mkdir(parents=True, exist_ok=True)
        if dst.exists():
            dst.unlink()
        src.replace(dst)


LEGACY_FIGURE_OUTPUTS = {
    "09_decay_waveform_and_fft_contrast": ["08_decay_waveform_and_fft_contrast", "02_fast_decay_frequency_check"],
    "11_dimscaling_split_panels_combined": ["09_dimscaling_split_panels_combined"],
    "12_cross_validation_predicted_vs_observed": ["11_cross_validation_predicted_vs_observed"],
    "13_cross_type_validation": ["16_cross_type_validation"],
    "14_equivalent_damping_ratio_vs_normalized_frequency_amplitude_sensitivity_beta": [
        "12_equivalent_damping_ratio_vs_normalized_frequency_amplitude_sensitivity_beta"
    ],
    "15_local_damping_ratio_vs_amplitude": ["13_local_damping_ratio_vs_amplitude"],
    "16_cycle_energy_dissipation_deltaE_over_E": ["14_cycle_energy_dissipation_deltaE_over_E"],
    "17_nonlinear_indicator_applicability_nli": ["15_nonlinear_indicator_applicability_nli"],
}


def _remove_legacy_figure_outputs(current_stem: str) -> None:
    for legacy_stem in LEGACY_FIGURE_OUTPUTS.get(current_stem, []):
        for out_dir, ext in [
            (WORKBENCH_PNG_OUT_DIR, ".png"),
            (WORKBENCH_TIFF_OUT_DIR, ".tiff"),
            (WORKBENCH_PDF_OUT_DIR, ".pdf"),
        ]:
            for suffix in ("", "_enhanced"):
                legacy_path = out_dir / f"{legacy_stem}{suffix}{ext}"
                if legacy_path.exists():
                    legacy_path.unlink()


def _workbench_format_path(path: Path, fmt: str = "png") -> Path:
    fmt_norm = fmt.lower()
    if fmt_norm == "png":
        out_dir = WORKBENCH_PNG_OUT_DIR
        ext = ".png"
    elif fmt_norm in {"tif", "tiff"}:
        out_dir = WORKBENCH_TIFF_OUT_DIR
        ext = ".tiff"
    elif fmt_norm == "pdf":
        out_dir = WORKBENCH_PDF_OUT_DIR
        ext = ".pdf"
    else:
        out_dir = WORKBENCH_OUT_DIR / fmt_norm
        ext = f".{fmt_norm}"
    out_dir.mkdir(parents=True, exist_ok=True)
    return out_dir / f"{path.stem}{ext}"


def prepare_dimscaling_context():
    h5_freq_means = load_h5_modal_freq_means(H5_PATH)
    compare_records, compare_freq_means = load_compare6(COMPARE_XLSX)
    rows = build_rows(compare_records, compare_freq_means, h5_freq_means)

    conflicts = []
    for r in rows:
        f_h5, f_ex = r["f_exp_h5"], r["f_exp_excel"]
        if np.isfinite(f_h5) and np.isfinite(f_ex) and f_ex > 0:
            diff = f_h5 - f_ex
            rel = abs(diff) / f_ex
        else:
            diff, rel = np.nan, np.nan
        conflicts.append({
            "exp_id": r["exp_id"],
            "mode": r["mode"],
            "freq_compare6": f_ex,
            "freq_h5": f_h5,
            "diff_h5_minus_compare6": diff,
            "rel_diff": rel,
        })

    stats_by_mode = {}
    rows_by_mode = {}
    for mode in MODE_ORDER:
        rows_mode = [r for r in rows if r["mode"] == mode]
        if not rows_mode:
            continue
        rows_ca = [r for r in rows_mode if get_fit_group(r["exp_id"]) == "CA"]
        rows_others = [r for r in rows_mode if get_fit_group(r["exp_id"]) == "OTHERS"]
        mode_stats = {}
        out_rows = []
        for fit_group, sub_rows in [("CA", rows_ca), ("OTHERS", rows_others)]:
            if not sub_rows:
                continue
            st, sub_out = fit_mode(sub_rows, outlier=OUTLIER, residual_thresh=RESIDUAL_THRESH, fit_group=fit_group)
            mode_stats[fit_group] = st
            out_rows.extend(sub_out)
        out_rows = sorted(out_rows, key=lambda z: exp_sort_key(z["exp_id"]))
        stats_by_mode[mode] = mode_stats
        rows_by_mode[mode] = out_rows

    missing_report_rows = []
    for r in rows:
        reason = ""
        m, l, rr = r.get("m", np.nan), r.get("l", np.nan), r.get("r", np.nan)
        f_ex, f_h5 = r.get("f_exp_excel", np.nan), r.get("f_exp_h5", np.nan)
        if not np.isfinite(f_ex):
            reason = "missing_compare6_frequency"
        elif not np.isfinite(m) or not np.isfinite(l) or not np.isfinite(rr) or m <= 0 or l <= 0 or rr <= 0:
            reason = "missing_or_invalid_m_l_r"
        else:
            k_obs = _calc_mode_k_obs(str(r.get("mode", "")), m, l, f_ex)
            reason = "invalid_k_obs_not_used_in_fit" if not np.isfinite(k_obs) else "used_in_fit"
        source_status = "both" if (np.isfinite(f_ex) and np.isfinite(f_h5)) else ("compare6_only" if np.isfinite(f_ex) else ("h5_only" if np.isfinite(f_h5) else "none"))
        missing_report_rows.append({
            "exp_id": r["exp_id"],
            "mode": r["mode"],
            "fit_group": get_fit_group(r["exp_id"]),
            "m": m,
            "l": l,
            "r": rr,
            "f_exp_compare6": f_ex,
            "f_exp_h5": f_h5,
            "freq_source_status": source_status,
            "fit_status": reason,
        })
    return {
        "rows": rows,
        "rows_by_mode": rows_by_mode,
        "stats_by_mode": stats_by_mode,
        "conflicts": conflicts,
        "missing_report_rows": missing_report_rows,
    }


def plot_dimscaling_split_panels(ctx: dict):
    print("[DIM] Exporting split-panels figure...")
    plot_split_panels_publication(ctx["rows_by_mode"], ctx["stats_by_mode"], DIM_OUT_DIR)
    _rename_if_exists(
        _workbench_format_path(DIM_OUT_DIR / "dimscaling_split_panels_combined.png", "png"),
        _workbench_format_path(DIM_OUT_DIR / "11_dimscaling_split_panels_combined.png", "png"),
    )
    _rename_if_exists(
        _workbench_format_path(DIM_OUT_DIR / "dimscaling_split_panels_combined.tiff", "tiff"),
        _workbench_format_path(DIM_OUT_DIR / "11_dimscaling_split_panels_combined.tiff", "tiff"),
    )
    _rename_if_exists(
        _workbench_format_path(DIM_OUT_DIR / "dimscaling_split_panels_combined.pdf", "pdf"),
        _workbench_format_path(DIM_OUT_DIR / "11_dimscaling_split_panels_combined.pdf", "pdf"),
    )
    _remove_legacy_figure_outputs("11_dimscaling_split_panels_combined")


def plot_dimscaling_normalized_freq(ctx: dict):
    print("[DIM] Plotting normalized frequency vs parameters...")
    plot_normalized_freq_vs_params(ctx["rows_by_mode"], ctx["stats_by_mode"], DIM_OUT_DIR, save=True, show=True)
    _rename_if_exists(
        _workbench_format_path(DIM_OUT_DIR / "dimscaling_normalized_freq_vs_params.png", "png"),
        _workbench_format_path(DIM_OUT_DIR / "10_normalized_frequency_ratios_vs_parameters.png", "png"),
    )
    _rename_if_exists(
        _workbench_format_path(DIM_OUT_DIR / "dimscaling_normalized_freq_vs_params.tiff", "tiff"),
        _workbench_format_path(DIM_OUT_DIR / "10_normalized_frequency_ratios_vs_parameters.tiff", "tiff"),
    )
    _rename_if_exists(
        _workbench_format_path(DIM_OUT_DIR / "dimscaling_normalized_freq_vs_params.pdf", "pdf"),
        _workbench_format_path(DIM_OUT_DIR / "10_normalized_frequency_ratios_vs_parameters.pdf", "pdf"),
    )


def run_all_dimscaling():
    ctx = prepare_dimscaling_context()
    print("[DONE] modes fitted:", list(ctx["stats_by_mode"].keys()))

    from collections import Counter
    cnt = Counter((row["mode"], row["fit_status"]) for row in ctx["missing_report_rows"])
    print("\n=== Missing/Fit Status Summary ===")
    for mode in MODE_ORDER:
        print(f"[{mode}]")
        for status in ["used_in_fit", "missing_compare6_frequency", "missing_or_invalid_m_l_r", "invalid_k_obs_not_used_in_fit"]:
            print(f"  {status}: {cnt.get((mode, status), 0)}")

    missing_csv = DIM_OUT_DIR / "03_dimscaling_missing_reason_report.csv"
    with missing_csv.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(
            f,
            fieldnames=[
                "exp_id", "mode", "fit_group", "m", "l", "r",
                "f_exp_compare6", "f_exp_h5",
                "freq_source_status", "fit_status",
            ],
        )
        w.writeheader()
        for row in ctx["missing_report_rows"]:
            w.writerow(row)
    print("[DONE] missing reason report:", missing_csv)

    plot_dimscaling_split_panels(ctx)

    import pandas as pd
    df_missing = pd.DataFrame(ctx["missing_report_rows"])
    display(df_missing[df_missing["fit_status"] != "used_in_fit"].head(50))
    plot_dimscaling_normalized_freq(ctx)

    def write_csv(path: Path, fieldnames: list[str], rows: list[dict]):
        with path.open("w", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(f, fieldnames=fieldnames)
            w.writeheader()
            for r in rows:
                w.writerow(r)

    stats_rows = []
    point_rows = []
    for mode in MODE_ORDER:
        if mode not in ctx["stats_by_mode"]:
            continue
        mode_stats = ctx["stats_by_mode"][mode]
        for fit_group, st in mode_stats.items():
            stats_rows.append({
                "mode": st.mode,
                "fit_group": fit_group,
                "n_total": st.n_total,
                "n_used": st.n_used,
                "outlier_policy": st.outlier_policy,
                "A": st.A,
                "B": st.B,
                "r2": st.r2,
                "rmse": st.rmse,
                "mae": st.mae,
            })
        for r in ctx["rows_by_mode"][mode]:
            point_rows.append({
                "exp_id": r["exp_id"],
                "mode": r["mode"],
                "fit_group": r.get("fit_group", get_fit_group(r["exp_id"])),
                "m": r["m"],
                "l": r["l"],
                "r": r["r"],
                "f_exp_compare6": r["f_exp_excel"],
                "f_exp_h5": r["f_exp_h5"],
                "f_exp_used": r["f_exp"],
                "f_pend": r.get("f_pend", np.nan),
                "f_fit": r.get("f_fit", np.nan),
                "fit_used": r.get("fit_used", False),
                "err_abs_hz": r.get("err_abs", np.nan),
                "err_rel": r.get("err_rel", np.nan),
                "A_mode": r.get("A", np.nan),
                "B_mode": r.get("B", np.nan),
            })

    write_csv(
        DIM_OUT_DIR / "03_dimscaling_fit_summary.csv",
        ["mode", "fit_group", "n_total", "n_used", "outlier_policy", "A", "B", "r2", "rmse", "mae"],
        stats_rows,
    )
    write_csv(
        DIM_OUT_DIR / "03_dimscaling_freq_conflict_report.csv",
        ["exp_id", "mode", "freq_compare6", "freq_h5", "diff_h5_minus_compare6", "rel_diff"],
        ctx["conflicts"],
    )
    write_csv(
        DIM_OUT_DIR / "03_dimscaling_pointwise_results.csv",
        [
            "exp_id", "mode", "fit_group", "m", "l", "r",
            "f_exp_compare6", "f_exp_h5", "f_exp_used",
            "f_pend", "f_fit", "fit_used", "err_abs_hz", "err_rel",
            "A_mode", "B_mode",
        ],
        point_rows,
    )

    print("[DONE] CSV exported to:", DIM_OUT_DIR)



# ===== extracted from notebook cell 06 =====

from pathlib import Path
import sys
import time
import math

import numpy as np
import openpyxl
import h5py
from scipy import signal
from matplotlib.figure import Figure
import matplotlib.pyplot as plt
from IPython.display import display, Image

THIS_DIR = Path.cwd().resolve()
CODE_ROOT = THIS_DIR.parent if THIS_DIR.name == "damping" else THIS_DIR
if str(CODE_ROOT) not in sys.path:
    sys.path.insert(0, str(CODE_ROOT))

from damping.damping_batch import batch_extract_damping
from damping.damping_stats import (
    compute_damping_statistics,
    plot_damping_summary_by_support,
    plot_drift_ratio_summary,
    plot_xi_amplitude,
)
from damping.damping_h5_schema import read_damping_results, read_cycle_damping
from damping.amplitude_anomaly import filter_result_rows_for_amplitude
from damping.extra_plots import plot_damping_drift_correlation

# 路径参数（与 bat 保持一致）
BASE_DIR = CODE_ROOT
PROJECT_DIR = BASE_DIR.parent
H5_PATH = PROJECT_DIR / "data" / "derived" / "modal_db.h5"
OUT_DIR = DAMP_OUT_DIR
COMPARE_XLSX = PROJECT_DIR / "data" / "metadata" / "COMPARE_6.xlsx"

# 计算参数
BANDWIDTH_PCT = 15.0
BUTTER_ORDER = 4
OVERWRITE = True
WORKERS = None  # None=自动

print('[CONFIG] H5_PATH =', H5_PATH)
print('[CONFIG] OUT_DIR =', OUT_DIR)
print('[CONFIG] COMPARE_XLSX =', COMPARE_XLSX)
print('[CONFIG] bandwidth_pct =', BANDWIDTH_PCT)
print('[CONFIG] butter_order =', BUTTER_ORDER)
print('[CONFIG] overwrite =', OVERWRITE)
print('[CONFIG] workers =', WORKERS)



# ===== extracted from notebook cell 07 =====

import matplotlib.pyplot as plt
from damping.advanced_plots import (
    plot_nli_applicability,
    plot_backbone_examples,
    plot_cycle_energy_dissipation,
    plot_selected_fast_decay_frequency_check,
)
from damping.signal_access import load_preferred_signal


WORD_PAGE_WIDTH = WORD_A4_TEXT_WIDTH_IN
FIGSIZE_TRIPTYCH = (WORD_A4_TEXT_WIDTH_IN, 4.25)
FIGSIZE_SPLIT = (WORD_A4_TEXT_WIDTH_IN, 8.25)
FIGSIZE_NORMALIZED = (WORD_A4_TEXT_WIDTH_IN, 7.0)
FIGSIZE_SUMMARY = (WORD_A4_TEXT_WIDTH_IN, 6.35)
FIGSIZE_WIDE = (WORD_A4_TEXT_WIDTH_IN, 4.35)
FIGSIZE_SQUARE_TRIPTYCH = (WORD_A4_TEXT_WIDTH_IN, 3.35)
FIGSIZE_FLAT_TRIPTYCH = (WORD_A4_TEXT_WIDTH_IN, 2.95)
FIGSIZE_FLAT_CORRELATION = (WORD_A4_TEXT_WIDTH_IN, 3.35)
FIGSIZE_FLAT_SINGLE = (WORD_A4_TEXT_WIDTH_IN, 1.95)
FIGSIZE_TALL = (WORD_A4_TEXT_WIDTH_IN, 6.45)
FIGSIZE_MID = (WORD_A4_TEXT_WIDTH_IN, 5.0)
FIGSIZE_TRIPTYCH_A4 = (WORD_A4_TEXT_WIDTH_IN, 3.9)
FIGSIZE_SIGNAL_QUALITY = (WORD_A4_TEXT_WIDTH_IN, 3.45)


def _apply_word_style(fig, title_size=15.0, label_size=14.0, tick_size=12.0, legend_size=11.2, text_size=10.5, suptitle_size=15.0):
    if getattr(fig, "_suptitle", None) is not None:
        fig._suptitle.remove()
        fig._suptitle = None
    for ax in fig.axes:
        keep_local_sizes = ax.get_gid() == "keep-local-fontsizes"
        if keep_local_sizes:
            continue
        if ax.get_title():
            ax.set_title(ax.get_title(), fontsize=title_size, pad=8)
        if ax.get_xlabel():
            ax.set_xlabel(ax.get_xlabel(), fontsize=label_size, labelpad=7)
        if ax.get_ylabel():
            ax.set_ylabel(ax.get_ylabel(), fontsize=label_size, labelpad=7)
        ax.tick_params(labelsize=tick_size)
        leg = ax.get_legend()
        if leg is not None:
            for txt in leg.get_texts():
                txt.set_fontsize(legend_size)
            if leg.get_title() is not None:
                leg.get_title().set_fontsize(legend_size)
        for txt in ax.texts:
            try:
                if txt.get_gid() == "keep-fontsize":
                    continue
                txt.set_fontsize(max(float(txt.get_fontsize()), text_size))
            except Exception:
                pass
    for leg in getattr(fig, "legends", []):
        for txt in leg.get_texts():
            txt.set_fontsize(legend_size)
        if leg.get_title() is not None:
            leg.get_title().set_fontsize(legend_size)


def _save_workbench_figure(
    fig,
    png: Path,
    size: tuple[float, float],
    dpi: int = 600,
    font_profile="auto",
) -> None:
    fig.set_size_inches(*size, forward=True)
    if font_profile is None:
        _apply_word_style(fig)
        style_figure(fig, theme=WORD_THEME, grid=False)
    else:
        style_figure(fig, theme=WORD_THEME, grid=False)
        if font_profile == "auto":
            apply_insert_text_style_for_figure(fig, latex_text_width_in=6.45, pad_inches=0.0)
        else:
            apply_insert_text_style(fig, font_profile)
    save_figure_bundle(
        fig,
        png.with_suffix(""),
        formats=("png",),
        dpi=dpi,
        split_format_dirs=True,
    )
    pdf_path = WORKBENCH_PDF_OUT_DIR / f"{png.stem}.pdf"
    fig.savefig(
        pdf_path,
        format="pdf",
        bbox_inches="tight",
        pad_inches=0.0,
        facecolor=fig.get_facecolor(),
    )


def _display_workbench_figure(png: Path) -> None:
    show_saved_figure(_workbench_format_path(png, "png"))


def _normalize_component_prefix(component: str) -> str:
    c = (component or "").strip()
    if c.startswith("TR_"):
        return "CSBD_" + c[3:]
    if c.startswith("CS_"):
        return "CSB_" + c[3:]
    return c


def load_compare6_params(compare_xlsx: Path) -> dict[tuple[str, str], tuple[float, float, float]]:
    wb = openpyxl.load_workbook(compare_xlsx, data_only=True)
    bucket: dict[tuple[str, str], list[tuple[float, float, float]]] = {}
    for ws in wb.worksheets:
        last_component = ""
        last_direction = ""
        last_m = np.nan
        last_l = np.nan
        last_r = np.nan
        for rr in range(2, ws.max_row + 1):
            c_comp = ws.cell(rr, 2).value
            c_dir = ws.cell(rr, 3).value
            c_mode = ws.cell(rr, 4).value
            c_m = ws.cell(rr, 13).value
            c_l = ws.cell(rr, 14).value
            c_r = ws.cell(rr, 15).value

            if isinstance(c_comp, str) and c_comp.strip():
                last_component = c_comp.strip()
            if isinstance(c_dir, str) and c_dir.strip():
                last_direction = c_dir.strip().upper()

            try:
                v = float(c_m) if c_m is not None else np.nan
                if np.isfinite(v):
                    last_m = v
            except Exception:
                pass
            try:
                v = float(c_l) if c_l is not None else np.nan
                if np.isfinite(v):
                    last_l = v
            except Exception:
                pass
            try:
                v = float(c_r) if c_r is not None else np.nan
                if np.isfinite(v):
                    last_r = v
            except Exception:
                pass

            comp = _normalize_component_prefix(last_component)
            direction = last_direction
            if not comp or not direction:
                continue
            if not (np.isfinite(last_m) and np.isfinite(last_l) and np.isfinite(last_r)):
                continue
            mode = _normalize_mode(str(c_mode) if c_mode is not None else "")
            if mode not in MODE_ORDER:
                continue
            key = _param_key(f"{comp}_{direction}", mode)
            bucket.setdefault(key, []).append((float(last_m), float(last_l), float(last_r)))

    out: dict[tuple[str, str], tuple[float, float, float]] = {}
    for k, arr in bucket.items():
        a = np.asarray(arr, dtype=np.float64)
        out[k] = (float(np.mean(a[:, 0])), float(np.mean(a[:, 1])), float(np.mean(a[:, 2])))
    return out


def _butter_bandpass(sig: np.ndarray, dt: float, f0: float, bw_pct: float = 15.0, order: int = 4) -> np.ndarray:
    fs = 1.0 / dt
    nyq = 0.5 * fs
    lo = max(0.05, f0 * (1.0 - bw_pct / 100.0))
    hi = min(nyq * 0.98, f0 * (1.0 + bw_pct / 100.0))
    if not (np.isfinite(lo) and np.isfinite(hi) and hi > lo):
        return sig.copy()
    b, a = signal.butter(order, [lo / nyq, hi / nyq], btype="band")
    return signal.filtfilt(b, a, sig)


def _compute_decay_cycles(row: dict) -> float:
    f0 = float(row.get("freq_target", np.nan))
    t_start = float(row.get("t_start", np.nan))
    t_end = float(row.get("t_end", np.nan))
    if not (np.isfinite(f0) and f0 > 0 and np.isfinite(t_start) and np.isfinite(t_end) and t_end > t_start):
        return float("nan")
    return float((t_end - t_start) * f0)


def _choose_decay_fft_cases(valid_rows: list[dict], all_rows: list[dict]) -> tuple[dict, dict]:
    valid_candidates = []
    for row in valid_rows:
        rr = dict(row)
        cyc = _compute_decay_cycles(rr)
        if not np.isfinite(cyc) or cyc <= 0:
            continue
        rr["decay_cycles"] = cyc
        valid_candidates.append(rr)
    if not valid_candidates:
        raise RuntimeError("No valid damping rows with usable decay windows were found.")

    all_candidates = []
    for row in all_rows:
        rr = dict(row)
        cyc = _compute_decay_cycles(rr)
        if not np.isfinite(cyc) or cyc <= 0:
            continue
        rr["decay_cycles"] = cyc
        all_candidates.append(rr)
    if not all_candidates:
        raise RuntimeError("No damping rows with usable decay windows were found.")

    rapid_pool = [
        r for r in all_candidates
        if str(r.get("exp_id", "")) == "CSBD_07_A" and str(r.get("channel_id", "")) == "ch_07"
    ]
    if not rapid_pool:
        raise RuntimeError("Requested rapid-decay case CSBD_07_A | ch_07 was not found in damping rows.")

    rapid = min(rapid_pool, key=lambda r: (r["decay_cycles"], -float(r.get("r_squared", np.nan))))

    def _normalize_channel_id(ch: str) -> str:
        return str(ch).replace("_", "").lower()

    requested_good_exp = "CA_07+13_Mid_RR"
    requested_good_channels = {"ch18", "ch_18"}
    good_pool = [
        r for r in valid_candidates
        if str(r.get("exp_id", "")) == requested_good_exp
        and _normalize_channel_id(str(r.get("channel_id", ""))) in {_normalize_channel_id(ch) for ch in requested_good_channels}
    ]
    if not good_pool:
        raise RuntimeError(
            f"Requested good-decay case {requested_good_exp} | ch18 was not found in valid damping rows."
        )
    good = max(good_pool, key=lambda r: (r["decay_cycles"], float(r.get("r_squared", np.nan))))
    return good, rapid


def _extract_decay_fft_payload(h5_path: Path, row: dict, bandwidth_pct: float, butter_order: int) -> dict:
    exp_id = str(row["exp_id"])
    ch = str(row["channel_id"])
    f0 = float(row["freq_target"])
    t_start = float(row["t_start"])
    t_end = float(row["t_end"])

    with h5py.File(h5_path, "r") as h5:
        g = h5["experiments"][exp_id]
        t = np.asarray(g["time"][:], dtype=np.float64)
        x = load_preferred_signal(g["channels"][ch])

    dt = float(np.median(np.diff(t)))
    xf = _butter_bandpass(x, dt, f0, bw_pct=bandwidth_pct, order=butter_order)
    env = np.abs(signal.hilbert(xf))

    mask = (t >= t_start) & (t <= t_end)
    if int(np.sum(mask)) < 32:
        raise RuntimeError(f"Decay window too short for {exp_id} | {ch}")

    # FFT still uses the analysis decay window, but plotting uses the full post-peak waveform.
    t_seg = np.asarray(t[mask] - t_start, dtype=np.float64)
    sig_seg = np.asarray(xf[mask], dtype=np.float64)
    env_seg = np.asarray(env[mask], dtype=np.float64)

    post_mask = t >= t_start
    t_post = np.asarray(t[post_mask] - t_start, dtype=np.float64)
    sig_post = np.asarray(xf[post_mask], dtype=np.float64)
    env_post = np.asarray(env[post_mask], dtype=np.float64)

    sig_peak = max(float(np.nanmax(np.abs(sig_post))), 1e-12)
    env_peak = max(float(np.nanmax(env_post)), 1e-12)
    sig_norm = sig_post / sig_peak
    env_norm = env_post / env_peak

    win = np.hanning(len(sig_seg))
    spec = np.abs(np.fft.rfft(sig_seg * win))
    freqs = np.fft.rfftfreq(len(sig_seg), d=dt)
    spec_norm = spec / max(float(np.nanmax(spec)), 1e-12)
    band_mask = (freqs >= 0.05) & (freqs <= max(2.0, f0 * 2.5))
    if np.any(band_mask):
        band_freqs = freqs[band_mask]
        band_spec = spec_norm[band_mask]
        fft_peak = float(band_freqs[int(np.argmax(band_spec))])
    else:
        band_freqs = freqs
        band_spec = spec_norm
        fft_peak = float("nan")

    return {
        "exp_id": exp_id,
        "channel_id": ch,
        "mode": str(row.get("mode", "")),
        "support_type": str(row.get("support_type", "")),
        "decay_cycles": float(row.get("decay_cycles", _compute_decay_cycles(row))),
        "freq_auto": f0,
        "freq_peakcount": float(row.get("freq_peakcount", np.nan)),
        "freq_fft_peak": fft_peak,
        "t_seg": t_seg,
        "sig_seg_norm": sig_seg / sig_peak,
        "env_seg_norm": env_seg / env_peak,
        "t_plot": t_post,
        "sig_norm": sig_norm,
        "env_norm": env_norm,
        "freqs": band_freqs,
        "spec_norm": band_spec,
        "r_squared": float(row.get("r_squared", np.nan)),
    }


def prepare_damping_context(run_batch: bool = False) -> dict:
    print("[DAMP] H5_PATH =", H5_PATH)
    print("[DAMP] OUT_DIR =", OUT_DIR)
    print("[DAMP] Existing damping results will be reused when run_batch = False.")

    if run_batch:
        def _bar(pct: float, width: int = 48) -> str:
            filled = int(width * max(0.0, min(1.0, pct)))
            return "[" + "#" * filled + "-" * (width - filled) + "]"

        t0 = time.time()

        def progress_cb(exp_id: str, current: int, total: int) -> None:
            pct = (current / total) if total > 0 else 0.0
            elapsed = time.time() - t0
            eta = (elapsed / max(current, 1)) * (total - current) if current > 0 else 0.0
            line = (
                f"\r{_bar(pct)} {pct*100:6.2f}% "
                f"({current:>3}/{total:<3}) "
                f"EXP={exp_id:<30} "
                f"elapsed={elapsed:7.1f}s eta={eta:7.1f}s"
            )
            print(line, end="", flush=True)

        print("[1/4] Running damping batch extraction...")
        summary = batch_extract_damping(
            h5_path=H5_PATH,
            bandwidth_pct=BANDWIDTH_PCT,
            butter_order=BUTTER_ORDER,
            workers=WORKERS,
            overwrite=OVERWRITE,
            progress_callback=progress_cb,
        )
        print()
        print("Batch summary:", summary)
    else:
        print("[DAMP] Skipping batch extraction and reading existing damping results from modal_db.h5.")

    rows_by_exp = read_damping_results(H5_PATH, exp_id=None)
    all_rows = []
    for eid, rs in rows_by_exp.items():
        for r in rs:
            rr = dict(r)
            if not str(rr.get("exp_id", "")).strip():
                rr["exp_id"] = str(eid)
            all_rows.append(rr)
    valid_rows = [r for r in all_rows if bool(r.get("valid", False))]
    param_map = load_compare6_params(COMPARE_XLSX)
    for rr in all_rows:
        exp_id = str(rr.get("exp_id", ""))
        mode = _normalize_mode(str(rr.get("mode", "")))
        params = param_map.get(_param_key(exp_id, mode))
        if params is None and mode == "planar_rotation":
            params = param_map.get(_param_key(exp_id, "axial"))
        rr["l_m"] = float(params[1]) if params is not None and np.isfinite(params[1]) else np.nan
    amp_valid_rows = filter_result_rows_for_amplitude(valid_rows, H5_PATH)
    supports = ["CA", "TB", "CSBD", "CSB"]
    support_color = dict(SUPPORT_COLORS)
    print(
        f"[INFO] damping rows: total={len(all_rows)}, valid={len(valid_rows)}, "
        f"amp-valid={len(amp_valid_rows)}, mapped-exp={len(param_map)}"
    )
    return {
        "all_rows": all_rows,
        "valid_rows": valid_rows,
        "amp_valid_rows": amp_valid_rows,
        "param_map": param_map,
        "supports": supports,
        "support_color": support_color,
    }


def plot_damping_summary_figure(ctx: dict) -> None:
    print("[DAMP] Plotting damping summary...")
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    fig = Figure(figsize=FIGSIZE_SUMMARY)
    plot_damping_summary_by_support(H5_PATH, fig)
    png = OUT_DIR / "S1_damping_summary_by_support_and_mode.png"
    _save_workbench_figure(fig, png, FIGSIZE_SUMMARY)
    _display_workbench_figure(png)


def plot_damping_freq_drift_figure(ctx: dict) -> None:
    print("[DAMP] Plotting normalized frequency-amplitude slope summary...")
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    fig = Figure(figsize=FIGSIZE_SUMMARY)
    plot_drift_ratio_summary(H5_PATH, fig)
    png = OUT_DIR / "S2_normalized_frequency_amplitude_sensitivity_beta_summary.png"
    _save_workbench_figure(fig, png, FIGSIZE_SUMMARY)
    _display_workbench_figure(png)


def plot_damping_xi_amplitude_figure(ctx: dict) -> None:
    print("[DAMP] Plotting damping ratio vs amplitude (mm/s)...")
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    fig = Figure(figsize=FIGSIZE_FLAT_TRIPTYCH)
    plot_xi_amplitude(H5_PATH, fig, exclude_amp_anomalies=True)
    png = OUT_DIR / "15_local_damping_ratio_vs_amplitude.png"
    _save_workbench_figure(fig, png, FIGSIZE_FLAT_TRIPTYCH)
    fig_enhanced = Figure(figsize=FIGSIZE_FLAT_TRIPTYCH)
    plot_xi_amplitude(H5_PATH, fig_enhanced, exclude_amp_anomalies=True, enhanced=True)
    png_enhanced = OUT_DIR / "15_local_damping_ratio_vs_amplitude_enhanced.png"
    _save_workbench_figure(fig_enhanced, png_enhanced, FIGSIZE_FLAT_TRIPTYCH)
    _remove_legacy_figure_outputs("15_local_damping_ratio_vs_amplitude")
    _display_workbench_figure(png_enhanced)


def plot_damping_statistics_refresh(ctx: dict) -> None:
    stats_dict = compute_damping_statistics(H5_PATH)
    print(f"[stats done] groups={len(stats_dict)}")


def _compute_eta_clean(ctx: dict) -> tuple[list[dict], list[str], dict[str, str]]:
    supports = ctx["supports"]
    support_color = ctx["support_color"]
    valid_rows = ctx["valid_rows"]
    param_map = ctx["param_map"]

    eta_rows = []
    g = 9.81
    for r in valid_rows:
        exp_id = str(r.get("exp_id", ""))
        mode = _normalize_mode(str(r.get("mode", "")))
        params = param_map.get(_param_key(exp_id, mode))
        if params is None and mode == "planar_rotation":
            params = param_map.get(_param_key(exp_id, "axial"))
        if params is None:
            continue
        m, l, rr = params
        f = float(r.get("freq_target", np.nan))
        if not (np.isfinite(m) and np.isfinite(l) and np.isfinite(rr) and np.isfinite(f)):
            continue
        if m <= 0 or l <= 0 or rr <= 0 or f <= 0:
            continue
        k_theta = _calc_mode_k_obs(mode, m, l, f)
        eta = k_theta / (m * g * rr)
        if np.isfinite(eta):
            eta_rows.append({"support_type": str(r.get("support_type", "")), "mode": str(r.get("mode", "")), "eta_eq": float(eta)})

    from collections import defaultdict

    def _clean_eta(rows, lo_win, hi_win):
        bucket = defaultdict(list)
        out = []
        for row in rows:
            v = float(row["eta_eq"])
            if np.isfinite(v) and (lo_win <= v <= hi_win):
                bucket[(row["support_type"], row["mode"])].append(v)
        bounds = {}
        for k, arr in bucket.items():
            a = np.asarray(arr, dtype=np.float64)
            if a.size >= 4:
                q1, q3 = np.quantile(a, [0.25, 0.75])
                iqr = q3 - q1
                bounds[k] = (q1 - 1.5 * iqr, q3 + 1.5 * iqr)
            else:
                bounds[k] = (lo_win, hi_win)
        for row in rows:
            k = (row["support_type"], row["mode"])
            v = float(row["eta_eq"])
            if not np.isfinite(v):
                continue
            if not (lo_win <= v <= hi_win):
                continue
            lo, hi = bounds.get(k, (lo_win, hi_win))
            if lo <= v <= hi:
                out.append(row)
        return out

    eta_clean = _clean_eta(eta_rows, -1.0, 2.0)
    if len(eta_clean) == 0:
        all_eta = np.asarray([float(r["eta_eq"]) for r in eta_rows if np.isfinite(float(r["eta_eq"]))], dtype=np.float64)
        if all_eta.size > 0:
            p1, p99 = np.percentile(all_eta, [1, 99])
            eta_clean = _clean_eta(eta_rows, float(p1), float(p99))
    return eta_clean, supports, support_color


def plot_damping_deltaE_figure(ctx: dict) -> None:
    print("[DAMP] Plotting cycle energy dissipation vs amplitude (mm/s)...")
    supports = ctx["supports"]
    support_color = ctx["support_color"]
    fig = Figure(figsize=FIGSIZE_FLAT_TRIPTYCH)
    ede_stats = plot_cycle_energy_dissipation(
        h5_path=H5_PATH,
        fig=fig,
        supports=supports,
        support_colors=support_color,
        exclude_amp_anomalies=True,
        bins=16,
    )
    png = OUT_DIR / "16_cycle_energy_dissipation_deltaE_over_E.png"
    _save_workbench_figure(fig, png, FIGSIZE_FLAT_TRIPTYCH)
    fig_enhanced = Figure(figsize=FIGSIZE_FLAT_TRIPTYCH)
    plot_cycle_energy_dissipation(
        h5_path=H5_PATH,
        fig=fig_enhanced,
        supports=supports,
        support_colors=support_color,
        exclude_amp_anomalies=True,
        bins=16,
        enhanced=True,
    )
    png_enhanced = OUT_DIR / "16_cycle_energy_dissipation_deltaE_over_E_enhanced.png"
    _save_workbench_figure(fig_enhanced, png_enhanced, FIGSIZE_FLAT_TRIPTYCH)
    _remove_legacy_figure_outputs("16_cycle_energy_dissipation_deltaE_over_E")
    _display_workbench_figure(png_enhanced)
    print("[DeltaE/E]", ede_stats)


def plot_damping_nli_figure(ctx: dict) -> None:
    print("[DAMP] Plotting NLI applicability...")
    supports = ctx["supports"]
    support_color = ctx["support_color"]
    fig = Figure(figsize=FIGSIZE_FLAT_SINGLE)
    nli_stats = plot_nli_applicability(
        h5_path=H5_PATH,
        fig=fig,
        bandwidth_pct=BANDWIDTH_PCT,
        butter_order=BUTTER_ORDER,
        support_colors=support_color,
        supports=supports,
        min_freq_ratio=0.2,
        min_abs_freq_hz=0.1,
        amplitude_floor_ratio=0.05,
        nli_hard_cap=100.0,
        window_cycles=3.0,
        exclude_amp_anomalies=True,
    )
    png = OUT_DIR / "17_nonlinear_indicator_applicability_nli.png"
    _save_workbench_figure(fig, png, FIGSIZE_FLAT_SINGLE)
    _remove_legacy_figure_outputs("17_nonlinear_indicator_applicability_nli")
    _display_workbench_figure(png)
    print("[NLI]", nli_stats)


def plot_damping_eta_eq_frozen(ctx: dict) -> None:
    # ⛔ FROZEN — parked for appendix review. Red background = do not use in paper.
    print("[DAMP][FROZEN] Plotting eta_eq inverse...")
    eta_clean, supports, support_color = _compute_eta_clean(ctx)
    fig = Figure(figsize=FIGSIZE_WIDE)
    fig.patch.set_facecolor('#FFCCCC')
    axm = fig.subplots(1, 1)
    for ax_item in (axm if hasattr(axm, '__iter__') else [axm]):
        if hasattr(ax_item, '__iter__'):
            for a in ax_item:
                a.set_facecolor('#FFE0E0')
        else:
            ax_item.set_facecolor('#FFE0E0')
    data = []
    labels = []
    for sup in supports:
        vals = [x["eta_eq"] for x in eta_clean if x["support_type"] == sup and np.isfinite(x["eta_eq"])]
        if not vals:
            continue
        data.append(vals)
        labels.append(sup)
    if data:
        bp = axm.boxplot(data, tick_labels=labels, patch_artist=True, showfliers=False)
        for patch, sup in zip(bp["boxes"], labels):
            patch.set_facecolor(support_color[sup]); patch.set_alpha(0.25)
        for i, sup in enumerate(labels, start=1):
            vals = np.asarray(data[i - 1], dtype=np.float64)
            xj = np.linspace(i - 0.12, i + 0.12, len(vals))
            axm.scatter(xj, vals, s=12, alpha=0.55, color=support_color[sup])
    axm.axhspan(0.1, 0.6, color="#adb5bd", alpha=0.30, zorder=0, label="steel friction reference band")
    axm.set_title("Equivalent Dissipation Parameter Inversion (eta_eq)")
    axm.set_ylabel("eta_eq")
    axm.grid(False)
    axm.legend(frameon=False, fontsize=8, loc="upper right")
    png = OUT_DIR / "A1_equivalent_dissipation_parameter_inversion_eta_eq.png"
    _save_workbench_figure(fig, png, FIGSIZE_WIDE)
    _display_workbench_figure(png)


def plot_damping_backbone_frozen(ctx: dict) -> None:
    # ⛔ FROZEN — parked for appendix review. Red background = do not use in paper.
    print("[DAMP][FROZEN] Plotting backbone examples...")
    supports = ctx["supports"]
    support_color = ctx["support_color"]
    fig = Figure(figsize=FIGSIZE_TALL)
    fig.patch.set_facecolor('#FFCCCC')
    axs = fig.subplots(2, 2)
    for ax_item in (axs if hasattr(axs, '__iter__') else [axs]):
        if hasattr(ax_item, '__iter__'):
            for a in ax_item:
                a.set_facecolor('#FFE0E0')
        else:
            ax_item.set_facecolor('#FFE0E0')
    backbone_notes = plot_backbone_examples(
        h5_path=H5_PATH,
        fig=fig,
        bandwidth_pct=BANDWIDTH_PCT,
        butter_order=BUTTER_ORDER,
        supports=supports,
        support_colors=support_color,
        amp_floor_ratio=0.08,
        freq_range=(0.1, 1.0),
        min_keep_ratio=0.80,
        exclude_amp_anomalies=True,
    )
    png = OUT_DIR / "A2_backbone_examples.png"
    _save_workbench_figure(fig, png, FIGSIZE_TALL)
    _display_workbench_figure(png)
    print("[Backbone]")
    for sup in supports:
        if sup in backbone_notes:
            print(f"  {sup}: {backbone_notes[sup]}")


def plot_decay_fft_contrast_group(ctx: dict) -> None:
    print("[Layer 1] Plotting integrated frequency-reliability triptych...")
    valid_rows = ctx["valid_rows"]
    all_rows = ctx["all_rows"]
    good_row, rapid_row = _choose_decay_fft_cases(valid_rows, all_rows)
    good_payload = _extract_decay_fft_payload(H5_PATH, good_row, BANDWIDTH_PCT, BUTTER_ORDER)
    rapid_payload = _extract_decay_fft_payload(H5_PATH, rapid_row, BANDWIDTH_PCT, BUTTER_ORDER)

    selected_fast_decay_exp_ids = [
        "CA_07_A", "CA_07_R", "CA_13_A", "CA_13_R",
        "TB_07_R", "TB_13_A", "TB_13_R",
        "CSBD_07_A", "CSBD_07_R", "CSBD_13_A", "CSBD_13_R",
        "CSB_07_A", "CSB_07_R", "CSB_13_A", "CSB_13_R",
    ]

    fig = plt.figure(figsize=FIGSIZE_SIGNAL_QUALITY)
    gs = fig.add_gridspec(2, 2, width_ratios=[1, 1.15], height_ratios=[1, 1], wspace=0.22, hspace=0.56)
    ax_t = fig.add_subplot(gs[0, 0])
    ax_f = fig.add_subplot(gs[1, 0])
    ax_panel = fig.add_subplot(gs[:, 1])
    for ax in (ax_t, ax_f, ax_panel):
        ax.set_gid("keep-local-fontsizes")

    def add_panel_label(ax, label):
        text = ax.text(
            0.018,
            0.975,
            f"({label})",
            transform=ax.transAxes,
            ha="left",
            va="top",
            fontsize=8.4,
            fontweight="bold",
            color="#1f2933",
            bbox=dict(facecolor="white", edgecolor="none", alpha=0.88, pad=1.5),
            zorder=30,
        )
        text.set_gid("keep-fontsize")

    good_color = "#7b8794"
    rapid_color = "#2aa6a4"
    accent_dark = "#1f2933"

    ax_t.set_facecolor("#fcfcfd")
    ax_f.set_facecolor("#fcfcfd")

    good_t_plot = good_payload["t_plot"]
    rapid_t_plot = rapid_payload["t_plot"]
    t_max_plot = max(float(good_t_plot[-1]), float(rapid_t_plot[-1]))

    ax_t.plot(good_t_plot, good_payload["sig_norm"], color=good_color, lw=1.15, alpha=0.92, zorder=2)
    ax_t.plot(good_t_plot, good_payload["env_norm"], color=good_color, lw=1.0, ls="--", alpha=0.95, zorder=3)
    ax_t.plot(rapid_t_plot, rapid_payload["sig_norm"], color=rapid_color, lw=1.35, alpha=0.9, zorder=4)
    ax_t.plot(rapid_t_plot, rapid_payload["env_norm"], color=rapid_color, lw=1.05, ls="--", alpha=0.98, zorder=5)
    ax_t.set_title("Waveform contrast", fontsize=9.5, pad=9)
    ax_t.set_xlabel("Time after detected peak (s)", fontsize=9.6, labelpad=4)
    ax_t.set_xlim(0.0, t_max_plot * 1.02)
    ax_t.set_ylabel("Normalized amplitude", fontsize=9.6, labelpad=8)
    ax_t.set_ylim(-1.05, 1.05)
    ax_t.grid(axis="x", color="#d9e2ec", alpha=0.55, linewidth=0.6, linestyle=":")
    ax_t.tick_params(labelsize=8.4)
    from matplotlib.lines import Line2D
    ax_t.legend(
        handles=[
            Line2D([0], [0], color=good_color, lw=1.45, label="Well-decayed"),
            Line2D([0], [0], color=rapid_color, lw=1.45, label="Rapidly decayed"),
        ],
        frameon=True,
        facecolor="white",
        edgecolor="#d7dde5",
        framealpha=0.9,
        fontsize=7.2,
        loc="lower right",
        bbox_to_anchor=(0.99, 0.01),
        ncol=1,
        handlelength=1.1,
        handletextpad=0.35,
        labelspacing=0.18,
        borderaxespad=0.0,
    )
    inset = ax_t.inset_axes([0.70, 0.60, 0.28, 0.30])
    inset.set_gid("keep-local-fontsizes")
    rapid_t = rapid_payload["t_plot"]
    rapid_sig = rapid_payload["sig_norm"]
    rapid_env = rapid_payload["env_norm"]
    good_t = good_payload["t_plot"]
    good_sig = good_payload["sig_norm"]
    good_env = good_payload["env_norm"]
    inset_end = min(float(max(rapid_t[-1], good_t[-1])), max(2.8, float(rapid_payload["t_seg"][-1]) * 0.55))
    rapid_mask = (rapid_t >= 0.0) & (rapid_t <= inset_end)
    good_mask = (good_t >= 0.0) & (good_t <= inset_end)
    inset.set_facecolor("white")
    inset.plot(good_t[good_mask], good_sig[good_mask], color=good_color, lw=1.0, alpha=0.9)
    inset.plot(good_t[good_mask], good_env[good_mask], color=good_color, lw=0.85, ls="--", alpha=0.95)
    inset.plot(rapid_t[rapid_mask], rapid_sig[rapid_mask], color=rapid_color, lw=1.15, alpha=0.95)
    inset.plot(rapid_t[rapid_mask], rapid_env[rapid_mask], color=rapid_color, lw=0.95, ls="--", alpha=0.95)
    inset.set_title("Zoom", fontsize=6.2, pad=1.2)
    inset.set_xlim(0.0, inset_end)
    inset.set_ylim(-1.02, 1.02)
    inset.tick_params(labelsize=5.8, direction="in", length=1.8, pad=1.0)
    inset.grid(axis="x", color="#e2e8f0", alpha=0.55, linewidth=0.45, linestyle=":")
    for spine in inset.spines.values():
        spine.set_linewidth(0.8)
        spine.set_edgecolor("black")

    from matplotlib.patches import Rectangle, ConnectionPatch
    zoom_rect = Rectangle((0.0, -1.0), inset_end, 2.0, linewidth=0.9, edgecolor="black", facecolor="none", linestyle="-", alpha=0.95, zorder=6)
    ax_t.add_patch(zoom_rect)
    con_top = ConnectionPatch(xyA=(inset_end, 1.0), coordsA=ax_t.transData, xyB=(0.0, 1.0), coordsB=inset.transAxes, color="black", lw=0.75, alpha=0.9)
    con_bottom = ConnectionPatch(xyA=(inset_end, -1.0), coordsA=ax_t.transData, xyB=(0.0, 0.0), coordsB=inset.transAxes, color="black", lw=0.75, alpha=0.9)
    fig.add_artist(con_top)
    fig.add_artist(con_bottom)

    ax_f.plot(good_payload["freqs"], good_payload["spec_norm"], color=good_color, lw=1.45, alpha=0.95)
    ax_f.plot(rapid_payload["freqs"], rapid_payload["spec_norm"], color=rapid_color, lw=1.5, alpha=0.9)
    ax_f.axvline(good_payload["freq_fft_peak"], color=good_color, lw=1.0, ls="--", alpha=0.85)
    ax_f.axvline(rapid_payload["freq_fft_peak"], color=rapid_color, lw=1.0, ls="--", alpha=0.85)
    if np.isfinite(rapid_payload["freq_peakcount"]):
        ax_f.axvline(rapid_payload["freq_peakcount"], color="#4a5568", lw=1.0, ls=":", alpha=0.95)
    x_hi = max(1.6, float(np.nanmax(good_payload["freqs"])), float(np.nanmax(rapid_payload["freqs"])), good_payload["freq_auto"] * 1.6, rapid_payload["freq_auto"] * 1.6)
    ax_f.set_xlim(0.05, x_hi)
    ax_f.set_title("FFT contrast", fontsize=9.5, pad=9)
    ax_f.set_xlabel("Frequency (Hz)", fontsize=9.6, labelpad=7)
    ax_f.set_ylabel("Normalized FFT amplitude", fontsize=9.6, labelpad=8)
    ax_f.grid(axis="x", color="#d9e2ec", alpha=0.55, linewidth=0.6, linestyle=":")
    ax_f.tick_params(labelsize=8.4)
    good_peak_label = ax_f.text(
        good_payload["freq_fft_peak"] + 0.03,
        0.92,
        f"{good_payload['freq_fft_peak']:.2f} Hz",
        color=good_color,
        fontsize=7.0,
        ha="left",
        va="center",
    )
    good_peak_label.set_gid("keep-fontsize")
    rapid_peak_label = ax_f.text(
        rapid_payload["freq_fft_peak"] + 0.03,
        0.92,
        f"{rapid_payload['freq_fft_peak']:.2f} Hz",
        color=rapid_color,
        fontsize=7.0,
        ha="left",
        va="center",
    )
    rapid_peak_label.set_gid("keep-fontsize")
    selected_decay_report = plot_selected_fast_decay_frequency_check(
        h5_path=H5_PATH,
        fig=fig,
        ax=ax_panel,
        exp_ids=selected_fast_decay_exp_ids,
        bandwidth_pct=BANDWIDTH_PCT,
        butter_order=BUTTER_ORDER,
        show_case_legend=False,
    )
    ax_panel.set_title("Fast-decay frequency check", fontsize=9.5, pad=9)
    ax_panel.xaxis.label.set_size(9.6)
    ax_panel.yaxis.label.set_size(9.6)
    ax_panel.xaxis.labelpad = 7
    ax_panel.yaxis.labelpad = 8
    ax_panel.tick_params(labelsize=8.4)
    add_panel_label(ax_t, "a")
    add_panel_label(ax_f, "b")
    add_panel_label(ax_panel, "c")

    png_cmp = OUT_DIR / "09_decay_waveform_and_fft_contrast.png"
    # `tight_layout` is not reliable here because the figure includes an inset axes
    # and connection patches. Use explicit margins to keep the geometry stable.
    fig.subplots_adjust(left=0.07, right=0.985, bottom=0.12, top=0.93)
    _save_workbench_figure(fig, png_cmp, FIGSIZE_SIGNAL_QUALITY)
    _remove_legacy_figure_outputs("09_decay_waveform_and_fft_contrast")
    _display_workbench_figure(png_cmp)
    plt.close(fig)
    print("[Layer 1 Frequency Reliability]")
    print(
        f"  good : {good_payload['exp_id']} | {good_payload['channel_id']} | {good_payload['mode']} | "
        f"cycles={good_payload['decay_cycles']:.2f} | auto={good_payload['freq_auto']:.3f} | FFT={good_payload['freq_fft_peak']:.3f}"
    )
    print(
        f"  rapid: {rapid_payload['exp_id']} | {rapid_payload['channel_id']} | {rapid_payload['mode']} | "
        f"cycles={rapid_payload['decay_cycles']:.2f} | auto={rapid_payload['freq_auto']:.3f} | FFT={rapid_payload['freq_fft_peak']:.3f}"
    )
    print("[Selected Fast-Decay Frequency Check]")
    for row in selected_decay_report:
        auto_txt = f"{row['freq_auto_hz']:.3f}"
        peak_txt = f"{row['freq_peakcount_hz']:.3f}" if np.isfinite(row.get("freq_peakcount_hz", np.nan)) else "NA"
        diff_txt = f"{row['rel_diff_pct']:.1f}%" if np.isfinite(row.get("rel_diff_pct", np.nan)) else "NA"
        print(
            f"  {row['exp_id']} | {row['channel_id']} | {row['mode']} | "
            f"auto={auto_txt} Hz | peak={peak_txt} Hz | diff={diff_txt} | "
            f"peaks={row['peak_count']} | cycles={row['decay_cycles']:.2f} | {row['reject_reason']}"
        )







def plot_damping_correlation_group(ctx: dict) -> None:
    print("[4/4] Generating damping vs normalized frequency-amplitude slope plots...")
    valid_rows = ctx["valid_rows"]
    mode_rows = {m: [r for r in valid_rows if str(r.get("mode", "")) == m] for m in MODE_ORDER}

    fig_corr, _ = plot_damping_drift_correlation(
        mode_rows,
        out_path=None,
        dpi=600,
    )
    png = OUT_DIR / "14_equivalent_damping_ratio_vs_normalized_frequency_amplitude_sensitivity_beta.png"
    _save_workbench_figure(fig_corr, png, FIGSIZE_FLAT_CORRELATION)
    fig_corr_enhanced, _ = plot_damping_drift_correlation(
        mode_rows,
        out_path=None,
        dpi=600,
        enhanced=True,
    )
    png_enhanced = OUT_DIR / "14_equivalent_damping_ratio_vs_normalized_frequency_amplitude_sensitivity_beta_enhanced.png"
    _save_workbench_figure(fig_corr_enhanced, png_enhanced, FIGSIZE_FLAT_CORRELATION)
    _remove_legacy_figure_outputs("14_equivalent_damping_ratio_vs_normalized_frequency_amplitude_sensitivity_beta")
    _display_workbench_figure(png_enhanced)
    plt.close(fig_corr)
    plt.close(fig_corr_enhanced)


def run_all_damping():
    ctx = prepare_damping_context(run_batch=RUN_DAMPING_BATCH)
    plot_decay_fft_contrast_group(ctx)
    plot_damping_summary_figure(ctx)
    plot_damping_statistics_refresh(ctx)
    plot_damping_freq_drift_figure(ctx)
    plot_damping_correlation_group(ctx)
    plot_damping_xi_amplitude_figure(ctx)
    plot_damping_deltaE_figure(ctx)
    plot_damping_nli_figure(ctx)
    plot_damping_eta_eq_frozen(ctx)
    plot_damping_backbone_frozen(ctx)
    print("[DONE] Damping workflow finished.")

