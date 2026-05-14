"""DimScaling frequency fitting and comparison plotting (H5 + COMPARE_6)."""

from __future__ import annotations

import argparse
import csv
import math
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Tuple

import h5py
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
from scipy.optimize import curve_fit

from core.modal_db import DEFAULT_H5_PATH, DEFAULT_METADATA_DIR, DEFAULT_PROCESSED_DIR

G_ACCEL = 9.81
MODE_ORDER = ["radial", "axial", "planar_rotation"]
MODE_ALIAS = {
    "径向平动": "radial",
    "轴向平动": "axial",
    "平面转动": "planar_rotation",
}
MODE_CN = {
    "radial": "Radial Translation",
    "axial": "Axial Translation",
    "planar_rotation": "In-plane rotation",
}
MODE_COLOR = {
    "radial": "#d00000",
    "axial": "#3a86ff",
    "planar_rotation": "#2a9d8f",
}
PIPE_FIXED_MASS_KG = 388.0
PIPE_YAW_INERTIA_BASE = 259.7
HANGER_ECCENTRICITY_M = 0.9
COMPONENT_ORDER_RAW = [
    "CA_07",
    "CA_13",
    "CA_07_Top",
    "CA_13_Top",
    "CA_07+13_Mid",
    "CA_13+07_Mid",
    "TB_07",
    "TB_13",
    "TB_07_Top",
    "TB_13_Top",
    "TB_07+07_Mid",
    "TB_07+13_Mid",
    "TB_13+07_Mid",
    "TB_07+07_MidTop",
    "TB_07+13_MidTop",
    "TB_13+07_MidTop",
    "TR_07",
    "TR_13",
    "CS_07",
    "CS_13",
]
DIRECTION_RANK = {"A": 0, "R": 1, "AA": 2, "AR": 3, "RA": 4, "RR": 5}


@dataclass
class FitStats:
    mode: str
    fit_group: str
    n_total: int
    n_used: int
    outlier_policy: str
    A: float
    B: float
    r2: float
    rmse: float
    mae: float


def _to_float(v) -> float | None:
    if v is None:
        return None
    try:
        fv = float(v)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(fv):
        return None
    return fv


def _compute_planar_rotation_inertia(mass_total: float) -> float:
    return PIPE_YAW_INERTIA_BASE + (mass_total - PIPE_FIXED_MASS_KG) * (HANGER_ECCENTRICITY_M ** 2)


def _calc_mode_f_pend(mode: str, mass_total: float, l_val: float) -> float:
    if not (np.isfinite(mass_total) and np.isfinite(l_val) and mass_total > 0 and l_val > 0):
        return np.nan
    return float((1.0 / (2.0 * np.pi)) * math.sqrt(G_ACCEL / l_val))


def _calc_mode_k_obs(mode: str, mass_total: float, l_val: float, f_exp: float) -> float:
    if not (np.isfinite(mass_total) and np.isfinite(l_val) and np.isfinite(f_exp)):
        return np.nan
    if mass_total <= 0 or l_val <= 0 or f_exp <= 0:
        return np.nan
    if mode == "planar_rotation":
        f_ideal = _calc_mode_f_pend("axial", mass_total, l_val)
        if not np.isfinite(f_ideal):
            return np.nan
        inertia_z = _compute_planar_rotation_inertia(mass_total)
        return float(4.0 * (np.pi ** 2) * inertia_z * (f_exp ** 2 - f_ideal ** 2))
    omega = 2.0 * np.pi * f_exp
    return float(mass_total * (l_val ** 2) * (omega * omega - G_ACCEL / l_val))


def _calc_mode_f_fit(mode: str, mass_total: float, l_val: float, k_fit: float) -> float:
    if not (np.isfinite(mass_total) and np.isfinite(l_val) and np.isfinite(k_fit)):
        return np.nan
    if mass_total <= 0 or l_val <= 0:
        return np.nan
    f_ideal = _calc_mode_f_pend(mode, mass_total, l_val)
    if not np.isfinite(f_ideal):
        return np.nan
    if mode == "planar_rotation":
        inertia_z = _compute_planar_rotation_inertia(mass_total)
        inside = f_ideal ** 2 + k_fit / (4.0 * (np.pi ** 2) * inertia_z)
        return float(math.sqrt(max(inside, 0.0)))
    inside = G_ACCEL / l_val + k_fit / (mass_total * (l_val ** 2))
    return float((1.0 / (2.0 * np.pi)) * math.sqrt(max(inside, 0.0)))


def _normalize_mode(raw_mode: str) -> str:
    text = (raw_mode or "").strip()
    for key, val in MODE_ALIAS.items():
        if key in text:
            return val
    if text in MODE_ORDER:
        return text
    return "unknown"


def _normalize_component_prefix(component: str) -> str:
    c = (component or "").strip()
    if c.startswith("TR_"):
        return "CSBD_" + c[3:]
    if c.startswith("CS_"):
        return "CSB_" + c[3:]
    return c


COMPONENT_ORDER = [_normalize_component_prefix(x) for x in COMPONENT_ORDER_RAW]
COMPONENT_RANK = {name: idx for idx, name in enumerate(COMPONENT_ORDER)}


def _split_exp_id(exp_id: str) -> tuple[str, str]:
    if "_" not in exp_id:
        return exp_id, ""
    comp, direction = exp_id.rsplit("_", 1)
    if direction in DIRECTION_RANK:
        return comp, direction
    return exp_id, ""


def exp_sort_key(exp_id: str) -> tuple[int, int, str]:
    comp, direction = _split_exp_id(exp_id)
    return (
        COMPONENT_RANK.get(comp, 10_000),
        DIRECTION_RANK.get(direction, 10_000),
        exp_id,
    )


def get_fit_group(exp_id: str) -> str:
    if exp_id.startswith("CA_"):
        return "CA"
    return "OTHERS"


def _param_key(exp_id: str, mode: str) -> tuple[str, str]:
    return str(exp_id), str(mode)


def _mean_positive(values: list[float]) -> float:
    arr = np.asarray([float(v) for v in values if v is not None and float(v) > 0], dtype=np.float64)
    if arr.size == 0:
        return np.nan
    return float(np.mean(arr))


def _select_params_for_mode(params_mean: Dict[tuple[str, str], dict], exp_id: str, mode: str) -> dict:
    default = {"m": np.nan, "l": np.nan, "r": np.nan}
    direct = params_mean.get(_param_key(exp_id, mode))
    axial = params_mean.get(_param_key(exp_id, "axial"))
    if mode == "planar_rotation" and axial is not None:
        return dict(axial)
    if direct is not None:
        return dict(direct)
    if axial is not None:
        return dict(axial)
    return dict(default)


def load_h5_modal_freq_means(h5_path: Path) -> Dict[Tuple[str, str], float]:
    result: Dict[Tuple[str, str], float] = {}
    bucket: Dict[Tuple[str, str], List[float]] = {}
    with h5py.File(h5_path, "r") as h5:
        exps = h5["experiments"]
        for exp_id in exps.keys():
            g = exps[exp_id]
            if "modal_freq_modes" not in g or "modal_freq_values" not in g:
                continue
            modes_raw = g["modal_freq_modes"][:]
            vals_raw = g["modal_freq_values"][:]
            for m_raw, f_raw in zip(modes_raw, vals_raw):
                mode = _normalize_mode(m_raw.decode("utf-8", errors="ignore") if isinstance(m_raw, bytes) else str(m_raw))
                if mode not in MODE_ORDER:
                    continue
                fv = _to_float(f_raw)
                if fv is None or fv <= 0:
                    continue
                bucket.setdefault((exp_id, mode), []).append(fv)
    for key, arr in bucket.items():
        result[key] = float(np.mean(np.asarray(arr, dtype=np.float64)))
    return result


def load_compare6_records(compare6_path: Path) -> tuple[list[dict], Dict[Tuple[str, str], float]]:
    wb = openpyxl.load_workbook(compare6_path, data_only=True)
    records: list[dict] = []
    freq_bucket: Dict[Tuple[str, str], List[float]] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        last_component = ""
        last_direction = ""
        last_m = None
        last_l = None
        last_r = None

        for r in range(2, ws.max_row + 1):
            c_component = ws.cell(r, 2).value
            c_direction = ws.cell(r, 3).value
            c_mode = ws.cell(r, 4).value
            c_freq = ws.cell(r, 5).value
            c_m = ws.cell(r, 13).value  # M
            c_l = ws.cell(r, 14).value  # N
            c_r = ws.cell(r, 15).value  # O

            if isinstance(c_component, str) and c_component.strip():
                last_component = c_component.strip()
            component = _normalize_component_prefix(last_component)
            if not component:
                continue

            if isinstance(c_direction, str) and c_direction.strip():
                last_direction = c_direction.strip().upper()
            direction = last_direction
            if not direction:
                continue

            m_val = _to_float(c_m)
            l_val = _to_float(c_l)
            r_val = _to_float(c_r)
            if m_val is not None:
                last_m = m_val
            if l_val is not None:
                last_l = l_val
            if r_val is not None:
                last_r = r_val

            mode = _normalize_mode(str(c_mode) if c_mode is not None else "")
            freq_val = _to_float(c_freq)
            exp_id = f"{component}_{direction}"

            rec = {
                "exp_id": exp_id,
                "component": component,
                "direction": direction,
                "mode": mode,
                "f_excel": freq_val,
                "m": last_m,
                "l": last_l,
                "r": last_r,
            }
            records.append(rec)

            if mode in MODE_ORDER and freq_val is not None and freq_val > 0:
                freq_bucket.setdefault((exp_id, mode), []).append(freq_val)

    freq_means = {k: float(np.mean(np.asarray(v, dtype=np.float64))) for k, v in freq_bucket.items()}
    return records, freq_means


def build_analysis_rows(
    compare_records: list[dict],
    compare_freq_means: Dict[Tuple[str, str], float],
    h5_freq_means: Dict[Tuple[str, str], float],
) -> list[dict]:
    param_by_mode: Dict[tuple[str, str], dict] = {}
    for rec in compare_records:
        mode = rec["mode"]
        if mode not in MODE_ORDER:
            continue
        key = _param_key(rec["exp_id"], mode)
        if key not in param_by_mode:
            param_by_mode[key] = {"m_list": [], "l_list": [], "r_list": []}
        for k in ["m", "l", "r"]:
            val = rec[k]
            if val is not None and val > 0:
                param_by_mode[key][f"{k}_list"].append(float(val))

    params_mean: Dict[tuple[str, str], dict] = {}
    for key, p in param_by_mode.items():
        params_mean[key] = {
            "m": _mean_positive(p["m_list"]),
            "l": _mean_positive(p["l_list"]),
            "r": _mean_positive(p["r_list"]),
        }

    all_keys = sorted(set(compare_freq_means.keys()) | set(h5_freq_means.keys()))
    rows: list[dict] = []
    for exp_id, mode in all_keys:
        p = _select_params_for_mode(params_mean, exp_id, mode)
        f_excel = compare_freq_means.get((exp_id, mode), np.nan)
        f_h5 = h5_freq_means.get((exp_id, mode), np.nan)
        rows.append(
            {
                "exp_id": exp_id,
                "mode": mode,
                "m": p["m"],
                "l": p["l"],
                "r": p["r"],
                "f_exp_excel": f_excel,
                "f_exp_h5": f_h5,
                "f_exp": f_excel,  # locked policy: COMPARE_6 as primary
            }
        )
    return rows


def _iqr_mask(values: np.ndarray) -> np.ndarray:
    q1 = np.quantile(values, 0.25)
    q3 = np.quantile(values, 0.75)
    iqr = q3 - q1
    lo = q1 - 1.5 * iqr
    hi = q3 + 1.5 * iqr
    return (values >= lo) & (values <= hi)


def fit_mode_dimscaling(rows_mode: list[dict], outlier: str, residual_thresh: float, fit_group: str = "ALL") -> tuple[FitStats, list[dict]]:
    # Keep only rows with valid params + primary frequency
    valid_rows = []
    for r in rows_mode:
        if not (np.isfinite(r["m"]) and np.isfinite(r["l"]) and np.isfinite(r["r"]) and np.isfinite(r["f_exp"])):
            continue
        if r["m"] <= 0 or r["l"] <= 0 or r["r"] <= 0 or r["f_exp"] <= 0:
            continue
        valid_rows.append(dict(r))

    if not valid_rows:
        stats = FitStats(mode=rows_mode[0]["mode"], fit_group=fit_group, n_total=len(rows_mode), n_used=0, outlier_policy=outlier, A=np.nan, B=np.nan, r2=np.nan, rmse=np.nan, mae=np.nan)
        return stats, rows_mode

    for r in valid_rows:
        r["k_obs"] = _calc_mode_k_obs(r["mode"], r["m"], r["l"], r["f_exp"])
        r["pi_k"] = r["k_obs"] / (r["m"] * G_ACCEL * r["r"]) if r["m"] > 0 and r["r"] > 0 else np.nan
        r["pi_l"] = r["l"] / r["r"]

    fit_rows = [r for r in valid_rows if np.isfinite(r["pi_k"]) and np.isfinite(r["pi_l"]) and r["pi_l"] > 0]
    if not fit_rows:
        stats = FitStats(mode=rows_mode[0]["mode"], fit_group=fit_group, n_total=len(rows_mode), n_used=0, outlier_policy=outlier, A=np.nan, B=np.nan, r2=np.nan, rmse=np.nan, mae=np.nan)
        return stats, rows_mode

    x_raw = np.asarray([r["pi_l"] for r in fit_rows], dtype=np.float64)
    y_raw = np.asarray([r["pi_k"] for r in fit_rows], dtype=np.float64)
    mask = np.ones_like(x_raw, dtype=bool)

    if outlier == "iqr" and len(y_raw) >= 4:
        mask = _iqr_mask(y_raw)
    elif outlier == "residual" and len(y_raw) >= 4:
        def _model(x, A, B):
            return A * np.power(x, B)

        A0 = float(np.nanmedian(y_raw)) if len(y_raw) else 1.0
        if not np.isfinite(A0) or A0 == 0:
            A0 = 1.0
        B0 = 0.5
        try:
            popt0, _ = curve_fit(_model, x_raw, y_raw, p0=[A0, B0], maxfev=30000)
            A0, B0 = float(popt0[0]), float(popt0[1])
        except Exception:
            pass
        rels = []
        for r in fit_rows:
            k_fit_0 = A0 * r["m"] * G_ACCEL * r["r"] * ((r["l"] / r["r"]) ** B0)
            f_fit_0 = _calc_mode_f_fit(r["mode"], r["m"], r["l"], k_fit_0)
            rel = abs(f_fit_0 - r["f_exp"]) / max(r["f_exp"], 1e-12)
            rels.append(rel)
        rel_arr = np.asarray(rels, dtype=np.float64)
        mask = rel_arr <= residual_thresh
        if np.count_nonzero(mask) < 2:
            mask = np.ones_like(mask, dtype=bool)

    used_fit_rows = [fit_rows[i] for i, keep in enumerate(mask) if keep]
    x = x_raw[mask]
    y = y_raw[mask]
    if len(x) < 2:
        used_fit_rows = list(fit_rows)
        x = x_raw
        y = y_raw

    def _model(xv, A, B):
        return A * np.power(xv, B)

    A_init = float(np.nanmedian(y)) if len(y) else 1.0
    if not np.isfinite(A_init) or A_init == 0:
        A_init = 1.0
    B_init = 0.5
    try:
        popt, _ = curve_fit(_model, x, y, p0=[A_init, B_init], maxfev=50000)
        A = float(popt[0])
        B = float(popt[1])
    except Exception:
        # Fallback: linear least squares on fixed B=0.5
        B = 0.5
        xb = np.power(x, B)
        denom = float(np.dot(xb, xb))
        A = float(np.dot(xb, y) / denom) if denom > 0 else float(A_init)

    used_keys = {(r["exp_id"], r["mode"]) for r in used_fit_rows}

    # Predict for all valid rows with final A/B
    for r in valid_rows:
        r["fit_used"] = (r["exp_id"], r["mode"]) in used_keys
        k_fit = A * r["m"] * G_ACCEL * r["r"] * ((r["l"] / r["r"]) ** B)
        r["k_fit"] = k_fit
        f_pend = _calc_mode_f_pend(r["mode"], r["m"], r["l"])
        f_fit = _calc_mode_f_fit(r["mode"], r["m"], r["l"], k_fit)
        r["f_pend"] = f_pend
        r["f_fit"] = f_fit
        r["err_abs"] = abs(f_fit - r["f_exp"])
        r["err_rel"] = abs(f_fit - r["f_exp"]) / max(r["f_exp"], 1e-12)

    y_true = np.asarray([r["f_exp"] for r in valid_rows], dtype=np.float64)
    y_pred = np.asarray([r["f_fit"] for r in valid_rows], dtype=np.float64)
    ss_res = float(np.sum((y_true - y_pred) ** 2))
    ss_tot = float(np.sum((y_true - np.mean(y_true)) ** 2))
    r2 = 1.0 - ss_res / ss_tot if ss_tot > 0 else np.nan
    rmse = float(np.sqrt(np.mean((y_true - y_pred) ** 2)))
    mae = float(np.mean(np.abs(y_true - y_pred)))

    stats = FitStats(
        mode=rows_mode[0]["mode"],
        fit_group=fit_group,
        n_total=len(rows_mode),
        n_used=len(used_fit_rows),
        outlier_policy=outlier,
        A=A,
        B=B,
        r2=r2,
        rmse=rmse,
        mae=mae,
    )

    valid_map = {(r["exp_id"], r["mode"]): r for r in valid_rows}
    out_rows = []
    for r in rows_mode:
        nr = dict(r)
        nr["f_pend"] = _calc_mode_f_pend(nr["mode"], nr.get("m", np.nan), nr.get("l", np.nan))
        vr = valid_map.get((r["exp_id"], r["mode"]))
        if vr is None:
            nr["f_fit"] = np.nan
            nr["err_abs"] = np.nan
            nr["err_rel"] = np.nan
            nr["fit_used"] = False
            nr["A"] = A
            nr["B"] = B
        else:
            nr["f_pend"] = vr["f_pend"]
            nr["f_fit"] = vr["f_fit"]
            nr["err_abs"] = vr["err_abs"]
            nr["err_rel"] = vr["err_rel"]
            nr["fit_used"] = bool(vr.get("fit_used", False))
            nr["A"] = A
            nr["B"] = B
        out_rows.append(nr)
    return stats, out_rows


def plot_compare(rows_mode: list[dict], stats: FitStats, out_dir: Path) -> None:
    mode = rows_mode[0]["mode"]
    rows_display = [r for r in rows_mode if np.isfinite(r.get("f_exp", np.nan)) and np.isfinite(r.get("f_pend", np.nan))]
    if not rows_display:
        return
    rows_display = sorted(rows_display, key=lambda x: exp_sort_key(x["exp_id"]))
    x = np.arange(len(rows_display))
    labels = [r["exp_id"] for r in rows_display]
    f_exp = np.asarray([r["f_exp"] for r in rows_display], dtype=np.float64)
    f_pend = np.asarray([r["f_pend"] for r in rows_display], dtype=np.float64)
    f_fit = np.asarray([r.get("f_fit", np.nan) for r in rows_display], dtype=np.float64)
    fit_ok = np.isfinite(f_fit)

    fig, ax = plt.subplots(figsize=(14, 6))
    ax.plot(x, f_exp, "o-", lw=1.2, ms=4, color="#1f77b4", label="Experimental Frequency (COMPARE_6)")
    ax.plot(x, f_pend, "s--", lw=1.0, ms=3.5, color="#6c757d", label="Pendulum Frequency")
    ax.plot(x, f_fit, "^-", lw=1.2, ms=4, color=MODE_COLOR.get(mode, "#d00000"), label="DimScaling Fitted Frequency")
    if np.count_nonzero(~fit_ok) > 0:
        ax.scatter(
            x[~fit_ok],
            f_exp[~fit_ok],
            marker="x",
            s=28,
            color="#ee9b00",
            label="Not used in fit",
            zorder=6,
        )
    ax.set_title(f"{MODE_CN.get(mode, mode)} Frequency Comparison (n={len(rows_display)}, fit={int(np.count_nonzero(fit_ok))})", fontsize=12)
    ax.set_ylabel("Frequency (Hz)", fontsize=10)
    ax.set_xlabel("Experiment", fontsize=10)
    ax.set_xticks(x)
    ax.set_xticklabels(labels, rotation=80, ha="right", fontsize=7)
    ax.tick_params(labelsize=8)
    ax.grid(True, alpha=0.2)
    ax.legend(fontsize=9)
    text = f"A={stats.A:.5g}, B={stats.B:.5g}, R2={stats.r2:.4f}, RMSE={stats.rmse:.4f}, MAE={stats.mae:.4f}"
    ax.text(0.01, 0.98, text, transform=ax.transAxes, ha="left", va="top", fontsize=8, bbox={"facecolor": "white", "alpha": 0.75, "edgecolor": "none", "pad": 2})
    fig.subplots_adjust(left=0.06, right=0.995, top=0.93, bottom=0.30)

    png = out_dir / f"dimscaling_freq_compare_{mode}.png"
    pdf = out_dir / f"dimscaling_freq_compare_{mode}.pdf"
    fig.savefig(png, dpi=320)
    fig.savefig(pdf, dpi=320)
    plt.close(fig)


def plot_error(rows_mode: list[dict], stats: FitStats, out_dir: Path) -> None:
    mode = rows_mode[0]["mode"]
    rows_valid = [r for r in rows_mode if np.isfinite(r.get("err_rel", np.nan))]
    if not rows_valid:
        return
    rows_valid = sorted(rows_valid, key=lambda x: exp_sort_key(x["exp_id"]))
    x = np.arange(len(rows_valid))
    labels = [r["exp_id"] for r in rows_valid]
    err_rel_pct = 100.0 * np.asarray([r["err_rel"] for r in rows_valid], dtype=np.float64)
    err_abs = np.asarray([r["err_abs"] for r in rows_valid], dtype=np.float64)

    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(14, 8), sharex=True)
    ax1.bar(x, err_rel_pct, color=MODE_COLOR.get(mode, "#d00000"), alpha=0.7)
    ax1.set_ylabel("Relative Error (%)", fontsize=10)
    ax1.set_title(f"{MODE_CN.get(mode, mode)} Fit Error", fontsize=12)
    ax1.grid(True, alpha=0.2)
    ax1.tick_params(labelsize=8)

    ax2.plot(x, err_abs, "o-", lw=1.1, ms=3.5, color="#495057")
    ax2.set_ylabel("Absolute Error (Hz)", fontsize=10)
    ax2.set_xlabel("Experiment", fontsize=10)
    ax2.grid(True, alpha=0.2)
    ax2.tick_params(labelsize=8)
    ax2.set_xticks(x)
    ax2.set_xticklabels(labels, rotation=80, ha="right", fontsize=7)
    text = f"RMSE={stats.rmse:.4f}, MAE={stats.mae:.4f}"
    ax2.text(0.01, 0.95, text, transform=ax2.transAxes, ha="left", va="top", fontsize=8, bbox={"facecolor": "white", "alpha": 0.75, "edgecolor": "none", "pad": 2})
    fig.subplots_adjust(left=0.07, right=0.995, top=0.92, bottom=0.30, hspace=0.15)

    png = out_dir / f"dimscaling_freq_error_{mode}.png"
    pdf = out_dir / f"dimscaling_freq_error_{mode}.pdf"
    fig.savefig(png, dpi=320)
    fig.savefig(pdf, dpi=320)
    plt.close(fig)


def plot_dual_axis_compare(rows_mode: list[dict], stats: FitStats, out_dir: Path) -> None:
    mode = rows_mode[0]["mode"]
    d = [r for r in rows_mode if np.isfinite(r.get("f_exp", np.nan)) and np.isfinite(r.get("f_pend", np.nan))]
    if not d:
        return
    d = sorted(d, key=lambda z: exp_sort_key(z["exp_id"]))
    x = np.arange(len(d))
    labels = [r["exp_id"] for r in d]

    f_exp = np.asarray([r["f_exp"] for r in d], dtype=np.float64)
    f_pend = np.asarray([r["f_pend"] for r in d], dtype=np.float64)
    f_fit = np.asarray([r.get("f_fit", np.nan) for r in d], dtype=np.float64)
    err_rel_pct = 100.0 * np.asarray([r.get("err_rel", np.nan) for r in d], dtype=np.float64)
    fit_mask = np.isfinite(f_fit)
    err_mask = np.isfinite(err_rel_pct)

    fig, ax1 = plt.subplots(figsize=(14, 6))

    # Left axis: all frequencies as scatter-only (no connecting lines)
    l1 = ax1.scatter(x, f_exp, s=16, marker="o", color="#1f77b4", alpha=0.9, label="Exp Frequency")
    l2 = ax1.scatter(x, f_pend, s=16, marker="s", color="#6c757d", alpha=0.85, label="Pendulum Freq")
    l3 = ax1.scatter(
        x[fit_mask],
        f_fit[fit_mask],
        s=20,
        marker="^",
        color=MODE_COLOR.get(mode, "#d00000"),
        alpha=0.9,
        label="Fitted Freq",
    )
    # Dumbbell stems: connect the local min/max among pendulum/fitted/experimental frequencies.
    y_stack = np.vstack([f_exp, f_pend, f_fit])
    y_min = np.nanmin(y_stack, axis=0)
    y_max = np.nanmax(y_stack, axis=0)
    ax1.vlines(x, ymin=y_min, ymax=y_max, color="gray", alpha=0.3, lw=1.5, zorder=0)
    if np.count_nonzero(~fit_mask) > 0:
        ax1.scatter(
            x[~fit_mask],
            f_exp[~fit_mask],
            marker="x",
            s=28,
            color="#ee9b00",
            label="Not used in fit",
            zorder=6,
        )

    ax1.set_ylabel("Frequency (Hz)", fontweight="bold")
    ax1.set_xlabel("Experiment", fontweight="bold")
    ax1.set_ylim(0.2, 0.9)
    ax1.set_xticks(x)
    ax1.set_xticklabels(labels, rotation=80, ha="right", fontsize=7)
    ax1.grid(True, alpha=0.2)

    # Right axis: lollipop error
    ax2 = ax1.twinx()
    error_color = "#e07a5f"
    ax2.vlines(x[err_mask], ymin=0, ymax=err_rel_pct[err_mask], color=error_color, alpha=0.5, lw=1.2)
    l4 = ax2.scatter(
        x[err_mask],
        err_rel_pct[err_mask],
        s=12,
        marker="o",
        color=error_color,
        alpha=0.85,
        label="Relative Error (%)",
    )
    ax2.set_ylabel("Relative Error (%)", color=error_color, fontweight="bold")
    ax2.tick_params(axis="y", colors=error_color)
    max_err = float(np.nanmax(err_rel_pct[err_mask])) if np.count_nonzero(err_mask) > 0 else 1.0
    if not np.isfinite(max_err) or max_err <= 0:
        max_err = 1.0
    ax2.set_ylim(0, max_err * 3.0)

    # Merged legend (single box)
    handles = [l1, l2, l3, l4]
    labels_legend = [h.get_label() for h in handles]
    ax1.legend(handles, labels_legend, loc="upper right", frameon=False, fontsize=9, ncol=2)

    ax1.set_title(f"{MODE_CN.get(mode, mode)}: Frequency & Error Dual-Axis (n={len(d)})")
    ax1.text(
        0.01,
        0.96,
        f"A={stats.A:.5g}, B={stats.B:.5g}\nR2={stats.r2:.4f}, RMSE={stats.rmse:.4f}\nMAE={stats.mae:.4f}",
        transform=ax1.transAxes,
        ha="left",
        va="top",
        fontsize=8,
        bbox={"facecolor": "white", "alpha": 0.8, "edgecolor": "none", "pad": 2},
    )

    fig.subplots_adjust(left=0.05, right=0.95, top=0.92, bottom=0.25)
    fig.savefig(out_dir / f"dimscaling_dual_{mode}.png", dpi=320)
    fig.savefig(out_dir / f"dimscaling_dual_{mode}.pdf", dpi=320)
    plt.close(fig)


def plot_dual_axis_combined(
    mode_rows: Dict[str, list[dict]],
    mode_stats: Dict[str, Dict[str, FitStats]],
    out_dir: Path,
    figsize: tuple[float, float] = (16, 11),
    save_dpi: int = 320,
    file_prefix: str = "dimscaling",
) -> None:
    present_modes = [m for m in MODE_ORDER if m in mode_rows and m in mode_stats]
    if not present_modes:
        return

    exp_union = sorted(
        {r["exp_id"] for m in present_modes for r in mode_rows[m]},
        key=exp_sort_key,
    )
    if not exp_union:
        return
    exp_to_idx = {eid: i for i, eid in enumerate(exp_union)}
    x = np.arange(len(exp_union), dtype=np.float64)

    fig, axes = plt.subplots(len(present_modes), 1, figsize=figsize, sharex=True)
    if len(present_modes) == 1:
        axes = [axes]

    for i, mode in enumerate(present_modes):
        ax1 = axes[i]
        ax2 = ax1.twinx()
        mode_stats_dict = mode_stats[mode]
        rows = mode_rows[mode]

        f_exp = np.full(len(exp_union), np.nan, dtype=np.float64)
        f_pend = np.full(len(exp_union), np.nan, dtype=np.float64)
        f_fit = np.full(len(exp_union), np.nan, dtype=np.float64)
        err_rel_pct = np.full(len(exp_union), np.nan, dtype=np.float64)

        for r in rows:
            idx = exp_to_idx.get(r["exp_id"])
            if idx is None:
                continue
            f_exp[idx] = float(r["f_exp"]) if np.isfinite(r.get("f_exp", np.nan)) else np.nan
            f_pend[idx] = float(r["f_pend"]) if np.isfinite(r.get("f_pend", np.nan)) else np.nan
            f_fit[idx] = float(r["f_fit"]) if np.isfinite(r.get("f_fit", np.nan)) else np.nan
            err_rel_pct[idx] = 100.0 * float(r["err_rel"]) if np.isfinite(r.get("err_rel", np.nan)) else np.nan

        exp_mask = np.isfinite(f_exp)
        pend_mask = np.isfinite(f_pend)
        fit_mask = np.isfinite(f_fit)
        err_mask = np.isfinite(err_rel_pct)

        l1 = ax1.scatter(x[exp_mask], f_exp[exp_mask], s=14, marker="o", color="#1f77b4", alpha=0.9, label="Exp Frequency")
        l2 = ax1.scatter(x[pend_mask], f_pend[pend_mask], s=14, marker="s", color="#6c757d", alpha=0.85, label="Pendulum Freq")
        l3 = ax1.scatter(
            x[fit_mask],
            f_fit[fit_mask],
            s=18,
            marker="^",
            color=MODE_COLOR.get(mode, "#d00000"),
            alpha=0.9,
            label="Fitted Freq",
        )

        # Dumbbell stems.
        y_stack = np.vstack([f_exp, f_pend, f_fit])
        y_min = np.nanmin(y_stack, axis=0)
        y_max = np.nanmax(y_stack, axis=0)
        stem_mask = np.isfinite(y_min) & np.isfinite(y_max)
        ax1.vlines(x[stem_mask], ymin=y_min[stem_mask], ymax=y_max[stem_mask], color="gray", alpha=0.3, lw=1.2, zorder=0)

        if np.count_nonzero(exp_mask & ~fit_mask) > 0:
            ax1.scatter(
                x[exp_mask & ~fit_mask],
                f_exp[exp_mask & ~fit_mask],
                marker="x",
                s=24,
                color="#ee9b00",
                label="Not used in fit",
                zorder=6,
            )

        ax1.set_ylim(0.2, 0.9)
        ax1.set_ylabel("Frequency (Hz)", fontweight="bold")
        ax1.grid(True, alpha=0.18, linewidth=0.6)
        ax1.tick_params(labelsize=8, length=3.2, width=0.8)
        for spine in ax1.spines.values():
            spine.set_linewidth(0.8)

        error_color = "#e07a5f"
        ax2.vlines(x[err_mask], ymin=0, ymax=err_rel_pct[err_mask], color=error_color, alpha=0.5, lw=1.0)
        l4 = ax2.scatter(
            x[err_mask],
            err_rel_pct[err_mask],
            s=10,
            marker="o",
            color=error_color,
            alpha=0.85,
            label="Relative Error (%)",
        )
        max_err = float(np.nanmax(err_rel_pct[err_mask])) if np.count_nonzero(err_mask) > 0 else 1.0
        if not np.isfinite(max_err) or max_err <= 0:
            max_err = 1.0
        ax2.set_ylim(0, max_err * 3.0)
        ax2.tick_params(axis="y", colors=error_color, labelsize=8, length=3.2, width=0.8)
        ax2.set_ylabel("Relative Error (%)", color=error_color, fontweight="bold")
        for spine in ax2.spines.values():
            spine.set_linewidth(0.8)

        ca_st = mode_stats_dict.get("CA")
        ot_st = mode_stats_dict.get("OTHERS")
        if ca_st is not None and ot_st is not None:
            stat_txt = (
                f"{MODE_CN.get(mode, mode)} | "
                f"CA(A={ca_st.A:.3g},B={ca_st.B:.3g},R2={ca_st.r2:.3f}) | "
                f"OTH(A={ot_st.A:.3g},B={ot_st.B:.3g},R2={ot_st.r2:.3f})"
            )
        elif ca_st is not None:
            stat_txt = f"{MODE_CN.get(mode, mode)} | CA(A={ca_st.A:.3g},B={ca_st.B:.3g},R2={ca_st.r2:.3f})"
        elif ot_st is not None:
            stat_txt = f"{MODE_CN.get(mode, mode)} | OTH(A={ot_st.A:.3g},B={ot_st.B:.3g},R2={ot_st.r2:.3f})"
        else:
            stat_txt = f"{MODE_CN.get(mode, mode)}"

        ax1.text(
            0.01,
            0.94,
            stat_txt,
            transform=ax1.transAxes,
            ha="left",
            va="top",
            fontsize=8,
            bbox={"facecolor": "white", "alpha": 0.8, "edgecolor": "none", "pad": 1.8},
        )

        if i == 0:
            handles = [l1, l2, l3, l4]
            labels = [h.get_label() for h in handles]
            ax1.legend(handles, labels, loc="upper right", frameon=False, fontsize=8, ncol=2, title="Legend", title_fontsize=8)

    axes[-1].set_xlabel("Experiment", fontweight="bold")
    axes[-1].set_xticks(x)
    axes[-1].set_xticklabels(exp_union, rotation=80, ha="right", fontsize=7)
    fig.suptitle("DimScaling Frequency Prediction and Relative Error", fontsize=11, fontweight="bold", y=0.98)
    fig.subplots_adjust(left=0.08, right=0.92, top=0.90, bottom=0.22, hspace=0.20)
    fig.savefig(out_dir / f"{file_prefix}_dual_combined.png", dpi=save_dpi)
    plt.close(fig)


def write_csv(path: Path, fieldnames: list[str], rows: list[dict]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for r in rows:
            writer.writerow(r)


def run_pipeline(
    h5_path: Path,
    compare_xlsx: Path,
    out_dir: Path,
    outlier: str = "none",
    residual_thresh: float = 0.20,
    combined_figsize: tuple[float, float] = (16, 11),
    save_dpi: int = 320,
    file_prefix: str = "dimscaling",
) -> dict:
    out_dir.mkdir(parents=True, exist_ok=True)
    plt.rcParams.update(
        {
            "font.size": 9,
            "axes.titlesize": 12,
            "axes.labelsize": 10,
            "xtick.labelsize": 8,
            "ytick.labelsize": 8,
            "grid.alpha": 0.2,
        }
    )

    h5_freq_means = load_h5_modal_freq_means(h5_path)
    compare_records, compare_freq_means = load_compare6_records(compare_xlsx)
    rows = build_analysis_rows(compare_records, compare_freq_means, h5_freq_means)

    conflict_rows: list[dict] = []
    for r in rows:
        f_h5 = r["f_exp_h5"]
        f_ex = r["f_exp_excel"]
        if np.isfinite(f_h5) and np.isfinite(f_ex) and f_ex > 0:
            diff = f_h5 - f_ex
            rel = abs(diff) / f_ex
        else:
            diff = np.nan
            rel = np.nan
        conflict_rows.append(
            {
                "exp_id": r["exp_id"],
                "mode": r["mode"],
                "freq_compare6": f_ex,
                "freq_h5": f_h5,
                "diff_h5_minus_compare6": diff,
                "rel_diff": rel,
            }
        )

    all_point_rows: list[dict] = []
    stats_rows: list[dict] = []
    mode_rows: Dict[str, list[dict]] = {}
    mode_stats: Dict[str, Dict[str, FitStats]] = {}
    for mode in MODE_ORDER:
        rows_mode = [r for r in rows if r["mode"] == mode]
        if not rows_mode:
            continue
        mode_rows_dict = {(r["exp_id"], r["mode"]): dict(r) for r in rows_mode}
        mode_stats[mode] = {}
        for fit_group in ["CA", "OTHERS"]:
            rows_group = [r for r in rows_mode if get_fit_group(r["exp_id"]) == fit_group]
            if not rows_group:
                continue
            stats, out_rows = fit_mode_dimscaling(rows_group, outlier, residual_thresh, fit_group=fit_group)
            mode_stats[mode][fit_group] = stats
            stats_rows.append(
                {
                    "mode": stats.mode,
                    "fit_group": stats.fit_group,
                    "n_total": stats.n_total,
                    "n_used": stats.n_used,
                    "outlier_policy": stats.outlier_policy,
                    "A": stats.A,
                    "B": stats.B,
                    "r2": stats.r2,
                    "rmse": stats.rmse,
                    "mae": stats.mae,
                }
            )
            for rr in out_rows:
                mode_rows_dict[(rr["exp_id"], rr["mode"])] = rr

        out_rows_mode = list(mode_rows_dict.values())
        mode_rows[mode] = out_rows_mode
        for r in out_rows_mode:
            all_point_rows.append(
                {
                    "exp_id": r["exp_id"],
                    "mode": r["mode"],
                    "fit_group": get_fit_group(r["exp_id"]),
                    "m": r["m"],
                    "l": r["l"],
                    "r": r["r"],
                    "f_exp_compare6": r["f_exp_excel"],
                    "f_exp_h5": r["f_exp_h5"],
                    "f_exp_used": r["f_exp"],
                    "f_pend": r.get("f_pend", np.nan),
                    "f_fit": r.get("f_fit", np.nan),
                    "fit_used": r.get("fit_used", False),
                    "err_abs_hz": r.get("err_abs", np.nan),
                    "err_rel": r.get("err_rel", np.nan),
                    "A_mode": r.get("A", np.nan),
                    "B_mode": r.get("B", np.nan),
                }
            )

    plot_dual_axis_combined(
        mode_rows,
        mode_stats,
        out_dir,
        figsize=combined_figsize,
        save_dpi=save_dpi,
        file_prefix=file_prefix,
    )

    write_csv(
        out_dir / f"{file_prefix}_fit_summary.csv",
        ["mode", "fit_group", "n_total", "n_used", "outlier_policy", "A", "B", "r2", "rmse", "mae"],
        stats_rows,
    )
    write_csv(
        out_dir / f"{file_prefix}_freq_conflict_report.csv",
        ["exp_id", "mode", "freq_compare6", "freq_h5", "diff_h5_minus_compare6", "rel_diff"],
        conflict_rows,
    )
    write_csv(
        out_dir / f"{file_prefix}_pointwise_results.csv",
        [
            "exp_id",
            "mode",
            "fit_group",
            "m",
            "l",
            "r",
            "f_exp_compare6",
            "f_exp_h5",
            "f_exp_used",
            "f_pend",
            "f_fit",
            "fit_used",
            "err_abs_hz",
            "err_rel",
            "A_mode",
            "B_mode",
        ],
        all_point_rows,
    )

    return {
        "out_dir": str(out_dir),
        "modes_processed": len(stats_rows),
        "rows_total": len(rows),
        "rows_pointwise": len(all_point_rows),
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="DimScaling frequency fitting and plotting.")
    parser.add_argument("--h5-path", type=Path, default=DEFAULT_H5_PATH)
    parser.add_argument("--compare-xlsx", type=Path, default=DEFAULT_METADATA_DIR / "COMPARE_6.xlsx")
    parser.add_argument("--out-dir", type=Path, default=DEFAULT_PROCESSED_DIR)
    parser.add_argument("--outlier", choices=["none", "iqr", "residual"], default="none")
    parser.add_argument("--residual-thresh", type=float, default=0.20, help="Used when --outlier residual.")
    args = parser.parse_args()

    summary = run_pipeline(
        h5_path=args.h5_path,
        compare_xlsx=args.compare_xlsx,
        out_dir=args.out_dir,
        outlier=args.outlier,
        residual_thresh=args.residual_thresh,
    )
    print(f"[DONE] output dir: {summary['out_dir']}")
    print(f"[DONE] modes processed: {summary['modes_processed']}")


if __name__ == "__main__":
    main()
